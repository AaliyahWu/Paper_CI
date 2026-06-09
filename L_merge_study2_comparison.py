"""
L_merge_study2_comparison.py
============================
跑完 I / J / K 三支後，一鍵把結果併成一張 Study Two 總表，並算出
「under-sampling 效果」（各 sampler 相對 none 的指標差），方便寫論文比較。

讀取來源（三支的 ak_best_export / ak_all_export 分頁，皆為原始數值，可直接運算）：
    results/Study2_undersampling_OF_3OCC.xlsx        (I, OF_maj,        對齊 A)
    results/Study2_undersampling_DF_4AE_3OCC.xlsx    (J, DF_maj,        對齊 B)
    results/Study2_undersampling_OFDF_4AE_3OCC.xlsx  (K, OF_maj+DF_maj,對齊 C)

輸出：results/Study2_master_comparison.xlsx
    1. master_best     三支 ak_best_export 直接縱向合併（統一 schema，逐 fold）
    2. master_all      三支 ak_all_export 合併（完整 config，逐 fold）
    3. pivot_AUC       FeatureSet × AE × Sampler（列）× OCC（欄）的平均 AUC
                       —— 一眼看出在哪個特徵集 / OCC 下，哪個 sampler 較好
    4. effect_vs_none  各 sampler 相對 none 的平均指標差（Δ）
                       —— Study Two 的核心結論：under-sampling「有沒有幫助」
    5. none_vs_baseline none 列摘要，供你回頭核對是否重現 A / B / C

設計重點：
    • 完全依賴統一的 ak_export schema（欄位名抓值），不受各檔額外 metadata 欄影響。
    • 對 OCC / AE / FeatureSet 都保留維度，三支可在同一張表中交叉比較。
    • effect_vs_none 以「同 (FeatureSet, AE, OCC) 內」為基準對齊 none，避免跨條件誤比。
"""

import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

RESULTS_DIR = Path("results")
OUTPUT_FILE = RESULTS_DIR / "Study2_master_comparison.xlsx"

SOURCES = [
    ("I", RESULTS_DIR / "Study2_undersampling_OF_3OCC.xlsx"),
    ("J", RESULTS_DIR / "Study2_undersampling_DF_4AE_3OCC.xlsx"),
    ("K", RESULTS_DIR / "Study2_undersampling_OFDF_4AE_3OCC.xlsx"),
]

METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]
SAMPLERS    = ["none", "ENN", "CNN", "TL"]
OCC_TYPES   = ["OCSVM", "LOF", "iForest"]
# 統一 schema（與 I/J/K 的 COMPARISON_EXPORT_COLS 一致）
EXPORT_COLS = [
    "Study", "Method", "FeatureSet", "Dataset", "AE", "Sampler", "OCC",
    "Config", "Fold", "ConfigPolicy", "MajKept", "MajRemoved", "RemovedRate",
    "SamplerStatus", "BaselineRef", "OCCScope", "SamplerScaleMode",
] + METRIC_COLS


def load_export(path, sheet):
    """讀取單一檔案的某個 export 分頁；缺檔/缺分頁則回傳空表並提示。"""
    if not path.exists():
        print(f"  [WARN] 找不到 {path.name}（請先跑出該支結果）")
        return pd.DataFrame(columns=EXPORT_COLS)
    try:
        df = pd.read_excel(path, sheet_name=sheet)
    except Exception as e:
        print(f"  [WARN] {path.name} 讀取分頁 {sheet} 失敗：{e}")
        return pd.DataFrame(columns=EXPORT_COLS)
    # 只保留統一欄位（容忍欄位順序差異），缺的欄補 NA
    for c in EXPORT_COLS:
        if c not in df.columns:
            df[c] = np.nan
    return df[EXPORT_COLS]


def gather(sheet):
    frames = []
    for tag, path in SOURCES:
        d = load_export(path, sheet)
        if not d.empty:
            print(f"  [OK] {tag}: {path.name} / {sheet} → {len(d)} 列")
        frames.append(d)
    if not frames or all(f.empty for f in frames):
        return pd.DataFrame(columns=EXPORT_COLS)
    return pd.concat(frames, ignore_index=True)


def build_pivot_auc(master_best):
    """FeatureSet × AE × Sampler（列）× OCC（欄）平均 AUC。"""
    if master_best.empty:
        return pd.DataFrame()
    g = (master_best.groupby(["FeatureSet", "AE", "Sampler", "OCC"])["AUC"]
                    .mean().reset_index())
    piv = g.pivot_table(index=["FeatureSet", "AE", "Sampler"],
                        columns="OCC", values="AUC")
    # OCC 欄依固定順序排列
    piv = piv.reindex(columns=[c for c in OCC_TYPES if c in piv.columns])
    return piv.reset_index()


def build_effect_vs_none(master_best):
    """各 sampler 相對 none 的平均指標差 Δ（同 FeatureSet×AE×OCC 內對齊）。"""
    if master_best.empty:
        return pd.DataFrame()
    grp_keys = ["FeatureSet", "AE", "OCC", "Sampler"]
    mean_tbl = master_best.groupby(grp_keys)[METRIC_COLS].mean().reset_index()
    none_tbl = (mean_tbl[mean_tbl["Sampler"] == "none"]
                .drop(columns=["Sampler"])
                .rename(columns={m: f"{m}_none" for m in METRIC_COLS}))
    merged = mean_tbl.merge(none_tbl, on=["FeatureSet", "AE", "OCC"], how="left")
    for m in METRIC_COLS:
        merged[f"Δ{m}"] = merged[m] - merged[f"{m}_none"]
    keep = ["FeatureSet", "AE", "OCC", "Sampler"] + METRIC_COLS + [f"Δ{m}" for m in METRIC_COLS]
    out = merged[keep]
    # none 自身的 Δ 恆為 0，保留作對照
    return out.sort_values(["FeatureSet", "AE", "OCC", "Sampler"]).reset_index(drop=True)


def build_none_vs_baseline(master_best):
    """抽出 none 列摘要，供核對是否重現 A/B/C（BaselineRef 已標明對應 baseline）。"""
    if master_best.empty:
        return pd.DataFrame()
    none_rows = master_best[master_best["Sampler"] == "none"]
    if none_rows.empty:
        return pd.DataFrame()
    g = (none_rows.groupby(["FeatureSet", "BaselineRef", "AE", "OCC"])[METRIC_COLS]
                  .mean().reset_index())
    return g


# ─────────────────────────── Excel 輸出 ──────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
NONE_FILL   = PatternFill("solid", fgColor="EEEEEE")
POS_FILL    = PatternFill("solid", fgColor="C6EFCE")   # Δ>0 綠（優於 none）
NEG_FILL    = PatternFill("solid", fgColor="F8CBAD")   # Δ<0 紅（劣於 none）
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left", vertical="center")
BORDER      = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))


def _w(cell, v, font=BODY_FONT, fill=None, align=CENTER, fmt=None):
    cell.value = v
    cell.border = BORDER
    cell.font = font
    cell.alignment = align
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt


def dump_table(ws, title, df, numeric_fmt="0.0000", delta_cols=None):
    """通用：把 DataFrame 寫成一張表（首列標題、表頭、斑馬紋；Δ 欄正負上色）。"""
    delta_cols = delta_cols or []
    if df.empty:
        _w(ws.cell(1, 1), f"{title}（無資料：請確認 I/J/K 是否已跑出結果）", font=BOLD_FONT, align=LEFT)
        return
    cols = list(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    _w(ws.cell(1, 1), title, font=Font(name="Arial", bold=True, size=12, color="1F3864"), align=CENTER)
    for c, h in enumerate(cols, 1):
        _w(ws.cell(2, c), h, font=HEADER_FONT, fill=HEADER_FILL)
    for r, (_, row) in enumerate(df.iterrows(), 3):
        base_fill = ALT_FILL if r % 2 == 0 else None
        if "Sampler" in cols and row.get("Sampler") == "none":
            base_fill = NONE_FILL
        for c, col in enumerate(cols, 1):
            val = row[col]
            fill = base_fill
            if col in delta_cols and isinstance(val, (int, float)) and not pd.isna(val):
                if val > 1e-9:   fill = POS_FILL
                elif val < -1e-9: fill = NEG_FILL
            is_num = isinstance(val, (int, float)) and not pd.isna(val)
            _w(ws.cell(r, c), val,
               font=BODY_FONT,
               fill=fill,
               align=LEFT if col in ["Dataset", "FeatureSet", "Method", "SamplerScaleMode"] else CENTER,
               fmt=numeric_fmt if (is_num and col not in ["Fold", "MajKept", "MajRemoved"]) else None)
    for c, col in enumerate(cols, 1):
        width = 14
        if col in ["FeatureSet", "SamplerScaleMode", "BaselineRef"]: width = 22
        if col == "Dataset": width = 26
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A3"


def main():
    print("=" * 64)
    print("併表：Study Two（I + J + K）→ Study2_master_comparison.xlsx")
    print("=" * 64)
    print("\n讀取 ak_best_export：")
    master_best = gather("ak_best_export")
    print("\n讀取 ak_all_export：")
    master_all  = gather("ak_all_export")

    if master_best.empty and master_all.empty:
        print("\n⚠️  三支結果都讀不到，請先跑 I / J / K。")
        return

    pivot_auc = build_pivot_auc(master_best)
    effect    = build_effect_vs_none(master_best)
    none_base = build_none_vs_baseline(master_best)
    delta_cols = [f"Δ{m}" for m in METRIC_COLS]

    wb = Workbook()
    dump_table(wb.active, "master_best（I+J+K 最佳 config 逐 fold 合併）", master_best)
    wb.active.title = "master_best"
    dump_table(wb.create_sheet("master_all"),
               "master_all（I+J+K 全 config 逐 fold 合併）", master_all)
    dump_table(wb.create_sheet("pivot_AUC"),
               "平均 AUC：FeatureSet × AE × Sampler（列）× OCC（欄）", pivot_auc)
    dump_table(wb.create_sheet("effect_vs_none"),
               "Under-sampling 效果：各 sampler 相對 none 的平均指標差 Δ（綠=優於 none，紅=劣於）",
               effect, delta_cols=delta_cols)
    dump_table(wb.create_sheet("none_vs_baseline"),
               "none 摘要（應重現 BaselineRef 指定的 A/B/C；供核對）", none_base)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 已輸出：{OUTPUT_FILE.resolve()}")
    if not effect.empty:
        print("\n── Under-sampling 效果預覽（ΔAUC，>0 表示優於 none）──")
        prev = effect[effect["Sampler"] != "none"][
            ["FeatureSet", "AE", "OCC", "Sampler", "ΔAUC"]]
        print(prev.round(4).to_string(index=False))


if __name__ == "__main__":
    main()
