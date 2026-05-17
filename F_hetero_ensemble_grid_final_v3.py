"""
F_hetero_ensemble_grid.py
=========================
異質 AE Ensemble（Heterogeneous AE Ensemble）
=============================================

【什麼是「異質」？】

 同質（D/E 已做）：
   同一 AE 架構（如 VAE），三個特徵空間（OF / DF / OF+DF）的 OCC 做 voting。
   問題：OF+DF 包含 DF，特徵高度相關，多樣性不足，ensemble 無法帶來互補效果。

 異質（本檔）：
   四種不同 AE 架構（AE / DAE / SAE / VAE），同一特徵空間（DF_maj），
   各自提取的深度特徵完全獨立，再對四個 OCC 做 voting。
   原理：
     • AE  學習線性壓縮（最小重建誤差）
     • DAE 學習去噪（對 noise 魯棒的表示）
     • SAE 學習稀疏表示（特徵分離性更強）
     • VAE 學習機率分佈（隨機潛在空間）
   → 四種 DF 捕捉同一資料的不同面向，voting 可真正互補彼此的缺陷。

【方法總覽】每個 fold 會產出以下 Method 欄位值：

  ── Baseline（複製 B 結果，供直接對照）──────────────────────────────
  B_AE     : 僅用 AE 的 DF_maj → OCC（與 B baseline 相同邏輯，但 config 可能不同）
  B_DAE    : 僅用 DAE 的 DF_maj → OCC
  B_SAE    : 僅用 SAE 的 DF_maj → OCC
  B_VAE    : 僅用 VAE 的 DF_maj → OCC
  → 共 4 種 baseline

  ── Hetero Vote（全部 4 種 AE 一起投票，4 種 scheme）─────────────────
  Hetero_Vote_eq        : 4 個 OCC normalized score 等權平均（w = 1/4 each），soft
  Hetero_Vote_istd      : 4 個 OCC normalized score inverse-std 加權，soft
  Hetero_Vote_hard_or   : 4 個 OCC binary 預測 OR  合併（≥1 票判 anomaly），hard
  Hetero_Vote_hard_maj  : 4 個 OCC binary 預測 strict majority 合併（≥3 票判 anomaly），hard
  → 共 4 種

  ── Hetero Pair（C(4,2)=6 種 AE pair × 4 種 scheme）───────────────────
  Hetero_Pair_AE_DAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_AE_SAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_AE_VAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_DAE_SAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_DAE_VAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_SAE_VAE_eq / _istd / _hard_or / _hard_and
  → 共 6 × 4 = 24 種

  總計：4（baseline）+ 4（Vote）+ 24（Pair）= 32 種 Method

【Voting 說明】
  Soft（連續分數加權平均）：
    eq   → 所有成員等權；istd → w_i ∝ 1/std(s_i_norm)，再歸一化使 Σw=1
    最終分數 → 用 train-maj 第 90 百分位切閾值（與 baseline 一致）
  Hard（二元票數）：
    各成員先用自己的 train-maj 第 90 百分位切成 0/1（即 baseline 邏輯），
    再用以下規則合併：
      hard_or   (2-AE pair / 4-AE vote 都用) : ≥1 票判 anomaly（recall 高）
      hard_and  (2-AE pair 用)               : ≥2 票判 anomaly（precision 高）
      hard_maj  (4-AE vote 用)               : ≥3 票判 anomaly（strict majority；超過半數）
    AUC 用 vote_count 當連續 ranking 分數
    → 同一 vote 系列（or/and 或 or/maj）的 AUC 必然相同，僅 F1/Recall/G-mean 不同

【Grid Search 空間】（與 B/C/D/E 完全一致）
  n_layers   : [1, 2, 3]
  bottleneck : ["1/4", "1/3", "1/2", "1/1", "2/1", "3/1", "4/1"]
  → 21 種 AE 架構 per (AE_type, OCC, fold)

【Config 選擇說明】
  USE_BEST_CONFIG = True（預設）：
    從 B_baseline_grid.xlsx 的 best_overall 分頁讀取 "Most Freq Config" 欄位，
    作為各 (AE_type, OCC_type) 的固定 config（每個 AE 可能使用不同 config）。
    → 每個 (AE, OCC) 只跑 1 種 config，計算量少，適合快速驗證。

  USE_BEST_CONFIG = False：
    對所有 21 種 config 做完整 grid search，輸出的 best_overall 取各
    (Dataset, Method, AE, OCC) 中平均 AUC 最高的 config，邏輯與 D/E 一致。

【輸出】results/F_hetero_ensemble_grid.xlsx
  分頁（格式與 D/E 一致）：
    all_per_fold  : 所有 (Dataset, Method, AE, OCC, Config, Fold) 原始結果
    all_summary   : mean ± std across folds（per Dataset × Method × AE × OCC × Config）
    all_overall   : 全域平均（per Method × AE × OCC × Config，與 D/E 的 all_overall 粒度一致）
    best_per_fold : 各 (Dataset, Method, AE, OCC) 中 AUC 最高 config 的 fold 資料
    best_summary  : best config 的 mean ± std
    best_overall  : Method × AE × OCC 的全域平均（不含 Config 分組），96 列
                    每列 = 17 datasets × 5 folds = 85 筆的 mean ± std；
                    Config 欄顯示最常被選中的 config（mode），僅供參考
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from itertools import combinations

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── 全域 seed（與 B/C/D/E 完全一致，只在 import 時設定一次）──────────
# 配合下方主迴圈採用 ae-major 順序（外圈 AE_type、內圈 cfg、OCC 全部延後），
# F 在 (ds, fold, AE_type, cfg) 呼叫 train_and_extract 時的 torch RNG 累積狀態
# 會與 B/C/D/E 完全相同 → AE 權重相同 → DF 相同
# → F 的 B_AE/B_DAE/B_SAE/B_VAE 與 B baseline 對應 cell 數值相等。
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────
DATA_ROOT       = Path("data")
RESULTS_DIR     = Path("results")
OUTPUT_FILE     = RESULTS_DIR / "F_hetero_ensemble_grid.xlsx"
B_GRID_XLSX     = Path("results/B baseline grid.xlsx")   # 讀取最佳 config 用
RESULTS_DIR.mkdir(exist_ok=True)

# 若 True → 每個 (AE, OCC) 使用 B grid 的最佳 config；否則 full grid search
USE_BEST_CONFIG = False

# ─────────────────────────── 搜尋空間 ────────────────────────────────
AE_TYPES          = ["AE", "DAE", "SAE", "VAE"]
OCC_TYPES         = ["OCSVM", "LOF", "iForest"]
N_LAYERS_LIST     = [1, 2, 3]
BOTTLENECK_RATIOS = {
    "1/4": 1/4, "1/3": 1/3, "1/2": 1/2, "1/1": 1.0,
    "2/1": 2.0, "3/1": 3.0, "4/1": 4.0,
}
EPS = 1e-12  # 與 D/E 一致（防止 inverse-std 除 0）

# ── AE 訓練超參數（與 B_baseline_grid.py 完全一致）─────────────────
AE_EPOCHS     = 100          # B 用 100，F 原本錯誤寫成 50
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1          # DAE 加入的 Gaussian noise 標準差
SAE_SPARSITY  = 1e-3         # SAE L1 稀疏懲罰係數
VAE_BETA      = 1.0          # VAE KL 權重（B 用 1.0；F 原本錯誤用 0.001）

# Pair combinations: C(4,2) = 6
AE_PAIRS = list(combinations(AE_TYPES, 2))

METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]
ALL_CONFIGS = [f"h{n}-{r}" for n in N_LAYERS_LIST for r in BOTTLENECK_RATIOS]

# 固定 Method 輸出順序，避免每次 df 出現順序不同造成 Excel 欄位不穩定。
BASELINE_METHODS = [f"B_{ae}" for ae in AE_TYPES]
HETERO_VOTE_METHODS = [
    "Hetero_Vote_eq", "Hetero_Vote_istd",
    "Hetero_Vote_hard_or", "Hetero_Vote_hard_maj",
]
HETERO_PAIR_METHODS = []
for ae_i, ae_j in AE_PAIRS:
    HETERO_PAIR_METHODS.extend([
        f"Hetero_Pair_{ae_i}_{ae_j}_eq",
        f"Hetero_Pair_{ae_i}_{ae_j}_istd",
        f"Hetero_Pair_{ae_i}_{ae_j}_hard_or",
        f"Hetero_Pair_{ae_i}_{ae_j}_hard_and",
    ])
METHODS = BASELINE_METHODS + HETERO_VOTE_METHODS + HETERO_PAIR_METHODS
AE_GROUP_ORDER = AE_TYPES + ["+".join(AE_TYPES)] + [f"{a}+{b}" for a, b in AE_PAIRS]

# ─────────────── AE 模型定義（與 B_baseline_grid.py 完全一致）───────────────
# 架構設計：dims = [input_dim] + [n_units] * n_layers（平坦隱層，不做逐層縮小）
# F 原本用 get_hidden_dims() 做錐形結構，與 B 不符，現已改回平坦結構。

class AEModel(nn.Module):
    """AE / DAE / SAE 共用架構（與 B/C/D/E 完全一致）。"""
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        for i in range(len(dims) - 1):
            d_in, d_out = dims[-(i+1)], dims[-(i+2)]
            act = nn.Sigmoid() if i == len(dims) - 2 else nn.ReLU()
            dec += [nn.Linear(d_in, d_out), act]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


class VAEModel(nn.Module):
    """VAE 架構（同 B_baseline_grid.py）"""
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        base = []
        for i in range(len(dims) - 2):
            base += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        self.enc_base  = nn.Sequential(*base) if base else nn.Identity()
        mid = dims[-2] if len(dims) >= 2 else input_dim
        self.fc_mu     = nn.Linear(mid, dims[-1])
        self.fc_logvar = nn.Linear(mid, dims[-1])
        dec_dims = dims[::-1]
        dec = []
        for i in range(len(dec_dims) - 1):
            act = nn.Sigmoid() if i == len(dec_dims) - 2 else nn.ReLU()
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), act]
        self.decoder = nn.Sequential(*dec)

    def reparameterize(self, mu, lv):
        return mu + torch.exp(0.5 * lv) * torch.randn_like(lv)

    def encode(self, x):
        h = self.enc_base(x)
        return self.fc_mu(h), self.fc_logvar(h)

    def forward(self, x):
        h  = self.enc_base(x)
        mu = self.fc_mu(h);  lv = self.fc_logvar(h)
        z  = self.reparameterize(mu, lv)
        return self.decoder(z), z, mu, lv

def train_and_extract(ae_type, X_maj_s, X_test_s, n_layers, n_units):
    """
    訓練指定 AE 並提取深度特徵（DF）。與 B_baseline_grid.py 完全對齊。

    參數：
      ae_type  : 'AE' | 'DAE' | 'SAE' | 'VAE'
      X_maj_s  : MinMaxScaled train-majority（shape: [n_maj, input_dim]）
      X_test_s : MinMaxScaled test（shape: [n_test, input_dim]）
      n_layers : 隱層數（1/2/3）
      n_units  : 每層寬度 = max(2, round(input_dim × ratio))（與 B 一致）

    各 AE 訓練細節（與 B 對齊）：
      AE  : MSE loss，encoder 輸出 z（ReLU）
      DAE : 加 noise（std=DAE_NOISE）→ encode → reconstruct 乾淨 x；提取用乾淨 x
      SAE : MSE + SAE_SPARSITY × z.abs().mean()（L1 稀疏，F 原本錯用 KL divergence）
      VAE : MSE + VAE_BETA × KL（VAE_BETA=1.0，F 原本錯用 0.001）；提取取 mu

    Random state：完全沿用全域 torch.manual_seed(42)，DataLoader 不指定 generator。
                  與 B/C/D/E 的 train_and_extract 使用方式相同。

    返回：(DF_train, DF_test)，shape = (n, n_units)
    """
    input_dim = X_maj_s.shape[1]
    model = VAEModel(input_dim, n_layers, n_units) if ae_type == "VAE" \
            else AEModel(input_dim, n_layers, n_units)

    optim  = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))

    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True,
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim.zero_grad()
            if ae_type == "DAE":
                xb_n = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
                xr, _ = model(xb_n)
                loss  = mse(xr, xb)
            elif ae_type == "SAE":
                xr, z = model(xb)
                loss  = mse(xr, xb) + SAE_SPARSITY * z.abs().mean()
            elif ae_type == "VAE":
                xr, _, mu, lv = model(xb)
                kl   = -0.5 * (1 + lv - mu.pow(2) - lv.exp()).mean()
                loss = mse(xr, xb) + VAE_BETA * kl
            else:  # AE
                xr, _ = model(xb)
                loss  = mse(xr, xb)
            loss.backward()
            optim.step()

    model.eval()

    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            if ae_type == "VAE":
                # 與 B/C/D/E 一致：跑完整 forward 取 mu，不用 model.encode()。
                # 這樣 reparameterize 內部的 randn_like 會被 torch RNG 消耗，
                # 與 B/C/D/E 同步推進，下一輪 VAE 訓練起點才會與它們對齊。
                _, _, mu, _ = model(xt)
                return mu.numpy()
            else:
                _, z = model(xt)
                return z.numpy()

    return extract(X_maj_s), extract(X_test_s)


# ─────────────────────────── OCC 訓練 ────────────────────────────────
def train_occ_scores(occ_type, X_maj, X_tst, n_nb_cap):
    n_nb = min(20, max(2, n_nb_cap))
    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(X_maj)
        sm = -clf.decision_function(X_maj)
        st = -clf.decision_function(X_tst)
    elif occ_type == "LOF":
        clf = LocalOutlierFactor(n_neighbors=n_nb, contamination=0.1, novelty=True)
        clf.fit(X_maj)
        sm = -clf.decision_function(X_maj)
        st = -clf.decision_function(X_tst)
    else:
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(X_maj)
        sm = -clf.decision_function(X_maj)
        st = -clf.decision_function(X_tst)
    return sm, st


def normalize_by_majority(sm, st):
    lo, hi = sm.min(), sm.max()
    if hi - lo < EPS:
        return np.zeros_like(sm), np.zeros_like(st)
    return (sm - lo) / (hi - lo), (st - lo) / (hi - lo)


def metrics_from(y_true, y_pred, s_cont):
    try:   auc = roc_auc_score(y_true, s_cont)
    except: auc = np.nan
    if y_pred.sum() == 0:
        return {"AUC": auc, "F1": 0.0, "Recall": 0.0, "G-mean": 0.0}
    f1  = f1_score(y_true, y_pred, pos_label=1, zero_division=0)
    rec = recall_score(y_true, y_pred, pos_label=1, zero_division=0)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    gm   = np.sqrt(rec * spec)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


def weighted_soft_vote(scores_norm_maj, scores_norm_tst, scheme, y_tst):
    """
    對多個已 normalize 的 OCC score 做加權 soft voting。

    參數：
      scores_norm_maj : list[np.ndarray]，各 OCC 在 train-maj 上的 normalized score
      scores_norm_tst : list[np.ndarray]，各 OCC 在 test 上的 normalized score
                        normalize 用 train-maj 的 min/max，不碰 test 統計，無 data leak
      scheme          : 'eq' | 'istd'

    加權方式：
      eq   → 所有 OCC 等權重：w_i = 1/n（n = OCC 數量）
      istd → inverse-std 加權：w_i ∝ 1 / std(scores_norm_maj[i])
             std 越小表示該 OCC 在訓練集上的分數越穩定，給予更高的信任度。
             公式：w_i = (1/std_i) / Σ(1/std_j)，確保 Σw_i = 1
             EPS 保護分母，避免 std=0 時除以零。

    閾值：
      取加權合併後的 train-maj score 的第 90 百分位數為 threshold
      （與 A/B/C/D/E 一致，確保所有方法可直接比較）

    返回：
      metrics dict，含 AUC / F1 / Recall / G-mean
    """
    n = len(scores_norm_maj)
    if scheme == "eq":
        weights = np.ones(n) / n
    else:  # istd
        stds = np.array([s.std() + EPS for s in scores_norm_maj])
        inv_stds = 1.0 / stds
        weights = inv_stds / inv_stds.sum()   # 歸一化，使 Σw_i = 1

    s_maj_w = sum(w * s for w, s in zip(weights, scores_norm_maj))
    s_tst_w = sum(w * s for w, s in zip(weights, scores_norm_tst))
    thr      = np.percentile(s_maj_w, 90)
    y_pred   = (s_tst_w >= thr).astype(int)
    return metrics_from(y_tst, y_pred, s_tst_w)


def hard_vote(y_pred_list, y_tst, rule):
    """
    對多個 OCC 的 binary 預測做 hard voting。

    參數：
      y_pred_list : list[np.ndarray]，各 OCC 在 test 上的 0/1 預測
                    （由各 OCC 用自己 train-maj 的第 90 百分位切閾值產生，
                    與 baseline B 邏輯一致）
      y_tst       : 真實標籤
      rule        : 二元預測的合併規則：
                    "or"   → ≥1 票判 anomaly（任一 OCC 偵測到就視為異常）
                    "and"  → 全票判 anomaly（所有 OCC 都同意才視為異常）
                    "maj"  → strict majority：票數必須超過半數才判 anomaly
                             例如 n=4 時 ≥3 票；n=3 時 ≥2 票，與 D 的 Vote3_hard 一致
                    "and" 對於 n=2 等於 "maj"；n>2 時更嚴格

    AUC 計算：
      用連續 vote_count 作為 ranking 分數（範圍 0~n）
      → 同一組成員的 _hard_or / _hard_and / _hard_maj 三者 AUC 必然相同
      → F1 / Recall / G-mean 因二元預測門檻不同而有差異

    回傳：
      metrics dict（AUC / F1 / Recall / G-mean）
    """
    n = len(y_pred_list)
    vote_count = np.sum([y.astype(int) for y in y_pred_list], axis=0)  # 0~n

    if rule == "or":
        thr_votes = 1
    elif rule == "and":
        thr_votes = n
    elif rule == "maj":
        thr_votes = int(np.floor(n / 2) + 1)   # strict majority: n=4 → 3; n=3 → 2; n=2 → 2
        # 2-pair 不使用 maj；pair 維持 hard_or / hard_and 兩種規則
    else:
        raise ValueError(f"未知的 rule: {rule}")

    y_pred = (vote_count >= thr_votes).astype(int)
    return metrics_from(y_tst, y_pred, vote_count.astype(float))


# ─────────────────────────── 資料讀取 ────────────────────────────────
def load_dat(path, minority_label=None):
    """
    解析 KEEL .dat。

    修正重點：
      - 類別特徵與 B/C/D/E 一樣使用 pd.Categorical(...).codes，不再把非數值全部補 0。
      - test fold 沿用 train fold 的 minority_label，避免 train/test 各自判斷少數類。

    回傳：X, y, minority_label；其中 y=1 表示 minority/anomaly，y=0 表示 majority/normal。
    """
    lines = Path(path).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"):
            continue
        if s.lower() == "@data":
            data_start = True
            continue
        if data_start:
            rows.append(s)
    if not rows:
        raise ValueError(f"No data found in {path}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    y_raw = df[label_col].astype(str).str.strip().values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    if minority_label is None:
        uniq, cnts = np.unique(y_raw, return_counts=True)
        minority_label = uniq[np.argmin(cnts)]
    y = (y_raw == minority_label).astype(int)
    return X, y, minority_label


def scan_folds(root):
    out = {}
    for ds in sorted(root.iterdir()):
        if not ds.is_dir(): continue
        files = list(ds.glob("*.dat"))
        tra = sorted([f for f in files if "tra.dat" in f.name])
        tst = sorted([f for f in files if "tst.dat" in f.name])
        pairs = list(zip(tra, tst))
        if pairs: out[ds.name] = pairs
    return out


# ─────────────────────────── Best config 讀取 ────────────────────────
def load_best_configs():
    """從 B_baseline_grid.xlsx 讀取各 (AE, OCC) 最佳 config"""
    try:
        df = pd.read_excel(B_GRID_XLSX, sheet_name="best_overall", header=1)
        best = {}
        for _, r in df.iterrows():
            ae, occ, cfg = r["AE"], r["OCC"], r["Most Freq Config"]
            best[(ae, occ)] = cfg
        return best
    except Exception as e:
        print(f"[警告] 無法讀取 {B_GRID_XLSX}：{e}，改用 h2-1/2 作為預設")
        return {}


def parse_config(cfg):
    """
    將 config 字串解析成 (n_layers, ratio_label)。
    例：'h2-1/2' → (2, '1/2')，'h3-4/1' → (3, '4/1')

    若格式不符（cfg 不是 'hN-ratio' 格式），回傳預設值 (2, '1/2')。
    預設值選 h2-1/2 的原因：B baseline grid 中此 config 在多數 (AE, OCC)
    組合下出現頻率最高，是最具代表性的「中性」config。
    """
    m = re.match(r"h(\d+)-(.+)", cfg)
    if not m:
        return 2, "1/2"
    return int(m.group(1)), m.group(2)


# ─────────────────────────── 主程式 ──────────────────────────────────
def run_experiment():
    """
    主實驗迴圈。

    修正重點：
      1. full grid search 時，順序改為 dataset → fold → config → AE feature → OCC，
         因此同一個 fold/config/AE 的 DF 只訓練一次，OCSVM/LOF/iForest 共用同一份 DF。
      2. USE_BEST_CONFIG=True 時，由於最佳 config 是 (AE, OCC) 專屬，會以
         (fold, AE, config) 快取，避免相同 AE/config 被不同 OCC 重複訓練。
      3. train/test label encoding 沿用 train fold 的 minority_label。
    """
    dataset_folds = scan_folds(DATA_ROOT)
    if not dataset_folds:
        print(f"[錯誤] 找不到資料集，請確認 {DATA_ROOT}/ 目錄")
        return pd.DataFrame()

    best_configs = load_best_configs() if USE_BEST_CONFIG else {}
    all_records = []

    def append_results(ds_name, fold_idx, occ_type, y_tst, n_nb_cap, ae_features, report_cfg):
        """將一組已訓練好的 AE features 丟給指定 OCC，並加入 baseline/vote/pair 結果。"""
        if len(ae_features) < 2:
            print(f"  [SKIP] {ds_name} Fold{fold_idx} {occ_type}: 只有 {len(ae_features)} 個 AE 成功")
            return

        raw_scores = {}
        for ae_type, (DF_maj_s, DF_tst_s, _) in ae_features.items():
            sm, st = train_occ_scores(occ_type, DF_maj_s, DF_tst_s, n_nb_cap)
            raw_scores[ae_type] = (sm, st)

        norm_scores = {
            ae: normalize_by_majority(sm, st)
            for ae, (sm, st) in raw_scores.items()
        }

        baseline_ypred = {}
        for ae_type, (sm, st) in raw_scores.items():
            thr = np.percentile(sm, 90)
            y_pred = (st >= thr).astype(int)
            m = metrics_from(y_tst, y_pred, st)
            baseline_ypred[ae_type] = y_pred
            all_records.append({
                "Dataset": ds_name,
                "AE": ae_type,
                "OCC": occ_type,
                "Config": ae_features[ae_type][2],
                "Fold": fold_idx,
                "Method": f"B_{ae_type}",
                **m,
            })

        avail_aes = [ae for ae in AE_TYPES if ae in norm_scores]
        if len(avail_aes) >= 2:
            for scheme in ("eq", "istd"):
                sm_list = [norm_scores[ae][0] for ae in avail_aes]
                st_list = [norm_scores[ae][1] for ae in avail_aes]
                m = weighted_soft_vote(sm_list, st_list, scheme, y_tst)
                all_records.append({
                    "Dataset": ds_name,
                    "AE": "+".join(avail_aes),
                    "OCC": occ_type,
                    "Config": report_cfg,
                    "Fold": fold_idx,
                    "Method": f"Hetero_Vote_{scheme}",
                    **m,
                })

            y_pred_list = [baseline_ypred[ae] for ae in avail_aes]
            for rule in ("or", "maj"):
                m = hard_vote(y_pred_list, y_tst, rule)
                all_records.append({
                    "Dataset": ds_name,
                    "AE": "+".join(avail_aes),
                    "OCC": occ_type,
                    "Config": report_cfg,
                    "Fold": fold_idx,
                    "Method": f"Hetero_Vote_hard_{rule}",
                    **m,
                })

        for ae_i, ae_j in AE_PAIRS:
            if ae_i not in norm_scores or ae_j not in norm_scores:
                continue
            for scheme in ("eq", "istd"):
                sm_list = [norm_scores[ae_i][0], norm_scores[ae_j][0]]
                st_list = [norm_scores[ae_i][1], norm_scores[ae_j][1]]
                m = weighted_soft_vote(sm_list, st_list, scheme, y_tst)
                all_records.append({
                    "Dataset": ds_name,
                    "AE": f"{ae_i}+{ae_j}",
                    "OCC": occ_type,
                    "Config": report_cfg,
                    "Fold": fold_idx,
                    "Method": f"Hetero_Pair_{ae_i}_{ae_j}_{scheme}",
                    **m,
                })

            y_pred_pair = [baseline_ypred[ae_i], baseline_ypred[ae_j]]
            for rule in ("or", "and"):
                m = hard_vote(y_pred_pair, y_tst, rule)
                all_records.append({
                    "Dataset": ds_name,
                    "AE": f"{ae_i}+{ae_j}",
                    "OCC": occ_type,
                    "Config": report_cfg,
                    "Fold": fold_idx,
                    "Method": f"Hetero_Pair_{ae_i}_{ae_j}_hard_{rule}",
                    **m,
                })

    all_cfgs = [f"h{n}-{r}" for n in N_LAYERS_LIST for r in BOTTLENECK_RATIOS]

    for ds_name, fold_pairs in dataset_folds.items():
        try:
            X0, _, _ = load_dat(str(fold_pairs[0][0]))
            input_dim = X0.shape[1]
        except Exception as e:
            print(f"[跳過] {ds_name}: {e}")
            continue

        print(f"\n[{ds_name}] input_dim={input_dim}, folds={len(fold_pairs)}")

        for fold_idx, (tra_path, tst_path) in enumerate(fold_pairs, 1):
            try:
                X_tra, y_tra, minority_label = load_dat(str(tra_path))
                X_tst, y_tst, _ = load_dat(str(tst_path), minority_label=minority_label)
            except Exception as e:
                print(f"  [ERROR] Fold{fold_idx} 讀檔失敗: {e}")
                continue

            if y_tst.sum() < 1:
                print(f"  [SKIP] Fold{fold_idx}: 測試集無少數類樣本")
                continue

            X_maj = X_tra[y_tra == 0]
            if len(X_maj) < 5:
                print(f"  [SKIP] Fold{fold_idx}: 訓練集正常樣本不足 ({len(X_maj)})")
                continue

            sc_orig = MinMaxScaler()
            X_maj_s = sc_orig.fit_transform(X_maj)
            X_tst_s = sc_orig.transform(X_tst)
            n_nb_cap = max(1, len(X_maj_s) - 1)

            if not USE_BEST_CONFIG:
                # ── Step 1: AE-major 訓練（外圈 AE_type、內圈 cfg）─────────
                # 跟 B/C/D/E 主迴圈順序完全一致；OCC 全部延後到 Step 2 才跑。
                # 這樣 (ds, fold, AE_type, cfg) 呼叫 train_and_extract 時，
                # torch RNG 累積消耗的狀態與 B/C/D/E 完全相同，產出的 DF 也相同。
                all_features = {}  # (ae_type, cfg_label) → (DF_maj_s, DF_tst_s, cfg_label)
                for ae_type in AE_TYPES:
                    for cfg in all_cfgs:
                        n_layers, ratio_label = parse_config(cfg)
                        n_units = max(2, round(input_dim * BOTTLENECK_RATIOS.get(ratio_label, 0.5)))
                        cfg_label = f"h{n_layers}-{ratio_label}"
                        try:
                            DF_maj, DF_tst = train_and_extract(
                                ae_type, X_maj_s, X_tst_s, n_layers, n_units
                            )
                            sc_df = MinMaxScaler()
                            all_features[(ae_type, cfg_label)] = (
                                sc_df.fit_transform(DF_maj),
                                sc_df.transform(DF_tst),
                                cfg_label,
                            )
                        except Exception as e:
                            print(f"  [WARN] {ae_type} {cfg_label} Fold{fold_idx}: {e}")

                # ── Step 2: AE 全部訓練完之後再跑 OCC + ensemble ───────────
                for cfg in all_cfgs:
                    n_layers, ratio_label = parse_config(cfg)
                    cfg_label = f"h{n_layers}-{ratio_label}"
                    ae_features = {
                        ae_type: all_features[(ae_type, cfg_label)]
                        for ae_type in AE_TYPES
                        if (ae_type, cfg_label) in all_features
                    }
                    for occ_type in OCC_TYPES:
                        append_results(ds_name, fold_idx, occ_type, y_tst, n_nb_cap, ae_features, cfg_label)

            else:
                # ── Step 1: 收集所有 OCC 共需要的 (AE, cfg) 組合 ──────────
                cfg_by_ae_per_occ = {
                    occ_type: {
                        ae_type: best_configs.get((ae_type, occ_type), "h2-1/2")
                        for ae_type in AE_TYPES
                    }
                    for occ_type in OCC_TYPES
                }
                unique_pairs = set()
                for cfg_by_ae in cfg_by_ae_per_occ.values():
                    for ae_type, cfg in cfg_by_ae.items():
                        unique_pairs.add((ae_type, cfg))

                # ── Step 2: AE-major 訓練（同 USE_BEST_CONFIG=False 的順序）─
                feature_cache = {}
                for ae_type in AE_TYPES:
                    cfgs_for_this_ae = sorted({cfg for (a, cfg) in unique_pairs if a == ae_type})
                    for cfg in cfgs_for_this_ae:
                        n_layers, ratio_label = parse_config(cfg)
                        n_units = max(2, round(input_dim * BOTTLENECK_RATIOS.get(ratio_label, 0.5)))
                        cfg_label = f"h{n_layers}-{ratio_label}"
                        try:
                            DF_maj, DF_tst = train_and_extract(
                                ae_type, X_maj_s, X_tst_s, n_layers, n_units
                            )
                            sc_df = MinMaxScaler()
                            feature_cache[(ae_type, cfg)] = (
                                sc_df.fit_transform(DF_maj),
                                sc_df.transform(DF_tst),
                                cfg_label,
                            )
                        except Exception as e:
                            print(f"  [WARN] {ae_type} {cfg_label} Fold{fold_idx}: {e}")

                # ── Step 3: 跑 OCC + ensemble ──────────────────────────────
                for occ_type in OCC_TYPES:
                    cfg_by_ae = cfg_by_ae_per_occ[occ_type]
                    report_cfg = ";".join(f"{ae}:{cfg}" for ae, cfg in cfg_by_ae.items())
                    ae_features = {
                        ae_type: feature_cache[(ae_type, cfg)]
                        for ae_type, cfg in cfg_by_ae.items()
                        if (ae_type, cfg) in feature_cache
                    }
                    append_results(ds_name, fold_idx, occ_type, y_tst, n_nb_cap, ae_features, report_cfg)

    return pd.DataFrame(all_records)


# ─────────────────────────── Excel 輸出 ──────────────────────────────
def _metric_display(mean_val, std_val):
    if pd.isna(mean_val):
        return "N/A"
    if pd.isna(std_val):
        return f"{mean_val:.4f} ± N/A"
    return f"{mean_val:.4f} ± {std_val:.4f}"


def _ordered(values, preferred_order):
    values = list(pd.Series(values).dropna().unique())
    preferred = [v for v in preferred_order if v in values]
    extra = sorted([v for v in values if v not in preferred])
    return preferred + extra


def _method_ae_pairs(df):
    """
    F 方法中 Method 與 AE 群組不是完整交叉乘積：
      - B_AE 只對應 AE
      - Hetero_Vote_* 只對應 AE+DAE+SAE+VAE
      - Hetero_Pair_AE_DAE_* 只對應 AE+DAE
    因此 wide pivot 只展開實際存在的 (Method, AE) 組合，避免產生超過 Excel
    欄位上限的大量空欄，也避免後續論文整合時誤把不存在的組合納入比較。
    """
    pairs = df[["Method", "AE"]].drop_duplicates()
    method_order = _ordered(pairs["Method"], METHODS)
    ae_order = _ordered(pairs["AE"], AE_GROUP_ORDER)
    return [(m, ae) for m in method_order for ae in ae_order
            if ((pairs["Method"] == m) & (pairs["AE"] == ae)).any()]


def compute_summary_long(df):
    result = (df.groupby(["Dataset", "Method", "AE", "OCC", "Config"])[METRIC_COLS]
                .agg(["mean", "std"]).round(4))
    result.columns = [f"{m}_{s}" for m, s in result.columns]
    return result.reset_index()


def compute_summary_wide(df, include_config=True):
    """
    產生 wide pivot，取代舊版 F 的 long all_summary / best_summary。

    include_config=True  : Dataset × (Method|AE|OCC|Config|Metric)
    include_config=False : Dataset × (Method|AE|OCC|Metric)

    儲存格格式統一為「mean ± std」，與 B/C/D/E 的 summary 類型一致；
    欄名用 | 扁平化，方便後續 pandas read_excel 後直接 concat 或篩選。
    """
    keys = ["Dataset", "Method", "AE", "OCC"] + (["Config"] if include_config else [])
    grouped = df.groupby(keys)[METRIC_COLS].agg(["mean", "std"]).round(4)

    datasets = _ordered(df["Dataset"], sorted(df["Dataset"].dropna().unique()))
    method_ae_pairs = _method_ae_pairs(df)
    occs = _ordered(df["OCC"], OCC_TYPES)
    configs = _ordered(df["Config"], ALL_CONFIGS) if include_config else [None]

    rows = []
    for ds in datasets:
        row = {"Dataset": ds}
        for method, ae in method_ae_pairs:
            for occ in occs:
                if include_config:
                    for cfg in configs:
                        for metric in METRIC_COLS:
                            col = f"{method}|{ae}|{occ}|{cfg}|{metric}"
                            try:
                                mean_val = grouped.loc[(ds, method, ae, occ, cfg), (metric, "mean")]
                                std_val = grouped.loc[(ds, method, ae, occ, cfg), (metric, "std")]
                                row[col] = _metric_display(mean_val, std_val)
                            except KeyError:
                                row[col] = "N/A"
                else:
                    for metric in METRIC_COLS:
                        col = f"{method}|{ae}|{occ}|{metric}"
                        try:
                            mean_val = grouped.loc[(ds, method, ae, occ), (metric, "mean")]
                            std_val = grouped.loc[(ds, method, ae, occ), (metric, "std")]
                            row[col] = _metric_display(mean_val, std_val)
                        except KeyError:
                            row[col] = "N/A"
        rows.append(row)
    return pd.DataFrame(rows)


def compute_overall_by_config(df):
    """
    all_overall：以 Method × AE × OCC × Config 為 key 計算全域平均。

    這個粒度與 D/E 的 all_overall 對齊，避免 full grid 時把 21 種 config
    先混在一起平均，造成後續跨方法 concat 或畫圖時粒度不一致。
    """
    result = (df.groupby(["Method", "AE", "OCC", "Config"])[METRIC_COLS]
                .agg(["mean", "std"]).round(4))
    result.columns = [f"{m}_{s}" for m, s in result.columns]
    result = result.reset_index()

    method_rank = {m: i for i, m in enumerate(METHODS)}
    ae_rank = {ae: i for i, ae in enumerate(AE_GROUP_ORDER)}
    occ_rank = {occ: i for i, occ in enumerate(OCC_TYPES)}
    cfg_rank = {cfg: i for i, cfg in enumerate(ALL_CONFIGS)}
    result["_method_rank"] = result["Method"].map(method_rank).fillna(9999)
    result["_ae_rank"] = result["AE"].map(ae_rank).fillna(9999)
    result["_occ_rank"] = result["OCC"].map(occ_rank).fillna(9999)
    result["_cfg_rank"] = result["Config"].map(cfg_rank).fillna(9999)
    result = result.sort_values([
        "_method_rank", "_ae_rank", "_occ_rank", "_cfg_rank",
        "Method", "AE", "OCC", "Config"
    ])

    metric_cols_ordered = [f"{m}_{s}" for m in METRIC_COLS for s in ("mean", "std")]
    return result[["Method", "AE", "OCC", "Config"] + metric_cols_ordered]


def compute_overall(df):
    """
    overall / best_overall：以 Method × AE × OCC 為唯一 key 計算全域平均。

    Config 不做分組，避免 full grid 時同一方法被 config 拆成大量列；
    Most Freq Config 僅顯示該組最常出現的 config，作為論文表格註記。
    """
    result = (df.groupby(["Method", "AE", "OCC"])[METRIC_COLS]
                .agg(["mean", "std"]).round(4))
    result.columns = [f"{m}_{s}" for m, s in result.columns]
    result = result.reset_index()

    cfg_mode = (df.groupby(["Method", "AE", "OCC"])["Config"]
                  .agg(lambda x: x.mode().iloc[0] if not x.mode().empty else "N/A")
                  .reset_index())
    result = result.merge(cfg_mode, on=["Method", "AE", "OCC"])
    result = result.rename(columns={"Config": "Most Freq Config"})

    # 固定排序，讓 A~F 論文整合時不會因為資料出現順序而改變列順序。
    method_rank = {m: i for i, m in enumerate(METHODS)}
    ae_rank = {ae: i for i, ae in enumerate(AE_GROUP_ORDER)}
    occ_rank = {occ: i for i, occ in enumerate(OCC_TYPES)}
    result["_method_rank"] = result["Method"].map(method_rank).fillna(9999)
    result["_ae_rank"] = result["AE"].map(ae_rank).fillna(9999)
    result["_occ_rank"] = result["OCC"].map(occ_rank).fillna(9999)
    result = result.sort_values(["_method_rank", "_ae_rank", "_occ_rank", "Method", "AE", "OCC"])

    metric_cols_ordered = [f"{m}_{s}" for m in METRIC_COLS for s in ("mean", "std")]
    return result[["Method", "AE", "OCC", "Most Freq Config"] + metric_cols_ordered]


def format_overall_by_config(df_ov):
    rows = []
    for _, vals in df_ov.iterrows():
        row = {
            "Method": vals["Method"],
            "AE": vals["AE"],
            "OCC": vals["OCC"],
            "Config": vals["Config"],
        }
        for metric in METRIC_COLS:
            row[metric] = _metric_display(vals[f"{metric}_mean"], vals[f"{metric}_std"])
        rows.append(row)
    return pd.DataFrame(rows)


def format_overall(df_ov):
    rows = []
    for _, vals in df_ov.iterrows():
        row = {
            "Method": vals["Method"],
            "AE": vals["AE"],
            "OCC": vals["OCC"],
            "Most Freq Config": vals["Most Freq Config"],
        }
        for metric in METRIC_COLS:
            row[metric] = _metric_display(vals[f"{metric}_mean"], vals[f"{metric}_std"])
        rows.append(row)
    return pd.DataFrame(rows)


def write_df(ws, df, title=""):
    ws.title = title if title else ws.title
    df_reset = df.reset_index(drop=True)
    headers = list(df_reset.columns)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(1, j, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(df_reset.itertuples(index=False), 2):
        for j, v in enumerate(row, 1):
            ws.cell(i, j, round(v, 4) if isinstance(v, float) else v)
    ws.freeze_panes = "A2"
    # 不自動調整所有超寬欄，避免大型 wide sheet 寫檔過慢；保留 Dataset / 前幾欄可讀性即可。
    if headers:
        ws.column_dimensions["A"].width = 28
        for j in range(2, min(len(headers), 20) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 18


def save_excel(df_all):
    """
    將結果寫入 Excel，修正舊版 F 與 B/C/D/E 格式不一致的問題。

    分頁設計：
      all_per_fold  : 原始長表，保留完整 fold/config 紀錄
      all_summary   : wide pivot，Dataset 為列，Method|AE|OCC|Config|Metric 為欄
      all_overall   : stackable long overall，保留 Config 粒度，與 D/E 的 all_overall 一致
      best_per_fold : 每 Dataset×Method×AE×OCC 選 5-fold 平均 AUC 最高的 Config
      best_summary  : wide pivot，Dataset 為列，Method|AE|OCC|Metric 為欄
      best_overall  : stackable long best overall，可直接與 B/C/D/E best_overall 整合
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("all_per_fold")
    write_df(ws1, df_all)

    ws2 = wb.create_sheet("all_summary")
    write_df(ws2, compute_summary_wide(df_all, include_config=True))

    ws3 = wb.create_sheet("all_overall")
    write_df(ws3, format_overall_by_config(compute_overall_by_config(df_all)))

    # per (Dataset, Method, AE, OCC)，取 5-fold 平均 AUC 最高的 Config，
    # 保留該 Config 的所有 fold。這與 B/C/D/E 的 best 選法一致，避免 per-fold leakage。
    best_cfg = (
        df_all.dropna(subset=["AUC"])
              .groupby(["Dataset", "Method", "AE", "OCC", "Config"])["AUC"]
              .mean()
              .reset_index()
              .sort_values("AUC", ascending=False)
              .drop_duplicates(["Dataset", "Method", "AE", "OCC"])
    )
    df_best_filtered = df_all.merge(
        best_cfg[["Dataset", "Method", "AE", "OCC", "Config"]],
        on=["Dataset", "Method", "AE", "OCC", "Config"],
    )

    ws4 = wb.create_sheet("best_per_fold")
    write_df(ws4, df_best_filtered)

    ws5 = wb.create_sheet("best_summary")
    write_df(ws5, compute_summary_wide(df_best_filtered, include_config=False))

    ws6 = wb.create_sheet("best_overall")
    write_df(ws6, format_overall(compute_overall(df_best_filtered)))

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果儲存至：{OUTPUT_FILE}")


# ─────────────────────────── Entry Point ─────────────────────────────
if __name__ == "__main__":
    # Method 數量計算
    n_vote_soft = 2                       # Hetero_Vote_eq + Hetero_Vote_istd
    n_vote_hard = 2                       # Hetero_Vote_hard_or + Hetero_Vote_hard_maj
    n_pair_soft = len(AE_PAIRS) * 2       # 6 pairs × 2 = 12
    n_pair_hard = len(AE_PAIRS) * 2       # 6 pairs × 2 (or, and) = 12
    n_base      = len(AE_TYPES)           # B_AE + B_DAE + B_SAE + B_VAE = 4
    n_total     = n_vote_soft + n_vote_hard + n_pair_soft + n_pair_hard + n_base

    print("=" * 72)
    print("F_hetero_ensemble_grid.py — 異質 AE Ensemble")
    print(f"AE 類型    ：{AE_TYPES}")
    print(f"OCC 方法   ：{OCC_TYPES}")
    print(f"Best config from B grid：{USE_BEST_CONFIG}")
    print(f"Method 數量：Baseline({n_base}) + Vote({n_vote_soft+n_vote_hard}) "
          f"+ Pair({n_pair_soft+n_pair_hard}) = {n_total} 種")
    print(f"  Baseline    : {[f'B_{ae}' for ae in AE_TYPES]}")
    print(f"  Vote (soft) : [Hetero_Vote_eq, Hetero_Vote_istd]")
    print(f"  Vote (hard) : [Hetero_Vote_hard_or (≥1), Hetero_Vote_hard_maj (≥3)]")
    print(f"  Pair (soft) : {len(AE_PAIRS)} pairs × [eq, istd]")
    print(f"  Pair (hard) : {len(AE_PAIRS)} pairs × [hard_or (≥1), hard_and (≥2)]")
    print("=" * 72)

    df_all = run_experiment()

    if df_all.empty:
        print("\n⚠️  沒有結果，請確認資料路徑與格式。")
    else:
        save_excel(df_all)

        print("\n── Best Overall（Hetero Vote/Pairs，跨所有 dataset 平均）──")
        hetero = df_all[df_all["Method"].str.startswith("Hetero_")]
        if not hetero.empty:
            print(hetero.groupby(["Method", "OCC"])["AUC"]
                        .mean().sort_values(ascending=False).round(4).to_string())

        print("\n── Baseline 對照（各 AE 單獨，跨所有 dataset 平均）──")
        base = df_all[df_all["Method"].str.startswith("B_")]
        if not base.empty:
            print(base.groupby(["Method", "OCC"])["AUC"]
                      .mean().sort_values(ascending=False).round(4).to_string())
