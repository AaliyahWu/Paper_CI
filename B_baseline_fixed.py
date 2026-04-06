"""
baseline B.py
=============
使用前測 (best_params.csv) 得出的最佳 AE × OCC 參數組合，
對 data/ 資料集做正式實驗。

AE 類型：AE、DAE、SAE、VAE（各自對應最佳架構）
OCC 方法：OCSVM、LOF、iForest（各自對應最佳架構）

在 test set 評估：AUC、F1、Recall、G-mean

輸出：results/B baseline.xlsx
  分頁：per_fold  /  summary  /  overall_mean

資料夾結構（DATA_ROOT）：
  data/
  └── <dataset_name>/
      ├── <prefix>-1tra.dat  ...  <prefix>-5tra.dat
      └── <prefix>-1tst.dat  ...  <prefix>-5tst.dat
"""

import os
import re
import copy
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT   = Path("data")
RESULTS_DIR = Path("results")
BEST_PARAMS = RESULTS_DIR / "best_params.csv"
OUTPUT_FILE = RESULTS_DIR / "B baseline.xlsx"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5

# AE 訓練超參數（與前測一致）
AE_EPOCHS     = 100
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1
SAE_SPARSITY  = 1e-3
VAE_BETA      = 1.0

AE_TYPES  = ["AE", "DAE", "SAE", "VAE"]
OCC_TYPES = ["OCSVM", "LOF", "iForest"]
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

# 瓶頸比例標籤 → 實際比例
BOTTLENECK_RATIOS = {"1/4": 0.25, "1/3": 1/3, "1/2": 0.5, "1/1": 1.0}


# ─────────────────────────── 讀取最佳參數 ────────────────────────────────────
def load_best_params():
    """讀取 best_params.csv，建立 {(AE, OCC): (n_layers, ratio_label)} 映射"""
    df = pd.read_csv(BEST_PARAMS)
    mapping = {}
    for _, row in df.iterrows():
        cfg = str(row["best_config"]).strip()   # e.g. "h2-1/3"
        m = re.match(r"h(\d+)-(.+)", cfg)
        if m:
            n_layers    = int(m.group(1))
            ratio_label = m.group(2)            # "1/4", "1/3", "1/2", "1/1"
            mapping[(str(row["AE"]).strip(), str(row["OCC"]).strip())] = (n_layers, ratio_label)
    return mapping


# ─────────────────────────── AE 模型定義 ─────────────────────────────────────
class AEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        for i in range(len(dims) - 1):
            d_in, d_out = dims[-(i+1)], dims[-(i+2)]
            act = nn.Sigmoid() if i == len(dims) - 2 else nn.ReLU()
            dec += [nn.Linear(d_in, d_out), act]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


class VAEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        base = []
        for i in range(len(dims) - 2):
            base += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        self.enc_base  = nn.Sequential(*base) if base else nn.Identity()
        mid = dims[-2] if len(dims) >= 2 else input_dim
        self.fc_mu     = nn.Linear(mid, dims[-1])
        self.fc_logvar = nn.Linear(mid, dims[-1])
        dec_dims = dims[::-1]
        dec = []
        for i in range(len(dec_dims) - 1):
            act = nn.Sigmoid() if i == len(dec_dims) - 2 else nn.ReLU()
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), act]
        self.decoder = nn.Sequential(*dec)

    def reparameterize(self, mu, lv):
        return mu + torch.exp(0.5 * lv) * torch.randn_like(lv)

    def forward(self, x):
        h  = self.enc_base(x)
        mu = self.fc_mu(h)
        lv = self.fc_logvar(h)
        z  = self.reparameterize(mu, lv)
        return self.decoder(z), z, mu, lv


# ─────────────────────────── AE 訓練 + 特徵提取 ──────────────────────────────
def train_and_extract(ae_type, X_maj_s, X_test_s, n_layers, n_units):
    """
    用 majority/normal 樣本（已 MinMaxScaled）訓練 AE，
    對 majority 和 test set 提取瓶頸特徵。
    回傳 (feat_maj, feat_test)
    """
    input_dim = X_maj_s.shape[1]
    model = VAEModel(input_dim, n_layers, n_units) if ae_type == "VAE" \
            else AEModel(input_dim, n_layers, n_units)

    optim  = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))
    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim.zero_grad()
            if ae_type == "DAE":
                xb_n = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
                xr, _ = model(xb_n)
                loss   = mse(xr, xb)
            elif ae_type == "SAE":
                xr, z = model(xb)
                loss   = mse(xr, xb) + SAE_SPARSITY * z.abs().mean()
            elif ae_type == "VAE":
                xr, _, mu, lv = model(xb)
                kl   = -0.5 * (1 + lv - mu.pow(2) - lv.exp()).mean()
                loss = mse(xr, xb) + VAE_BETA * kl
            else:
                xr, _ = model(xb)
                loss   = mse(xr, xb)
            loss.backward()
            optim.step()

    model.eval()

    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            if ae_type == "VAE":
                _, _, mu, _ = model(xt)
                return mu.numpy()
            else:
                _, z = model(xt)
                return z.numpy()

    return extract(X_maj_s), extract(X_test_s)


# ─────────────────────────── 評估指標 ────────────────────────────────────────
def gmean_score(y_true, y_pred_binary):
    """G-mean，正類=1（minority/anomaly），負類=0（majority/normal）"""
    cm = confusion_matrix(y_true, y_pred_binary, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return float(np.sqrt(sens * spec))
    return 0.0


def run_occ_eval(occ_type, feat_maj, feat_test, y_test, n_neighbors_cap):
    """
    在 AE 提取後的特徵空間訓練 OCC（只用 majority/normal 特徵），
    對 test set 計算 AUC、F1、Recall、G-mean。

    y_test  : 1=minority(anomaly), 0=majority(normal)

    Threshold 設定（訓練集 majority 分數第 90 百分位數）：
      與 Baseline A 相同做法。predict() 的內建 threshold 在某些 fold
      會使 test set 預測為 anomaly 的數量為 0（AUC > 0.6 但 F1=0）。
      改用 training majority anomaly score 的第 90 百分位數作為 threshold，
      對應 contamination=0.1 語義，跨 fold 穩定一致。
    """
    scaler      = MinMaxScaler()
    feat_maj_s  = scaler.fit_transform(feat_maj)
    feat_test_s = scaler.transform(feat_test)

    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)

    elif occ_type == "LOF":
        k = min(20, n_neighbors_cap)
        clf = LocalOutlierFactor(n_neighbors=k, novelty=True, contamination=0.1)
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)

    else:  # iForest
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)

    # threshold：training majority 分數的第 90 百分位數（與 Baseline A 一致）
    threshold = np.percentile(scores_maj, 90)
    y_pred = (scores_test >= threshold).astype(int)

    try:
        auc = roc_auc_score(y_test, scores_test) if len(np.unique(y_test)) >= 2 else float("nan")
    except Exception:
        auc = float("nan")

    f1  = f1_score(y_test, y_pred, pos_label=1, zero_division=0)
    rec = recall_score(y_test, y_pred, pos_label=1, zero_division=0)
    gm  = gmean_score(y_test, y_pred)

    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── KEEL .dat 解析 ──────────────────────────────────
def parse_keel_dat(filepath):
    """
    解析 KEEL .dat，回傳：
      X: float ndarray
      y: int ndarray，0=majority(normal), 1=minority(anomaly)
    """
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []

    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"):
            continue
        if s.lower() == "@data":
            data_start = True
            continue
        if data_start:
            rows.append(s)

    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)

    label_col = df.columns[-1]
    y_raw = df[label_col].values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    unique, counts = np.unique(y_raw, return_counts=True)
    minority_label = unique[np.argmin(counts)]
    y = (y_raw == minority_label).astype(int)   # minority=1 (anomaly), majority=0 (normal)

    return X, y


# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment(best_params_map):
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    all_records = []

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n{'='*65}")
        print(f"▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            file_prefix = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]

            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)

            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案")
                continue

            try:
                X_tra, y_tra = parse_keel_dat(tra_file)
                X_tst, y_tst = parse_keel_dat(tst_file)

                input_dim = X_tra.shape[1]
                X_maj = X_tra[y_tra == 0]                  # majority/normal samples

                if len(X_maj) < 5:
                    print(f"  [SKIP] Fold {fold}: 訓練集正常樣本不足 ({len(X_maj)})")
                    continue
                if y_tst.sum() == 0:
                    print(f"  [SKIP] Fold {fold}: 測試集無少數類樣本")
                    continue

                scaler    = MinMaxScaler()
                X_maj_s   = scaler.fit_transform(X_maj)
                X_tst_s   = scaler.transform(X_tst)
                n_nb_cap  = max(1, len(X_maj) - 1)

            except Exception as e:
                print(f"  [ERROR] Fold {fold} 資料載入失敗: {e}")
                continue

            # ── 每個 (AE, OCC) 組合：載入對應最佳架構，訓練 AE，評估 OCC ──
            for ae_type in AE_TYPES:
                for occ_type in OCC_TYPES:
                    key = (ae_type, occ_type)
                    if key not in best_params_map:
                        print(f"  [SKIP] {ae_type}×{occ_type}: best_params.csv 無此組合")
                        continue

                    n_layers, ratio_label = best_params_map[key]
                    ratio   = BOTTLENECK_RATIOS.get(ratio_label, 1/3)
                    n_units = max(2, round(input_dim * ratio))
                    cfg_label = f"h{n_layers}-{ratio_label}"

                    try:
                        feat_maj, feat_tst = train_and_extract(
                            ae_type, X_maj_s, X_tst_s, n_layers, n_units)
                        metrics = run_occ_eval(occ_type, feat_maj, feat_tst, y_tst, n_nb_cap)
                    except Exception as e:
                        print(f"  [ERROR] {ae_type}×{occ_type}({cfg_label}) Fold{fold}: {e}")
                        metrics = {m: float("nan") for m in METRIC_COLS}

                    row = {
                        "Dataset": ds_name,
                        "AE":      ae_type,
                        "OCC":     occ_type,
                        "Config":  cfg_label,
                        "Fold":    fold,
                        **metrics,
                    }
                    all_records.append(row)
                    print(f"  {ae_type:4s}×{occ_type:8s}({cfg_label}) Fold{fold}  "
                          f"AUC={metrics['AUC']:.4f}  F1={metrics['F1']:.4f}  "
                          f"Recall={metrics['Recall']:.4f}  G-mean={metrics['G-mean']:.4f}")

    return pd.DataFrame(all_records)


# ─────────────────────────── Excel 樣式定義 ──────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")   # 深藍（標題）
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")   # 淡藍（副標題）
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")   # 淺灰（交替行）

AE_FILL = {
    "AE":  PatternFill("solid", fgColor="DAEEF3"),   # 淡青
    "DAE": PatternFill("solid", fgColor="E2EFDA"),   # 淡綠
    "SAE": PatternFill("solid", fgColor="FFF2CC"),   # 淡黃
    "VAE": PatternFill("solid", fgColor="FCE4D6"),   # 淡橙
}

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def sc(cell, value, font=None, fill=None, align=None, fmt=None):
    cell.value  = value
    cell.border = THIN_BORDER
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    if fmt:   cell.number_format = fmt


def col_w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ─────────────────────────── per_fold 分頁 ───────────────────────────────────
def write_per_fold(ws, df):
    ws.title = "per_fold"
    headers = ["Dataset", "AE", "OCC", "Config", "Fold"] + METRIC_COLS

    for c, h in enumerate(headers, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        fill = AE_FILL.get(row["AE"])
        for c, col in enumerate(headers, 1):
            val  = row[col]
            cell = ws.cell(r, c)
            sc(cell, val,
               font=BODY_FONT,
               fill=fill,
               align=LEFT_ALIGN if c <= 2 else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS else None)

    for i, w in enumerate([28, 6, 10, 10, 6, 10, 10, 10, 10], 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


# ─────────────────────────── summary 分頁 ────────────────────────────────────
def write_summary(ws, df):
    """
    列：Dataset
    欄群組：AE (4) × OCC (3) = 12 組，每組 4 個指標 → 48 欄
    顯示 mean ± std（across 5 folds）
    """
    ws.title = "summary"

    grouped = (df.groupby(["Dataset", "AE", "OCC"])[METRIC_COLS]
                 .agg(["mean", "std"])
                 .reset_index())

    # Row 1: Dataset (span 2 rows) | AE groups (span OCC*Metric cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for ae in AE_TYPES:
        ae_span = len(OCC_TYPES) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + ae_span - 1)
        sc(ws.cell(1, col), ae,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=AE_FILL.get(ae),
           align=CENTER_ALIGN)

        # Row 2: OCC sub-headers (span Metric cols each)
        for occ in OCC_TYPES:
            occ_span = len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + occ_span - 1)
            sc(ws.cell(2, col), occ,
               font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            col += occ_span

    # Row 3: Metric sub-headers
    metric_header_row = 3
    col = 2
    for ae in AE_TYPES:
        for occ in OCC_TYPES:
            for metric in METRIC_COLS:
                sc(ws.cell(metric_header_row, col), metric,
                   font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                col += 1

    # Data rows (start at row 4)
    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, metric_header_row + 1):
        fill = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=fill, align=LEFT_ALIGN)
        col = 2
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                sub = grouped[
                    (grouped["Dataset"] == ds) &
                    (grouped["AE"]      == ae)  &
                    (grouped["OCC"]     == occ)
                ]
                for metric in METRIC_COLS:
                    try:
                        m = sub[(metric, "mean")].values[0]
                        s = sub[(metric, "std")].values[0]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, col), display,
                       font=BODY_FONT,
                       fill=AE_FILL.get(ae) if fill is None else fill,
                       align=CENTER_ALIGN)
                    col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(AE_TYPES) * len(OCC_TYPES) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B4"


# ─────────────────────────── overall_mean 分頁 ───────────────────────────────
def write_overall_mean(ws, df):
    """
    列：AE × OCC（12 組）
    欄：Config + 4 指標（mean ± std，across all datasets & folds）
    """
    ws.title = "overall_mean"

    total_cols = 3 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    title = ws["A1"]
    title.value     = "Overall Mean（across all datasets & folds）"
    title.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    title.alignment = CENTER_ALIGN

    headers = ["AE", "OCC", "Best Config"] + METRIC_COLS
    for c, h in enumerate(headers, 1):
        sc(ws.cell(2, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["AE", "OCC"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for ae in AE_TYPES:
        for occ in OCC_TYPES:
            fill = AE_FILL.get(ae)
            sc(ws.cell(r, 1), ae,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
            sc(ws.cell(r, 2), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)

            cfg_series = df[(df["AE"] == ae) & (df["OCC"] == occ)]["Config"]
            cfg = cfg_series.iloc[0] if not cfg_series.empty else "N/A"
            sc(ws.cell(r, 3), cfg, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)

            for c, metric in enumerate(METRIC_COLS, 4):
                try:
                    m = overall.loc[(ae, occ), (metric, "mean")]
                    s = overall.loc[(ae, occ), (metric, "std")]
                    display = f"{m:.4f} ± {s:.4f}"
                except Exception:
                    display = "N/A"
                sc(ws.cell(r, c), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            r += 1

    for i, w in enumerate([8, 12, 14, 22, 22, 22, 22], 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


# ─────────────────────────── Excel 存檔 ──────────────────────────────────────
def save_excel(df):
    wb = Workbook()
    ws_fold    = wb.active
    ws_summary = wb.create_sheet()
    ws_overall = wb.create_sheet()

    write_per_fold(ws_fold, df)
    write_summary(ws_summary, df)
    write_overall_mean(ws_overall, df)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 65)
    print("Baseline B：AE 特徵提取 × OCC（最佳參數）正式評估")
    print("AE  類型：AE / DAE / SAE / VAE")
    print("OCC 方法：OCSVM / LOF / iForest")
    print("=" * 65)

    best_params_map = load_best_params()
    print(f"\n已載入最佳參數（共 {len(best_params_map)} 組）：")
    for (ae, occ), (nl, rl) in sorted(best_params_map.items()):
        print(f"  {ae:4s} × {occ:8s}  =>  h{nl}-{rl}")

    print()
    df_results = run_experiment(best_params_map)

    if df_results.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_results)
        print("\n── Overall Mean（所有資料集平均）──")
        print(df_results.groupby(["AE", "OCC"])[METRIC_COLS].mean().round(4).to_string())
