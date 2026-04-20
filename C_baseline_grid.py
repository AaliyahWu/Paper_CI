"""
C_baseline_grid.py
==================
對每個 dataset × fold × (AE, OCC) 枚舉全部 12 種參數組合，
所有組合結果完整保留，並另設「best」分頁組存 AUC 最高的結果。

每個 fold 流程：
  1. 載入 KEEL .dat
  2. MinMaxScale 原始特徵（fit on X_maj）
  3. 對每個 AE 架構組合：訓練 AE，提取深度特徵
  4. 拼接原始特徵 + 深度特徵：hstack([X_maj_s, feat_maj])
  5. 對拼接特徵再做 MinMaxScale
  6. 訓練 OCC，記錄所有組合結果；另取 AUC 最高者存入 best

參數搜尋空間：
  n_layers  : [1, 2, 3]
  bottleneck: ["1/4", "1/3", "1/2", "1/1"]
  → 12 種組合 per (AE, OCC, fold)

AE 類型：AE, DAE, SAE, VAE
OCC 方法：OCSVM, LOF, iForest

輸出：results/C baseline grid.xlsx
  分頁（全部組合）：
    all_per_fold      所有 12 種組合每 fold 的原始結果
    all_summary       所有組合 mean ± std（across folds，per dataset）
    all_overall       所有組合全域平均
  分頁（最佳組合）：
    best_per_fold     每 (Dataset, AE, OCC, fold) 中 AUC 最高的組合
    best_summary      最佳組合 mean ± std（across folds，per dataset）
    best_overall      最佳組合全域平均
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT   = Path("data")
RESULTS_DIR = Path("results")
OUTPUT_FILE = RESULTS_DIR / "C baseline grid.xlsx"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5

AE_EPOCHS     = 100
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1
SAE_SPARSITY  = 1e-3
VAE_BETA      = 1.0

AE_TYPES    = ["AE", "DAE", "SAE", "VAE"]
OCC_TYPES   = ["OCSVM", "LOF", "iForest"]
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

N_LAYERS_LIST     = [1, 2, 3]
BOTTLENECK_RATIOS = {
    "1/4": 0.25,   # DF 佔 combined 20%  (與 B 相同)
    "1/3": 1/3,    # DF 佔 combined 25%  (與 B 相同)
    "1/2": 0.5,    # DF 佔 combined 33%  (與 B 相同)
    "1/1": 1.0,    # DF 佔 combined 50%  (與 B 相同)
    "2/1": 2.0,    # DF 佔 combined 67%  (C 新增)
    "3/1": 3.0,    # DF 佔 combined 75%  (C 新增)
    "4/1": 4.0,    # DF 佔 combined 80%  (C 新增)
}
ALL_CONFIGS = [f"h{nl}-{rl}" for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]
# 共 3 × 7 = 21 種組合


# ─────────────────────────── AE 模型定義 ─────────────────────────────────────
class AEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        for i in range(len(dims) - 1):
            d_in, d_out = dims[-(i+1)], dims[-(i+2)]
            act = nn.Sigmoid() if i == len(dims) - 2 else nn.ReLU()
            dec += [nn.Linear(d_in, d_out), act]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


class VAEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        base = []
        for i in range(len(dims) - 2):
            base += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        self.enc_base  = nn.Sequential(*base) if base else nn.Identity()
        mid = dims[-2] if len(dims) >= 2 else input_dim
        self.fc_mu     = nn.Linear(mid, dims[-1])
        self.fc_logvar = nn.Linear(mid, dims[-1])
        dec_dims = dims[::-1]
        dec = []
        for i in range(len(dec_dims) - 1):
            act = nn.Sigmoid() if i == len(dec_dims) - 2 else nn.ReLU()
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), act]
        self.decoder = nn.Sequential(*dec)

    def reparameterize(self, mu, lv):
        return mu + torch.exp(0.5 * lv) * torch.randn_like(lv)

    def forward(self, x):
        h  = self.enc_base(x)
        mu = self.fc_mu(h)
        lv = self.fc_logvar(h)
        z  = self.reparameterize(mu, lv)
        return self.decoder(z), z, mu, lv


# ─────────────────────────── AE 訓練 + 特徵提取 ──────────────────────────────
def train_and_extract(ae_type, X_maj_s, X_test_s, n_layers, n_units):
    input_dim = X_maj_s.shape[1]
    model = VAEModel(input_dim, n_layers, n_units) if ae_type == "VAE" \
            else AEModel(input_dim, n_layers, n_units)

    optim  = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))
    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim.zero_grad()
            if ae_type == "DAE":
                xb_n = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
                xr, _ = model(xb_n)
                loss   = mse(xr, xb)
            elif ae_type == "SAE":
                xr, z = model(xb)
                loss   = mse(xr, xb) + SAE_SPARSITY * z.abs().mean()
            elif ae_type == "VAE":
                xr, _, mu, lv = model(xb)
                kl   = -0.5 * (1 + lv - mu.pow(2) - lv.exp()).mean()
                loss = mse(xr, xb) + VAE_BETA * kl
            else:
                xr, _ = model(xb)
                loss   = mse(xr, xb)
            loss.backward()
            optim.step()

    model.eval()

    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            if ae_type == "VAE":
                _, _, mu, _ = model(xt)
                return mu.numpy()
            else:
                _, z = model(xt)
                return z.numpy()

    return extract(X_maj_s), extract(X_test_s)


# ─────────────────────────── 評估指標 ────────────────────────────────────────
def gmean_score(y_true, y_pred_binary):
    cm = confusion_matrix(y_true, y_pred_binary, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return float(np.sqrt(sens * spec))
    return 0.0


def run_occ_eval(occ_type, X_comb_maj, X_comb_test, y_test, n_neighbors_cap):
    scaler        = MinMaxScaler()
    X_comb_maj_s  = scaler.fit_transform(X_comb_maj)
    X_comb_test_s = scaler.transform(X_comb_test)

    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(X_comb_maj_s)
        scores_maj  = -clf.decision_function(X_comb_maj_s)
        scores_test = -clf.decision_function(X_comb_test_s)
    elif occ_type == "LOF":
        k = min(20, n_neighbors_cap)
        clf = LocalOutlierFactor(n_neighbors=k, novelty=True, contamination=0.1)
        clf.fit(X_comb_maj_s)
        scores_maj  = -clf.decision_function(X_comb_maj_s)
        scores_test = -clf.decision_function(X_comb_test_s)
    else:
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(X_comb_maj_s)
        scores_maj  = -clf.decision_function(X_comb_maj_s)
        scores_test = -clf.decision_function(X_comb_test_s)

    threshold = np.percentile(scores_maj, 90)
    y_pred    = (scores_test >= threshold).astype(int)

    try:
        auc = roc_auc_score(y_test, scores_test) if len(np.unique(y_test)) >= 2 else float("nan")
    except Exception:
        auc = float("nan")

    f1  = f1_score(y_test, y_pred, pos_label=1, zero_division=0)
    rec = recall_score(y_test, y_pred, pos_label=1, zero_division=0)
    gm  = gmean_score(y_test, y_pred)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── KEEL .dat 解析 ──────────────────────────────────
def parse_keel_dat(filepath):
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"):
            continue
        if s.lower() == "@data":
            data_start = True
            continue
        if data_start:
            rows.append(s)
    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    y_raw     = df[label_col].values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    unique, counts = np.unique(y_raw, return_counts=True)
    minority_label = unique[np.argmin(counts)]
    y = (y_raw == minority_label).astype(int)
    return X, y


# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment():
    """
    回傳：
      df_all  — 所有 12 種組合每筆結果
      df_best — 每 (Dataset, AE, OCC, fold) 中 AUC 最高的那筆
    """
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    param_configs = [(nl, rl) for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]
    all_records  = []
    best_records = []

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n{'='*65}")
        print(f"▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            file_prefix  = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]

            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)

            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案")
                continue

            try:
                X_tra, y_tra = parse_keel_dat(tra_file)
                X_tst, y_tst = parse_keel_dat(tst_file)
                input_dim = X_tra.shape[1]
                X_maj     = X_tra[y_tra == 0]

                if len(X_maj) < 5:
                    print(f"  [SKIP] Fold {fold}: 訓練集正常樣本不足 ({len(X_maj)})")
                    continue
                if y_tst.sum() == 0:
                    print(f"  [SKIP] Fold {fold}: 測試集無少數類樣本")
                    continue

                scaler_orig = MinMaxScaler()
                X_maj_s     = scaler_orig.fit_transform(X_maj)
                X_tst_s     = scaler_orig.transform(X_tst)
                n_nb_cap    = max(1, len(X_maj) - 1)

            except Exception as e:
                print(f"  [ERROR] Fold {fold} 資料載入失敗: {e}")
                continue

            for ae_type in AE_TYPES:
                # 每個 AE 架構只訓練一次，快取拼接特徵
                comb_features = {}
                for n_layers, ratio_label in param_configs:
                    ratio   = BOTTLENECK_RATIOS[ratio_label]
                    n_units = max(2, round(input_dim * ratio))
                    try:
                        feat_maj, feat_tst = train_and_extract(
                            ae_type, X_maj_s, X_tst_s, n_layers, n_units)
                        X_comb_maj  = np.hstack([X_maj_s, feat_maj])
                        X_comb_test = np.hstack([X_tst_s, feat_tst])
                        combined_dim = X_comb_maj.shape[1]
                        comb_features[(n_layers, ratio_label)] = (X_comb_maj, X_comb_test, combined_dim)
                    except Exception as e:
                        print(f"  [ERROR] AE={ae_type} h{n_layers}-{ratio_label} Fold{fold}: {e}")

                for occ_type in OCC_TYPES:
                    best_auc = -1.0
                    best_row = None

                    for (n_layers, ratio_label), (X_comb_maj, X_comb_test, combined_dim) in comb_features.items():
                        cfg_label = f"h{n_layers}-{ratio_label}"
                        try:
                            metrics = run_occ_eval(
                                occ_type, X_comb_maj, X_comb_test, y_tst, n_nb_cap)
                        except Exception as e:
                            print(f"  [ERROR] {ae_type}×{occ_type}({cfg_label}) Fold{fold}: {e}")
                            metrics = {m: float("nan") for m in METRIC_COLS}

                        row = {
                            "Dataset":      ds_name,
                            "AE":           ae_type,
                            "OCC":          occ_type,
                            "Config":       cfg_label,
                            "Fold":         fold,
                            "Combined_Dim": combined_dim,
                            **metrics,
                        }
                        all_records.append(row)

                        auc = metrics["AUC"] if not np.isnan(metrics["AUC"]) else -1.0
                        if auc > best_auc:
                            best_auc = auc
                            best_row = row

                    if best_row is not None:
                        best_records.append(best_row)
                        print(
                            f"  {ae_type:4s}×{occ_type:8s} best={best_row['Config']:12s} "
                            f"dim={best_row['Combined_Dim']:3d} Fold{fold}  "
                            f"AUC={best_row['AUC']:.4f}  F1={best_row['F1']:.4f}  "
                            f"Recall={best_row['Recall']:.4f}  G-mean={best_row['G-mean']:.4f}"
                        )

    return pd.DataFrame(all_records), pd.DataFrame(best_records)


# ─────────────────────────── Excel 樣式定義 ──────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

AE_FILL = {
    "AE":  PatternFill("solid", fgColor="DAEEF3"),
    "DAE": PatternFill("solid", fgColor="E2EFDA"),
    "SAE": PatternFill("solid", fgColor="FFF2CC"),
    "VAE": PatternFill("solid", fgColor="FCE4D6"),
}

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def sc(cell, value, font=None, fill=None, align=None, fmt=None):
    cell.value  = value
    cell.border = THIN_BORDER
    if font:  cell.font         = font
    if fill:  cell.fill         = fill
    if align: cell.alignment    = align
    if fmt:   cell.number_format = fmt


def col_w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ─────────────────────────── per_fold 分頁 ───────────────────────────────────
def write_per_fold(ws, df, title):
    ws.title = title
    headers = ["Dataset", "AE", "OCC", "Config", "Fold", "Combined_Dim"] + METRIC_COLS

    for c, h in enumerate(headers, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        fill = AE_FILL.get(row["AE"])
        for c, col in enumerate(headers, 1):
            sc(ws.cell(r, c), row[col],
               font=BODY_FONT, fill=fill,
               align=LEFT_ALIGN if c <= 2 else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS else None)

    for i, w in enumerate([28, 6, 10, 14, 6, 12, 10, 10, 10, 10], 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


# ─────────────────────────── summary 分頁 ────────────────────────────────────
def write_summary_all(ws, df):
    """all_summary：Dataset | AE × OCC × Config × Metric"""
    ws.title = "all_summary"
    configs = ALL_CONFIGS

    grouped = (df.groupby(["Dataset", "AE", "OCC", "Config"])[METRIC_COLS]
                 .agg(["mean", "std"])
                 .reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for ae in AE_TYPES:
        ae_span = len(OCC_TYPES) * len(configs) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + ae_span - 1)
        sc(ws.cell(1, col), ae,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=AE_FILL.get(ae), align=CENTER_ALIGN)
        for occ in OCC_TYPES:
            occ_span = len(configs) * len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + occ_span - 1)
            sc(ws.cell(2, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            for cfg in configs:
                cfg_span = len(METRIC_COLS)
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + cfg_span - 1)
                sc(ws.cell(3, col), cfg, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                for metric in METRIC_COLS:
                    sc(ws.cell(4, col), metric, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                    col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 5):
        fill = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=fill, align=LEFT_ALIGN)
        col = 2
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                for cfg in configs:
                    sub = grouped[
                        (grouped["Dataset"] == ds) &
                        (grouped["AE"]      == ae)  &
                        (grouped["OCC"]     == occ) &
                        (grouped["Config"]  == cfg)
                    ]
                    for metric in METRIC_COLS:
                        try:
                            m = sub[(metric, "mean")].values[0]
                            s = sub[(metric, "std")].values[0]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, col), display,
                           font=BODY_FONT,
                           fill=AE_FILL.get(ae) if fill is None else fill,
                           align=CENTER_ALIGN)
                        col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(AE_TYPES) * len(OCC_TYPES) * len(configs) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B5"


def write_summary_best(ws, df):
    """best_summary：Dataset | AE × OCC × Metric"""
    ws.title = "best_summary"

    grouped = (df.groupby(["Dataset", "AE", "OCC"])[METRIC_COLS]
                 .agg(["mean", "std"])
                 .reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for ae in AE_TYPES:
        ae_span = len(OCC_TYPES) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + ae_span - 1)
        sc(ws.cell(1, col), ae,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=AE_FILL.get(ae), align=CENTER_ALIGN)
        for occ in OCC_TYPES:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + len(METRIC_COLS) - 1)
            sc(ws.cell(2, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            col += len(METRIC_COLS)

    col = 2
    for ae in AE_TYPES:
        for occ in OCC_TYPES:
            for metric in METRIC_COLS:
                sc(ws.cell(3, col), metric, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 4):
        fill = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=fill, align=LEFT_ALIGN)
        col = 2
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                sub = grouped[
                    (grouped["Dataset"] == ds) &
                    (grouped["AE"]      == ae)  &
                    (grouped["OCC"]     == occ)
                ]
                for metric in METRIC_COLS:
                    try:
                        m = sub[(metric, "mean")].values[0]
                        s = sub[(metric, "std")].values[0]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, col), display,
                       font=BODY_FONT,
                       fill=AE_FILL.get(ae) if fill is None else fill,
                       align=CENTER_ALIGN)
                    col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(AE_TYPES) * len(OCC_TYPES) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B4"


# ─────────────────────────── overall 分頁 ────────────────────────────────────
def write_overall_all(ws, df):
    """all_overall：AE × OCC × Config 全域平均"""
    ws.title = "all_overall"
    configs = ALL_CONFIGS

    total_cols = 3 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（all configs, all datasets & folds）- Baseline C Grid: OF + DF → OCC"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["AE", "OCC", "Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["AE", "OCC", "Config"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for ae in AE_TYPES:
        for occ in OCC_TYPES:
            for cfg in configs:
                fill = AE_FILL.get(ae)
                sc(ws.cell(r, 1), ae,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 2), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 3), cfg, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                for i, metric in enumerate(METRIC_COLS, 4):
                    try:
                        m = overall.loc[(ae, occ, cfg), (metric, "mean")]
                        s = overall.loc[(ae, occ, cfg), (metric, "std")]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, i), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                r += 1

    for i, w in enumerate([8, 12, 14] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


def write_overall_best(ws, df):
    """best_overall：AE × OCC 全域平均 + 最常選中的 Config"""
    ws.title = "best_overall"

    total_cols = 3 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（best config per fold, all datasets）- Baseline C Grid: OF + DF → OCC"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["AE", "OCC", "Most Freq Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["AE", "OCC"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for ae in AE_TYPES:
        for occ in OCC_TYPES:
            fill = AE_FILL.get(ae)
            sc(ws.cell(r, 1), ae,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
            sc(ws.cell(r, 2), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            cfg_series = df[(df["AE"] == ae) & (df["OCC"] == occ)]["Config"]
            cfg = cfg_series.mode().iloc[0] if not cfg_series.empty else "N/A"
            sc(ws.cell(r, 3), cfg, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            for i, metric in enumerate(METRIC_COLS, 4):
                try:
                    m = overall.loc[(ae, occ), (metric, "mean")]
                    s = overall.loc[(ae, occ), (metric, "std")]
                    display = f"{m:.4f} ± {s:.4f}"
                except Exception:
                    display = "N/A"
                sc(ws.cell(r, i), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            r += 1

    for i, w in enumerate([8, 12, 18] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


# ─────────────────────────── Excel 存檔 ──────────────────────────────────────
def save_excel(df_all, df_best):
    wb = Workbook()

    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()
    ws6 = wb.create_sheet()

    write_per_fold(     ws1, df_all,  "all_per_fold")
    write_summary_all(  ws2, df_all)
    write_overall_all(  ws3, df_all)
    write_per_fold(     ws4, df_best, "best_per_fold")
    write_summary_best( ws5, df_best)
    write_overall_best( ws6, df_best)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 65)
    print("Baseline C Grid Search：OF_maj + DF_maj → OCC（全參數搜尋）")
    print("AE  類型：AE / DAE / SAE / VAE")
    print("OCC 方法：OCSVM / LOF / iForest")
    print(f"搜尋空間：n_layers={N_LAYERS_LIST} × bottleneck={list(BOTTLENECK_RATIOS.keys())}")
    print(f"          共 {len(N_LAYERS_LIST) * len(BOTTLENECK_RATIOS)} 種組合 per (AE, OCC, fold)")
    print("選擇準則：AUC 最高")
    print("分頁：all_per_fold / all_summary / all_overall")
    print("      best_per_fold / best_summary / best_overall")
    print("=" * 65)

    df_all, df_best = run_experiment()

    if df_all.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_all, df_best)
        print("\n── Best Config Overall Mean（所有資料集平均）──")
        print(df_best.groupby(["AE", "OCC"])[METRIC_COLS].mean().round(4).to_string())
