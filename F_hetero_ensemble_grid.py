"""
F_hetero_ensemble_grid.py
=========================
異質 AE Ensemble（Heterogeneous AE Ensemble）
=============================================

【什麼是「異質」？】

 同質（D/E 已做）：
   同一 AE 架構（如 VAE），三個特徵空間（OF / DF / OF+DF）的 OCC 做 voting。
   問題：OF+DF 包含 DF，特徵高度相關，多樣性不足，ensemble 無法帶來互補效果。

 異質（本檔）：
   四種不同 AE 架構（AE / DAE / SAE / VAE），同一特徵空間（DF_maj），
   各自提取的深度特徵完全獨立，再對四個 OCC 做 voting。
   原理：
     • AE  學習線性壓縮（最小重建誤差）
     • DAE 學習去噪（對 noise 魯棒的表示）
     • SAE 學習稀疏表示（特徵分離性更強）
     • VAE 學習機率分佈（隨機潛在空間）
   → 四種 DF 捕捉同一資料的不同面向，voting 可真正互補彼此的缺陷。

【方法總覽】每個 fold 會產出以下 Method 欄位值：

  ── Baseline（複製 B 結果，供直接對照）──────────────────────────────
  B_AE     : 僅用 AE 的 DF_maj → OCC（與 B baseline 相同邏輯，但 config 可能不同）
  B_DAE    : 僅用 DAE 的 DF_maj → OCC
  B_SAE    : 僅用 SAE 的 DF_maj → OCC
  B_VAE    : 僅用 VAE 的 DF_maj → OCC
  → 共 4 種 baseline

  ── Hetero Vote（全部 4 種 AE 一起投票，4 種 scheme）─────────────────
  Hetero_Vote_eq        : 4 個 OCC normalized score 等權平均（w = 1/4 each），soft
  Hetero_Vote_istd      : 4 個 OCC normalized score inverse-std 加權，soft
  Hetero_Vote_hard_or   : 4 個 OCC binary 預測 OR  合併（≥1 票判 anomaly），hard
  Hetero_Vote_hard_maj  : 4 個 OCC binary 預測 majority 合併（≥2 票判 anomaly），hard
  → 共 4 種

  ── Hetero Pair（C(4,2)=6 種 AE pair × 4 種 scheme）───────────────────
  Hetero_Pair_AE_DAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_AE_SAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_AE_VAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_DAE_SAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_DAE_VAE_eq / _istd / _hard_or / _hard_and
  Hetero_Pair_SAE_VAE_eq / _istd / _hard_or / _hard_and
  → 共 6 × 4 = 24 種

  總計：4（baseline）+ 4（Vote）+ 24（Pair）= 32 種 Method

【Voting 說明】
  Soft（連續分數加權平均）：
    eq   → 所有成員等權；istd → w_i ∝ 1/std(s_i_norm)，再歸一化使 Σw=1
    最終分數 → 用 train-maj 第 90 百分位切閾值（與 baseline 一致）
  Hard（二元票數）：
    各成員先用自己的 train-maj 第 90 百分位切成 0/1（即 baseline 邏輯），
    再用以下規則合併：
      hard_or   (2-AE pair / 4-AE vote 都用) : ≥1 票判 anomaly（recall 高）
      hard_and  (2-AE pair 用)               : ≥2 票判 anomaly（precision 高）
      hard_maj  (4-AE vote 用)               : ≥2 票判 anomaly（多數決，與 D 一致）
    AUC 用 vote_count 當連續 ranking 分數
    → 同一 vote 系列（or/and 或 or/maj）的 AUC 必然相同，僅 F1/Recall/G-mean 不同

【Grid Search 空間】（與 B/C/D/E 完全一致）
  n_layers   : [1, 2, 3]
  bottleneck : ["1/4", "1/3", "1/2", "1/1", "2/1", "3/1", "4/1"]
  → 21 種 AE 架構 per (AE_type, OCC, fold)

【Config 選擇說明】
  USE_BEST_CONFIG = True（預設）：
    從 B_baseline_grid.xlsx 的 best_overall 分頁讀取 "Most Freq Config" 欄位，
    作為各 (AE_type, OCC_type) 的固定 config（每個 AE 可能使用不同 config）。
    → 每個 (AE, OCC) 只跑 1 種 config，計算量少，適合快速驗證。

  USE_BEST_CONFIG = False：
    對所有 21 種 config 做完整 grid search，輸出的 best_overall 取各
    (Dataset, Method, AE, OCC) 中平均 AUC 最高的 config，邏輯與 D/E 一致。

【輸出】results/F_hetero_ensemble_grid.xlsx
  分頁（格式與 D/E 一致）：
    all_per_fold  : 所有 (Dataset, Method, AE, OCC, Config, Fold) 原始結果
    all_summary   : mean ± std across folds（per Dataset × Method × AE × OCC × Config）
    all_overall   : 全域平均（across 所有 dataset 與 fold）
    best_per_fold : 各 (Dataset, Method, AE, OCC) 中 AUC 最高 config 的 fold 資料
    best_summary  : best config 的 mean ± std
    best_overall  : best config 的全域平均，格式 "mean ± std"（與 D/E 一致）
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from itertools import combinations

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────
DATA_ROOT       = Path("data")
RESULTS_DIR     = Path("results")
OUTPUT_FILE     = RESULTS_DIR / "F_hetero_ensemble_grid.xlsx"
B_GRID_XLSX     = Path("results/B baseline grid.xlsx")   # 讀取最佳 config 用
RESULTS_DIR.mkdir(exist_ok=True)

# 若 True → 每個 (AE, OCC) 使用 B grid 的最佳 config；否則 full grid search
USE_BEST_CONFIG = False

# ─────────────────────────── 搜尋空間 ────────────────────────────────
AE_TYPES          = ["AE", "DAE", "SAE", "VAE"]
OCC_TYPES         = ["OCSVM", "LOF", "iForest"]
N_LAYERS_LIST     = [1, 2, 3]
BOTTLENECK_RATIOS = {
    "1/4": 1/4, "1/3": 1/3, "1/2": 1/2, "1/1": 1.0,
    "2/1": 2.0, "3/1": 3.0, "4/1": 4.0,
}
EPS = 1e-8

# ── AE 訓練超參數（與 B_baseline_grid.py 完全一致）─────────────────
AE_EPOCHS     = 100          # B 用 100，F 原本錯誤寫成 50
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1          # DAE 加入的 Gaussian noise 標準差
SAE_SPARSITY  = 1e-3         # SAE L1 稀疏懲罰係數
VAE_BETA      = 1.0          # VAE KL 權重（B 用 1.0；F 原本錯誤用 0.001）

# Pair combinations: C(4,2) = 6
AE_PAIRS = list(combinations(AE_TYPES, 2))

METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

# ─────────────── AE 模型定義（與 B_baseline_grid.py 完全一致）───────────────
# 架構設計：dims = [input_dim] + [n_units] * n_layers（平坦隱層，不做逐層縮小）
# F 原本用 get_hidden_dims() 做錐形結構，與 B 不符，現已改回平坦結構。

class AEModel(nn.Module):
    """AE / DAE / SAE 共用架構（同 B_baseline_grid.py）"""
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        dec_dims = dims[::-1]
        for i in range(len(dec_dims) - 1):
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), nn.ReLU()]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


class VAEModel(nn.Module):
    """VAE 架構（同 B_baseline_grid.py）"""
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        base = []
        for i in range(len(dims) - 2):
            base += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        self.enc_base  = nn.Sequential(*base) if base else nn.Identity()
        mid = dims[-2] if len(dims) >= 2 else input_dim
        self.fc_mu     = nn.Linear(mid, dims[-1])
        self.fc_logvar = nn.Linear(mid, dims[-1])
        dec_dims = dims[::-1]
        dec = []
        for i in range(len(dec_dims) - 1):
            act = nn.Sigmoid() if i == len(dec_dims) - 2 else nn.ReLU()
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), act]
        self.decoder = nn.Sequential(*dec)

    def reparameterize(self, mu, lv):
        return mu + torch.exp(0.5 * lv) * torch.randn_like(lv)

    def encode(self, x):
        h = self.enc_base(x)
        return self.fc_mu(h), self.fc_logvar(h)

    def forward(self, x):
        h  = self.enc_base(x)
        mu = self.fc_mu(h);  lv = self.fc_logvar(h)
        z  = self.reparameterize(mu, lv)
        return self.decoder(z), z, mu, lv

def train_and_extract(ae_type, X_maj_s, X_test_s, n_layers, n_units):
    """
    訓練指定 AE 並提取深度特徵（DF）。與 B_baseline_grid.py 完全對齊。

    參數：
      ae_type  : 'AE' | 'DAE' | 'SAE' | 'VAE'
      X_maj_s  : MinMaxScaled train-majority（shape: [n_maj, input_dim]）
      X_test_s : MinMaxScaled test（shape: [n_test, input_dim]）
      n_layers : 隱層數（1/2/3）
      n_units  : 每層寬度 = max(2, round(input_dim × ratio))（與 B 一致）

    各 AE 訓練細節（與 B 對齊）：
      AE  : MSE loss，encoder 輸出 z（ReLU）
      DAE : 加 noise（std=DAE_NOISE）→ encode → reconstruct 乾淨 x；提取用乾淨 x
      SAE : MSE + SAE_SPARSITY × z.abs().mean()（L1 稀疏，F 原本錯用 KL divergence）
      VAE : MSE + VAE_BETA × KL（VAE_BETA=1.0，F 原本錯用 0.001）；提取取 mu

    返回：(DF_train, DF_test)，shape = (n, n_units)
    """
    input_dim = X_maj_s.shape[1]
    model = VAEModel(input_dim, n_layers, n_units) if ae_type == "VAE" \
            else AEModel(input_dim, n_layers, n_units)

    optim  = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))
    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True,
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim.zero_grad()
            if ae_type == "DAE":
                xb_n = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
                xr, _ = model(xb_n)
                loss  = mse(xr, xb)
            elif ae_type == "SAE":
                xr, z = model(xb)
                loss  = mse(xr, xb) + SAE_SPARSITY * z.abs().mean()
            elif ae_type == "VAE":
                xr, _, mu, lv = model(xb)
                kl   = -0.5 * (1 + lv - mu.pow(2) - lv.exp()).mean()
                loss = mse(xr, xb) + VAE_BETA * kl
            else:  # AE
                xr, _ = model(xb)
                loss  = mse(xr, xb)
            loss.backward()
            optim.step()

    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            if ae_type == "VAE":
                mu, _ = model.encode(xt)   # 取 mu，不取樣
                return mu.numpy()
            else:
                _, z = model(xt)
                return z.numpy()

    return extract(X_maj_s), extract(X_test_s)


# ─────────────────────────── OCC 訓練 ────────────────────────────────
def train_occ_scores(occ_type, X_maj, X_tst, n_nb_cap):
    n_nb = min(n_nb_cap, max(2, len(X_maj) - 1))
    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(X_maj)
        sm = -clf.decision_function(X_maj)
        st = -clf.decision_function(X_tst)
    elif occ_type == "LOF":
        clf = LocalOutlierFactor(n_neighbors=n_nb, contamination=0.1, novelty=True)
        clf.fit(X_maj)
        sm = -clf.decision_function(X_maj)
        st = -clf.decision_function(X_tst)
    else:
        clf = IsolationForest(contamination=0.1, random_state=42)
        clf.fit(X_maj)
        sm = -clf.decision_function(X_maj)
        st = -clf.decision_function(X_tst)
    return sm, st


def normalize_by_majority(sm, st):
    lo, hi = sm.min(), sm.max()
    if hi - lo < EPS:
        return np.zeros_like(sm), np.zeros_like(st)
    return (sm - lo) / (hi - lo), (st - lo) / (hi - lo)


def metrics_from(y_true, y_pred, s_cont):
    try:   auc = roc_auc_score(y_true, s_cont)
    except: auc = np.nan
    if y_pred.sum() == 0:
        return {"AUC": auc, "F1": 0.0, "Recall": 0.0, "G-mean": 0.0}
    f1  = f1_score(y_true, y_pred, zero_division=0)
    rec = recall_score(y_true, y_pred, zero_division=0)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    gm   = np.sqrt(rec * spec)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


def weighted_soft_vote(scores_norm_maj, scores_norm_tst, scheme, y_tst):
    """
    對多個已 normalize 的 OCC score 做加權 soft voting。

    參數：
      scores_norm_maj : list[np.ndarray]，各 OCC 在 train-maj 上的 normalized score
      scores_norm_tst : list[np.ndarray]，各 OCC 在 test 上的 normalized score
                        normalize 用 train-maj 的 min/max，不碰 test 統計，無 data leak
      scheme          : 'eq' | 'istd'

    加權方式：
      eq   → 所有 OCC 等權重：w_i = 1/n（n = OCC 數量）
      istd → inverse-std 加權：w_i ∝ 1 / std(scores_norm_maj[i])
             std 越小表示該 OCC 在訓練集上的分數越穩定，給予更高的信任度。
             公式：w_i = (1/std_i) / Σ(1/std_j)，確保 Σw_i = 1
             EPS 保護分母，避免 std=0 時除以零。

    閾值：
      取加權合併後的 train-maj score 的第 90 百分位數為 threshold
      （與 A/B/C/D/E 一致，確保所有方法可直接比較）

    返回：
      metrics dict，含 AUC / F1 / Recall / G-mean
    """
    n = len(scores_norm_maj)
    if scheme == "eq":
        weights = np.ones(n) / n
    else:  # istd
        stds = np.array([s.std() + EPS for s in scores_norm_maj])
        inv_stds = 1.0 / stds
        weights = inv_stds / inv_stds.sum()   # 歸一化，使 Σw_i = 1

    s_maj_w = sum(w * s for w, s in zip(weights, scores_norm_maj))
    s_tst_w = sum(w * s for w, s in zip(weights, scores_norm_tst))
    thr      = np.percentile(s_maj_w, 90)
    y_pred   = (s_tst_w >= thr).astype(int)
    return metrics_from(y_tst, y_pred, s_tst_w)


def hard_vote(y_pred_list, y_tst, rule):
    """
    對多個 OCC 的 binary 預測做 hard voting。

    參數：
      y_pred_list : list[np.ndarray]，各 OCC 在 test 上的 0/1 預測
                    （由各 OCC 用自己 train-maj 的第 90 百分位切閾值產生，
                    與 baseline B 邏輯一致）
      y_tst       : 真實標籤
      rule        : 二元預測的合併規則：
                    "or"   → ≥1 票判 anomaly（任一 OCC 偵測到就視為異常）
                    "and"  → 全票判 anomaly（所有 OCC 都同意才視為異常）
                    "maj"  → 多數決：≥ ceil(n/2) 票判 anomaly
                             例如 n=4 時 ≥2 票，與 D 的 Vote3_hard 邏輯一致
                             （D 用 ≥2/3 即 ceil(3/2)=2，本函式 n=4 時 ceil(4/2)=2）
                    "and" 對於 n=2 等於 "maj"；n>2 時更嚴格

    AUC 計算：
      用連續 vote_count 作為 ranking 分數（範圍 0~n）
      → 同一組成員的 _hard_or / _hard_and / _hard_maj 三者 AUC 必然相同
      → F1 / Recall / G-mean 因二元預測門檻不同而有差異

    回傳：
      metrics dict（AUC / F1 / Recall / G-mean）
    """
    n = len(y_pred_list)
    vote_count = np.sum([y.astype(int) for y in y_pred_list], axis=0)  # 0~n

    if rule == "or":
        thr_votes = 1
    elif rule == "and":
        thr_votes = n
    elif rule == "maj":
        thr_votes = int(np.ceil(n / 2))   # n=4 → 2; n=3 → 2; n=2 → 1
        # n=2 majority = ≥1 等於 OR；建議 2-pair 不要使用 maj
    else:
        raise ValueError(f"未知的 rule: {rule}")

    y_pred = (vote_count >= thr_votes).astype(int)
    return metrics_from(y_tst, y_pred, vote_count.astype(float))


# ─────────────────────────── 資料讀取 ────────────────────────────────
def load_dat(path):
    lines = open(path, encoding="utf-8", errors="ignore").readlines()
    attrs = [l.split()[1] for l in lines if l.strip().lower().startswith("@attribute")]
    rows  = [l.strip().split(",") for l in lines
             if l.strip() and not l.startswith("@") and not l.startswith("%")]
    df = pd.DataFrame(rows, columns=attrs)
    class_col = attrs[-1]
    y_raw = df[class_col].str.strip().values
    X = df.iloc[:, :-1].apply(pd.to_numeric, errors="coerce").fillna(0).values.astype(np.float32)
    uniq, cnts = np.unique(y_raw, return_counts=True)
    minority = uniq[np.argmin(cnts)]
    y = (y_raw == minority).astype(int)
    return X, y


def scan_folds(root):
    out = {}
    for ds in sorted(root.iterdir()):
        if not ds.is_dir(): continue
        files = list(ds.glob("*.dat"))
        tra = sorted([f for f in files if "tra.dat" in f.name])
        tst = sorted([f for f in files if "tst.dat" in f.name])
        pairs = list(zip(tra, tst))
        if pairs: out[ds.name] = pairs
    return out


# ─────────────────────────── Best config 讀取 ────────────────────────
def load_best_configs():
    """從 B_baseline_grid.xlsx 讀取各 (AE, OCC) 最佳 config"""
    try:
        df = pd.read_excel(B_GRID_XLSX, sheet_name="best_overall", header=1)
        best = {}
        for _, r in df.iterrows():
            ae, occ, cfg = r["AE"], r["OCC"], r["Most Freq Config"]
            best[(ae, occ)] = cfg
        return best
    except Exception as e:
        print(f"[警告] 無法讀取 {B_GRID_XLSX}：{e}，改用 h2-1/2 作為預設")
        return {}


def parse_config(cfg):
    """
    將 config 字串解析成 (n_layers, ratio_label)。
    例：'h2-1/2' → (2, '1/2')，'h3-4/1' → (3, '4/1')

    若格式不符（cfg 不是 'hN-ratio' 格式），回傳預設值 (2, '1/2')。
    預設值選 h2-1/2 的原因：B baseline grid 中此 config 在多數 (AE, OCC)
    組合下出現頻率最高，是最具代表性的「中性」config。
    """
    m = re.match(r"h(\d+)-(.+)", cfg)
    if not m:
        return 2, "1/2"
    return int(m.group(1)), m.group(2)


# ─────────────────────────── 主程式 ──────────────────────────────────
def run_experiment():
    """
    主實驗迴圈。

    每個 (dataset, occ_type, config, fold) 的處理步驟：
      1. 讀資料、MinMaxScale（fit on train, transform both）
      2. 各 AE 各自訓練並提取 DF_maj / DF_tst（再做第二次 MinMaxScale）
      3. 各 AE 的 DF 分別訓 OCC，得到 raw anomaly score
      4. raw score 做 train-maj min-max normalize（無 data leak）
      5. 輸出：
         - B_{ae_type}       : 各 AE 單獨的 OCC 結果（baseline 對照）
         - Hetero_Vote_eq/istd : 4 AE 全部參與 soft voting
         - Hetero_Pair_*_eq/istd : 6 對 AE pair soft voting

    返回：
      df_all : 所有 (method, config, fold) 的完整記錄，用於 all_per_fold 輸出。
               best_per_fold 在 save_excel() 內從 df_all 動態計算，
               做法：per (Dataset, Method, AE, OCC) 取 fold 平均 AUC 最高的 config。
    """
    dataset_folds = scan_folds(DATA_ROOT)
    if not dataset_folds:
        print(f"[錯誤] 找不到資料集，請確認 {DATA_ROOT}/ 目錄")
        return pd.DataFrame()

    best_configs = load_best_configs() if USE_BEST_CONFIG else {}
    all_records  = []

    for ds_name, fold_pairs in dataset_folds.items():
        try:
            X0, y0 = load_dat(str(fold_pairs[0][0]))
            input_dim = X0.shape[1]
        except Exception as e:
            print(f"[跳過] {ds_name}: {e}"); continue

        n_total = sum(1 for tp in fold_pairs for _ in [tp])  # noqa: 僅為佔位，實際用 len(fold_pairs)
        print(f"\n[{ds_name}] input_dim={input_dim}, folds={len(fold_pairs)}")

        for occ_type in OCC_TYPES:
            # ── USE_BEST_CONFIG=True：各 AE 用自己的 best config，fold 為內層 ──
            # 【修正】不能以 config 為外層迴圈：
            #   因為各 AE 的 best config 不同（如 AE→h1-1/2，VAE→h1-1/3），
            #   每輪 cfg 下只有 1 個 AE 可以跑，觸發 len(ae_features)<2 全部跳過。
            #   正確做法：fold 內讓所有 AE 各自用自己的 best config 獨立訓練。
            if USE_BEST_CONFIG and best_configs:
                ae_best_cfg = {
                    ae_type: best_configs.get((ae_type, occ_type), "h2-1/2")
                    for ae_type in AE_TYPES
                }
                from collections import Counter
                report_cfg = Counter(ae_best_cfg.values()).most_common(1)[0][0]

                for fold_idx, (tra_path, tst_path) in enumerate(fold_pairs, 1):
                    try:
                        X_tra, y_tra = load_dat(str(tra_path))
                        X_tst, y_tst = load_dat(str(tst_path))
                    except: continue
                    if y_tst.sum() < 1: continue

                    sc_orig = MinMaxScaler()
                    X_tra_s = sc_orig.fit_transform(X_tra)
                    X_tst_s = sc_orig.transform(X_tst)
                    X_maj_s = X_tra_s[y_tra == 0]
                    if len(X_maj_s) < 5: continue
                    n_nb_cap = min(20, len(X_maj_s) - 1)

                    # 各 AE 各自用自己的 best config 訓練，彼此獨立
                    ae_features = {}  # ae_type → (DF_maj_s, DF_tst_s, cfg_label)
                    for ae_type in AE_TYPES:
                        cfg = ae_best_cfg[ae_type]
                        n_layers, ratio_label = parse_config(cfg)
                        n_units = max(2, round(input_dim * BOTTLENECK_RATIOS.get(ratio_label, 0.5)))
                        cfg_label_ae = f"h{n_layers}-{ratio_label}"
                        try:
                            DF_maj, DF_tst = train_and_extract(
                                ae_type, X_maj_s, X_tst_s, n_layers, n_units)
                            sc_df = MinMaxScaler()
                            ae_features[ae_type] = (
                                sc_df.fit_transform(DF_maj),
                                sc_df.transform(DF_tst),
                                cfg_label_ae,
                            )
                        except Exception as e:
                            print(f"  [WARN] {ae_type} {cfg_label_ae} fold{fold_idx}: {e}")

                    if len(ae_features) < 2:
                        print(f"  [SKIP] {occ_type} fold{fold_idx}: 只有 {len(ae_features)} 個 AE 成功")
                        continue

                    # OCC scores
                    raw_scores = {}
                    for ae_type, (DF_maj_s, DF_tst_s, _) in ae_features.items():
                        sm, st = train_occ_scores(occ_type, DF_maj_s, DF_tst_s, n_nb_cap)
                        raw_scores[ae_type] = (sm, st)

                    norm_scores = {
                        ae: normalize_by_majority(sm, st)
                        for ae, (sm, st) in raw_scores.items()
                    }

                    # B_ Baseline：各 AE 單獨結果（供直接比較）
                    # 同時保留各 AE 的 binary y_pred 供後續 hard voting 使用
                    baseline_ypred = {}   # ae_type → y_pred (np.ndarray)
                    for ae_type, (sm, st) in raw_scores.items():
                        thr    = np.percentile(sm, 90)
                        y_pred = (st >= thr).astype(int)
                        m = metrics_from(y_tst, y_pred, st)
                        baseline_ypred[ae_type] = y_pred
                        ae_cfg = ae_features[ae_type][2]
                        all_records.append({
                            "Dataset": ds_name, "AE": ae_type, "OCC": occ_type,
                            "Config": ae_cfg, "Fold": fold_idx,
                            "Method": f"B_{ae_type}", **m,
                        })

                    # ── Hetero Vote（全部可用 AE）──────────────────────────
                    avail_aes = [ae for ae in AE_TYPES if ae in norm_scores]

                    # Soft：eq + istd
                    for scheme in ("eq", "istd"):
                        sm_list = [norm_scores[ae][0] for ae in avail_aes]
                        st_list = [norm_scores[ae][1] for ae in avail_aes]
                        m = weighted_soft_vote(sm_list, st_list, scheme, y_tst)
                        all_records.append({
                            "Dataset": ds_name, "AE": "+".join(avail_aes),
                            "OCC": occ_type, "Config": report_cfg,
                            "Fold": fold_idx, "Method": f"Hetero_Vote_{scheme}", **m,
                        })

                    # Hard：or（≥1 票）+ maj（多數決，n=4 時 ≥2 票）
                    y_pred_list = [baseline_ypred[ae] for ae in avail_aes]
                    for rule in ("or", "maj"):
                        m = hard_vote(y_pred_list, y_tst, rule)
                        all_records.append({
                            "Dataset": ds_name, "AE": "+".join(avail_aes),
                            "OCC": occ_type, "Config": report_cfg,
                            "Fold": fold_idx,
                            "Method": f"Hetero_Vote_hard_{rule}", **m,
                        })

                    # ── Hetero Pair（C(4,2) = 6 種）─────────────────────────
                    for ae_i, ae_j in AE_PAIRS:
                        if ae_i not in norm_scores or ae_j not in norm_scores:
                            continue
                        # Soft：eq + istd
                        for scheme in ("eq", "istd"):
                            sm_list = [norm_scores[ae_i][0], norm_scores[ae_j][0]]
                            st_list = [norm_scores[ae_i][1], norm_scores[ae_j][1]]
                            m = weighted_soft_vote(sm_list, st_list, scheme, y_tst)
                            all_records.append({
                                "Dataset": ds_name, "AE": f"{ae_i}+{ae_j}",
                                "OCC": occ_type, "Config": report_cfg,
                                "Fold": fold_idx,
                                "Method": f"Hetero_Pair_{ae_i}_{ae_j}_{scheme}", **m,
                            })
                        # Hard：or（≥1 票）+ and（≥2 票，2-AE pair 等於全票一致）
                        y_pred_pair = [baseline_ypred[ae_i], baseline_ypred[ae_j]]
                        for rule in ("or", "and"):
                            m = hard_vote(y_pred_pair, y_tst, rule)
                            all_records.append({
                                "Dataset": ds_name, "AE": f"{ae_i}+{ae_j}",
                                "OCC": occ_type, "Config": report_cfg,
                                "Fold": fold_idx,
                                "Method": f"Hetero_Pair_{ae_i}_{ae_j}_hard_{rule}", **m,
                            })

            else:
                # ── USE_BEST_CONFIG=False：全 grid search，所有 AE 同一 config ──
                all_cfgs = sorted([
                    f"h{n}-{r}" for n in N_LAYERS_LIST for r in BOTTLENECK_RATIOS
                ])
                for cfg in all_cfgs:
                    n_layers, ratio_label = parse_config(cfg)
                    n_units = max(2, round(input_dim * BOTTLENECK_RATIOS.get(ratio_label, 0.5)))
                    cfg_label = f"h{n_layers}-{ratio_label}"

                    for fold_idx, (tra_path, tst_path) in enumerate(fold_pairs, 1):
                        try:
                            X_tra, y_tra = load_dat(str(tra_path))
                            X_tst, y_tst = load_dat(str(tst_path))
                        except: continue
                        if y_tst.sum() < 1: continue

                        sc_orig = MinMaxScaler()
                        X_tra_s = sc_orig.fit_transform(X_tra)
                        X_tst_s = sc_orig.transform(X_tst)
                        X_maj_s = X_tra_s[y_tra == 0]
                        if len(X_maj_s) < 5: continue
                        n_nb_cap = min(20, len(X_maj_s) - 1)

                        ae_features = {}
                        for ae_type in AE_TYPES:
                            try:
                                DF_maj, DF_tst = train_and_extract(
                                    ae_type, X_maj_s, X_tst_s, n_layers, n_units)
                                sc_df = MinMaxScaler()
                                ae_features[ae_type] = (
                                    sc_df.fit_transform(DF_maj),
                                    sc_df.transform(DF_tst),
                                    cfg_label,
                                )
                            except Exception as e:
                                print(f"  [WARN] {ae_type} {cfg_label} fold{fold_idx}: {e}")

                        if len(ae_features) < 2:
                            continue

                        raw_scores = {
                            ae: train_occ_scores(occ_type, f[0], f[1], n_nb_cap)
                            for ae, f in ae_features.items()
                        }
                        norm_scores = {
                            ae: normalize_by_majority(sm, st)
                            for ae, (sm, st) in raw_scores.items()
                        }

                        # B_ Baseline + 保留 binary y_pred 給 hard voting 用
                        baseline_ypred = {}
                        for ae_type, (sm, st) in raw_scores.items():
                            thr    = np.percentile(sm, 90)
                            y_pred = (st >= thr).astype(int)
                            m = metrics_from(y_tst, y_pred, st)
                            baseline_ypred[ae_type] = y_pred
                            all_records.append({
                                "Dataset": ds_name, "AE": ae_type, "OCC": occ_type,
                                "Config": cfg_label, "Fold": fold_idx,
                                "Method": f"B_{ae_type}", **m,
                            })

                        # Hetero Vote：soft (eq/istd) + hard (or/maj)
                        avail_aes = [ae for ae in AE_TYPES if ae in norm_scores]
                        for scheme in ("eq", "istd"):
                            sm_list = [norm_scores[ae][0] for ae in avail_aes]
                            st_list = [norm_scores[ae][1] for ae in avail_aes]
                            m = weighted_soft_vote(sm_list, st_list, scheme, y_tst)
                            all_records.append({
                                "Dataset": ds_name, "AE": "+".join(avail_aes),
                                "OCC": occ_type, "Config": cfg_label,
                                "Fold": fold_idx, "Method": f"Hetero_Vote_{scheme}", **m,
                            })
                        y_pred_list = [baseline_ypred[ae] for ae in avail_aes]
                        for rule in ("or", "maj"):
                            m = hard_vote(y_pred_list, y_tst, rule)
                            all_records.append({
                                "Dataset": ds_name, "AE": "+".join(avail_aes),
                                "OCC": occ_type, "Config": cfg_label,
                                "Fold": fold_idx,
                                "Method": f"Hetero_Vote_hard_{rule}", **m,
                            })

                        # Hetero Pair：soft (eq/istd) + hard (or/and)
                        for ae_i, ae_j in AE_PAIRS:
                            if ae_i not in norm_scores or ae_j not in norm_scores:
                                continue
                            for scheme in ("eq", "istd"):
                                sm_list = [norm_scores[ae_i][0], norm_scores[ae_j][0]]
                                st_list = [norm_scores[ae_i][1], norm_scores[ae_j][1]]
                                m = weighted_soft_vote(sm_list, st_list, scheme, y_tst)
                                all_records.append({
                                    "Dataset": ds_name, "AE": f"{ae_i}+{ae_j}",
                                    "OCC": occ_type, "Config": cfg_label,
                                    "Fold": fold_idx,
                                    "Method": f"Hetero_Pair_{ae_i}_{ae_j}_{scheme}", **m,
                                })
                            y_pred_pair = [baseline_ypred[ae_i], baseline_ypred[ae_j]]
                            for rule in ("or", "and"):
                                m = hard_vote(y_pred_pair, y_tst, rule)
                                all_records.append({
                                    "Dataset": ds_name, "AE": f"{ae_i}+{ae_j}",
                                    "OCC": occ_type, "Config": cfg_label,
                                    "Fold": fold_idx,
                                    "Method": f"Hetero_Pair_{ae_i}_{ae_j}_hard_{rule}", **m,
                                })
    return pd.DataFrame(all_records)


# ─────────────────────────── Excel 輸出 ──────────────────────────────
def compute_summary(df):
    result = (df.groupby(["Dataset","Method","AE","OCC","Config"])[METRIC_COLS]
                .agg(["mean","std"]).round(4))
    result.columns = [f"{m}_{s}" for m, s in result.columns]
    return result.reset_index()

def compute_overall(df):
    result = (df.groupby(["Method","AE","OCC","Config"])[METRIC_COLS]
                .agg(["mean","std"]).round(4))
    result.columns = [f"{m}_{s}" for m, s in result.columns]
    return result.reset_index()

def save_excel(df_all):
    """
    將結果寫入 Excel，格式與 D/E 一致。

    best_per_fold 的 config 選擇邏輯（動態計算，非 USE_BEST_CONFIG 的靜態讀取）：
      per (Dataset, Method, AE, OCC)，計算各 Config 在所有 fold 的平均 AUC，
      取平均 AUC 最高的 Config，保留該 Config 的所有 fold 記錄。
      → 與 B/C/D/E baseline_grid 的 best_overall 選法完全一致。

    注意：USE_BEST_CONFIG 控制的是「每個 AE 用哪個 config 去跑實驗」；
    best_per_fold 這裡是「從跑完的結果中，哪個 config 最好」，兩者是不同層次。
    """
    wb = Workbook()
    wb.remove(wb.active)

    def write_df(ws, df, title=""):
        ws.title = title if title else ws.title
        df_reset = df.reset_index() if isinstance(df.index, pd.MultiIndex) else df
        headers = list(df_reset.columns)
        for j, h in enumerate(headers, 1):
            ws.cell(1, j, h).font = Font(bold=True)
        for i, row in enumerate(df_reset.itertuples(index=False), 2):
            for j, v in enumerate(row, 1):
                ws.cell(i, j, round(v, 4) if isinstance(v, float) else v)

    ws1 = wb.create_sheet("all_per_fold")
    write_df(ws1, df_all)

    ws2 = wb.create_sheet("all_summary")
    write_df(ws2, compute_summary(df_all))

    ws3 = wb.create_sheet("all_overall")
    write_df(ws3, compute_overall(df_all))

    # ── best_per_fold：動態選最高 AUC config ─────────────────────────
    # per (Dataset, Method, AE, OCC)，取 fold 平均 AUC 最高的 Config，
    # 保留該 Config 的所有 fold 記錄。與 B/C/D/E 的 best_overall 邏輯一致。
    best_cfg = (
        df_all.groupby(["Dataset", "Method", "AE", "OCC", "Config"])["AUC"]
              .mean()
              .reset_index()
              .sort_values("AUC", ascending=False)
              .drop_duplicates(["Dataset", "Method", "AE", "OCC"])
    )
    df_best_filtered = df_all.merge(
        best_cfg[["Dataset", "Method", "AE", "OCC", "Config"]],
        on=["Dataset", "Method", "AE", "OCC", "Config"],
    )

    ws4 = wb.create_sheet("best_per_fold")
    write_df(ws4, df_best_filtered)

    ws5 = wb.create_sheet("best_summary")
    write_df(ws5, compute_summary(df_best_filtered))

    ws6 = wb.create_sheet("best_overall")
    overall = compute_overall(df_best_filtered)
    # Format as "mean ± std" strings like D/E
    def fmt_overall(df_ov):
        rows = []
        for _, vals in df_ov.iterrows():
            row = {"Method": vals["Method"], "AE": vals["AE"],
                   "OCC": vals["OCC"], "Config": vals["Config"]}
            for col in METRIC_COLS:
                m = vals[f"{col}_mean"]; s = vals[f"{col}_std"]
                row[col] = f"{m:.4f} ± {s:.4f}"
            rows.append(row)
        return pd.DataFrame(rows)
    write_df(ws6, fmt_overall(overall))

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果儲存至：{OUTPUT_FILE}")


# ─────────────────────────── Entry Point ─────────────────────────────
if __name__ == "__main__":
    # Method 數量計算
    n_vote_soft = 2                       # Hetero_Vote_eq + Hetero_Vote_istd
    n_vote_hard = 2                       # Hetero_Vote_hard_or + Hetero_Vote_hard_maj
    n_pair_soft = len(AE_PAIRS) * 2       # 6 pairs × 2 = 12
    n_pair_hard = len(AE_PAIRS) * 2       # 6 pairs × 2 (or, and) = 12
    n_base      = len(AE_TYPES)           # B_AE + B_DAE + B_SAE + B_VAE = 4
    n_total     = n_vote_soft + n_vote_hard + n_pair_soft + n_pair_hard + n_base

    print("=" * 72)
    print("F_hetero_ensemble_grid.py — 異質 AE Ensemble")
    print(f"AE 類型    ：{AE_TYPES}")
    print(f"OCC 方法   ：{OCC_TYPES}")
    print(f"Best config from B grid：{USE_BEST_CONFIG}")
    print(f"Method 數量：Baseline({n_base}) + Vote({n_vote_soft+n_vote_hard}) "
          f"+ Pair({n_pair_soft+n_pair_hard}) = {n_total} 種")
    print(f"  Baseline    : {[f'B_{ae}' for ae in AE_TYPES]}")
    print(f"  Vote (soft) : [Hetero_Vote_eq, Hetero_Vote_istd]")
    print(f"  Vote (hard) : [Hetero_Vote_hard_or (≥1), Hetero_Vote_hard_maj (≥2)]")
    print(f"  Pair (soft) : {len(AE_PAIRS)} pairs × [eq, istd]")
    print(f"  Pair (hard) : {len(AE_PAIRS)} pairs × [hard_or (≥1), hard_and (≥2)]")
    print("=" * 72)

    df_all = run_experiment()

    if df_all.empty:
        print("\n⚠️  沒有結果，請確認資料路徑與格式。")
    else:
        save_excel(df_all)

        print("\n── Best Overall（Hetero Vote/Pairs，跨所有 dataset 平均）──")
        hetero = df_all[df_all["Method"].str.startswith("Hetero_")]
        if not hetero.empty:
            print(hetero.groupby(["Method", "OCC"])["AUC"]
                        .mean().sort_values(ascending=False).round(4).to_string())

        print("\n── Baseline 對照（各 AE 單獨，跨所有 dataset 平均）──")
        base = df_all[df_all["Method"].str.startswith("B_")]
        if not base.empty:
            print(base.groupby(["Method", "OCC"])["AUC"]
                      .mean().sort_values(ascending=False).round(4).to_string())
