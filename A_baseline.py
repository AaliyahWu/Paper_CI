"""
baseline A_scaled.py
====================
【修正版】在原始特徵輸入 OCC 前加上 MinMaxScaler，
與 Baseline B/C 的前處理保持一致，確保公平比較。

修正內容：
  - run_experiment() 中，在取出 X_train_maj 後：
      scaler = MinMaxScaler()
      X_train_maj = scaler.fit_transform(X_train_maj)   # fit on train majority
      X_tst_scaled = scaler.transform(X_tst)            # transform test
  - evaluate() 接收已 scaled 的資料，邏輯不變

輸出：results/A baseline.xlsx
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import (
    roc_auc_score, f1_score, recall_score, confusion_matrix
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT = Path("data")          # 修改為你的資料根目錄
OUTPUT_DIR = Path("results")
OUTPUT_FILE = OUTPUT_DIR / "A baseline.xlsx"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5

# ─────────────────────────── OCC 模型 ────────────────────────────────────────
MODELS = {
    "OCSVM":           OneClassSVM(nu=0.1, kernel="rbf"),
    "LOF":             LocalOutlierFactor(novelty=True, contamination=0.1),
    "IsolationForest": IsolationForest(contamination=0.1, random_state=42),
}

# ─────────────────────────── KEEL .dat 解析 ──────────────────────────────────
def parse_keel_dat(filepath):
    """解析 KEEL .dat 檔案，回傳 (X: ndarray, y: ndarray, class_values: list)"""
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    
    data_start = False
    class_values = []
    rows = []

    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("%"):
            continue

        low = stripped.lower()

        if low.startswith("@outputs") or low.startswith("@output"):
            # 最後一欄為 class，不需特別處理
            pass
        elif low.startswith("@attribute"):
            # 抓 class 的合法值，例如 @attribute Class {positive, negative}
            match = re.search(r'\{(.+)\}', stripped)
            if match and ("class" in low or "Class" in stripped):
                class_values = [v.strip() for v in match.group(1).split(",")]
        elif low == "@data":
            data_start = True
            continue

        if data_start and stripped:
            rows.append(stripped)

    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = []
    for row in rows:
        parts = [p.strip() for p in row.split(",")]
        records.append(parts)

    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    labels_raw = df[label_col].values

    # 數值欄位；類別欄位（無法轉數值）改用 label encoding
    feature_df = df.iloc[:, :-1].copy()
    for col in feature_df.columns:
        converted = pd.to_numeric(feature_df[col], errors="coerce")
        if converted.isna().all():
            # 純類別欄位：轉為整數編碼
            feature_df[col] = pd.Categorical(feature_df[col]).codes.astype(float)
        else:
            feature_df[col] = converted
    X = feature_df.values.astype(float)

    return X, labels_raw, class_values


def encode_labels(labels_raw, class_values):
    """
    minority（少數類）= anomaly = 1（正類，評估目標）
    majority（多數類）= normal  = 0（負類，OCC 訓練對象）
    """
    unique, counts = np.unique(labels_raw, return_counts=True)
    minority_class = unique[np.argmin(counts)]

    y = np.where(labels_raw == minority_class, 1, 0).astype(int)
    return y, minority_class


# ─────────────────────────── 評估指標 ────────────────────────────────────────
def gmean_score(y_true, y_pred):
    cm = confusion_matrix(y_true, y_pred, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return np.sqrt(sens * spec)
    return 0.0


def evaluate(model, X_train_maj, X_test, y_test):
    """
    訓練 OCC（只用 majority/normal 樣本），在 test set 評估。
    y_test : 1=minority(anomaly), 0=majority(normal)

    Threshold 設定（訓練集 majority 分數第 90 百分位數）：
      predict() 的內建 threshold 在某些 fold 會導致 test set 中
      預測為 anomaly 的數量為 0，使 AUC > 0.6 但 F1=0。
      改用訓練集 majority anomaly score 的第 90 百分位數作為 threshold，
      對應 contamination=0.1 語義（約 10% 的 normal 樣本超過此門檻），
      直接套用於 test scores，跨 fold 穩定一致。
    """
    model.fit(X_train_maj)

    # 高分 = 越像 anomaly
    scores_train = -model.decision_function(X_train_maj)
    scores_test  = -model.decision_function(X_test)

    # threshold：training majority 分數的第 90 百分位數
    threshold = np.percentile(scores_train, 90)
    y_pred = (scores_test >= threshold).astype(int)

    try:
        auc = roc_auc_score(y_test, scores_test)
    except Exception:
        auc = float("nan")

    f1  = f1_score(y_test, y_pred, pos_label=1, zero_division=0)
    rec = recall_score(y_test, y_pred, pos_label=1, zero_division=0)
    gm  = gmean_score(y_test, y_pred)

    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment():
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    all_records = []   # per_fold 資料

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            # 實際檔名格式：{ds_name去掉-fold及其後綴}-{fold}tra.dat
            # 例：abalone19-5-fold/ → abalone19-5-{fold}tra.dat
            # 例：abalone19-5-fold_Encoding/ → abalone19-5-{fold}tra.dat
            file_prefix = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]

            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)

            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案")
                continue

            try:
                X_tra, y_tra_raw, cv = parse_keel_dat(tra_file)
                X_tst, y_tst_raw, _  = parse_keel_dat(tst_file)

                y_tra, min_cls = encode_labels(y_tra_raw, cv)
                y_tst, _       = encode_labels(y_tst_raw, cv)

                X_train_maj = X_tra[y_tra == 0]

                if len(X_train_maj) == 0:
                    print(f"  [SKIP] Fold {fold}: 訓練集無多數類樣本")
                    continue

                # ── MinMaxScale（與 Baseline B/C 保持一致）──────────────────
                # fit 只用訓練集多數類，transform 同時套用到 test set
                scaler = MinMaxScaler()
                X_train_maj = scaler.fit_transform(X_train_maj)
                X_tst_scaled = scaler.transform(X_tst)

            except Exception as e:
                print(f"  [ERROR] Fold {fold} 解析失敗: {e}")
                continue

            for model_name, model_proto in MODELS.items():
                import copy
                model = copy.deepcopy(model_proto)
                try:
                    metrics = evaluate(model, X_train_maj, X_tst_scaled, y_tst)
                except Exception as e:
                    print(f"  [ERROR] {model_name} Fold {fold}: {e}")
                    metrics = {"AUC": float("nan"), "F1": float("nan"),
                               "Recall": float("nan"), "G-mean": float("nan")}

                row = {
                    "Dataset": ds_name,
                    "Model":   model_name,
                    "Fold":    fold,
                    **metrics
                }
                all_records.append(row)
                print(f"  {model_name:18s} Fold{fold}  "
                      f"AUC={metrics['AUC']:.4f}  F1={metrics['F1']:.4f}  "
                      f"Recall={metrics['Recall']:.4f}  G-mean={metrics['G-mean']:.4f}")

    return pd.DataFrame(all_records)


# ─────────────────────────── Excel 輸出 ──────────────────────────────────────
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

# 顏色定義
HEADER_FILL   = PatternFill("solid", fgColor="2F5597")   # 深藍
SUBHDR_FILL   = PatternFill("solid", fgColor="BDD7EE")   # 淡藍
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")   # 超淡藍
OVERALL_FILL  = PatternFill("solid", fgColor="FFD700")   # 金黃
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT   = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT     = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center")
RIGHT_ALIGN   = Alignment(horizontal="right",  vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
MED_BORDER = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium")
)


def style_cell(cell, value, font=None, fill=None, align=None, border=None, fmt=None):
    cell.value = value
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fmt:    cell.number_format = fmt


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def write_per_fold(ws, df):
    """分頁 per_fold：每一行對應一個 (Dataset, Model, Fold) 的四個指標"""
    ws.title = "per_fold"

    headers = ["Dataset", "Model", "Fold"] + METRIC_COLS
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, col in enumerate(headers, 1):
            val = row[col]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = BODY_FONT
            cell.alignment = CENTER_ALIGN if c_idx >= 3 else LEFT_ALIGN
            cell.border    = THIN_BORDER
            if fill: cell.fill = fill
            if col in METRIC_COLS:
                cell.number_format = "0.0000"

    # 欄寬
    widths = [28, 16, 8, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, get_column_letter(i), w)

    ws.freeze_panes = "A2"


def write_summary(ws, df):
    """
    分頁 summary：
    列：Dataset
    欄：Model × Metric（mean ± std over 5 folds）
    """
    ws.title = "summary"

    model_names = list(MODELS.keys())

    # 計算 mean & std
    grouped = df.groupby(["Dataset", "Model"])[METRIC_COLS]
    stats = grouped.agg(["mean", "std"]).reset_index()

    # Header row 1：Dataset | [Model × Metric columns]
    col = 1
    ws.cell(row=1, column=col, value="Dataset").font = HEADER_FONT
    ws.cell(row=1, column=col).fill = HEADER_FILL
    ws.cell(row=1, column=col).alignment = CENTER_ALIGN
    ws.cell(row=1, column=col).border = THIN_BORDER
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    for mname in model_names:
        span = len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        cell = ws.cell(row=1, column=col, value=mname)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        for mc, metric in enumerate(METRIC_COLS):
            c = ws.cell(row=2, column=col + mc, value=metric)
            c.font      = SUBHDR_FONT
            c.fill      = SUBHDR_FILL
            c.alignment = CENTER_ALIGN
            c.border    = THIN_BORDER
        col += span

    # Data rows
    datasets = df["Dataset"].unique()
    for r_idx, ds in enumerate(datasets, 3):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        cell = ws.cell(row=r_idx, column=1, value=ds)
        cell.font      = BODY_FONT
        cell.alignment = LEFT_ALIGN
        cell.border    = THIN_BORDER
        if fill: cell.fill = fill

        col = 2
        for mname in model_names:
            sub = stats[(stats["Dataset"] == ds) & (stats["Model"] == mname)]
            for metric in METRIC_COLS:
                try:
                    mean_val = sub[(metric, "mean")].values[0]
                    std_val  = sub[(metric, "std")].values[0]
                    display  = f"{mean_val:.4f} ± {std_val:.4f}"
                except Exception:
                    display = "N/A"
                c = ws.cell(row=r_idx, column=col, value=display)
                c.font      = BODY_FONT
                c.alignment = CENTER_ALIGN
                c.border    = THIN_BORDER
                if fill: c.fill = fill
                col += 1

    # 欄寬
    set_col_width(ws, "A", 28)
    for i in range(2, 2 + len(model_names) * len(METRIC_COLS)):
        set_col_width(ws, get_column_letter(i), 18)

    ws.freeze_panes = "B3"


def write_overall_mean(ws, df):
    """
    分頁 overall_mean：
    所有 Dataset、所有 Fold 的總平均（per Model × Metric）
    """
    ws.title = "overall_mean"

    model_names = list(MODELS.keys())
    overall = df.groupby("Model")[METRIC_COLS].agg(["mean", "std"])

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Overall Mean (across all datasets & folds)"
    title_cell.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    title_cell.alignment = CENTER_ALIGN

    # Header
    headers = ["Model"] + METRIC_COLS
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER

    for r_idx, mname in enumerate(model_names, 3):
        fill = OVERALL_FILL if r_idx == 3 else (ALT_FILL if r_idx % 2 == 0 else None)
        ws.cell(row=r_idx, column=1, value=mname).font = BOLD_FONT
        ws.cell(row=r_idx, column=1).alignment = CENTER_ALIGN
        ws.cell(row=r_idx, column=1).border = THIN_BORDER
        if fill: ws.cell(row=r_idx, column=1).fill = fill

        for c_idx, metric in enumerate(METRIC_COLS, 2):
            try:
                mean_val = overall.loc[mname, (metric, "mean")]
                std_val  = overall.loc[mname, (metric, "std")]
                display  = f"{mean_val:.4f} ± {std_val:.4f}"
            except Exception:
                display = "N/A"
            c = ws.cell(row=r_idx, column=c_idx, value=display)
            c.font      = BODY_FONT
            c.alignment = CENTER_ALIGN
            c.border    = THIN_BORDER
            if fill: c.fill = fill

    # 欄寬
    set_col_width(ws, "A", 20)
    for i in range(2, 2 + len(METRIC_COLS)):
        set_col_width(ws, get_column_letter(i), 22)

    ws.freeze_panes = "A3"


def save_excel(df):
    wb = Workbook()

    ws_fold    = wb.active
    ws_summary = wb.create_sheet()
    ws_overall = wb.create_sheet()

    write_per_fold(ws_fold, df)
    write_summary(ws_summary, df)
    write_overall_mean(ws_overall, df)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("Baseline A：OCC 三模型評估（OCSVM / LOF / Isolation Forest）")
    print("=" * 60)

    df_results = run_experiment()

    if df_results.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_results)
        print("\n── Summary（所有資料集平均）──")
        print(df_results.groupby("Model")[METRIC_COLS].mean().round(4).to_string())
