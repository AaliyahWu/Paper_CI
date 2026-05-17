"""
Ev3_ensemble_pairs_grid.py
==========================
Ensemble 方法二：pairs of OCCs by weighted voting（對應投影片 "b. pairs of OCCs by weighted voting"）

修正說明（相對於 Ev2）：
  [修正 1] 加入真正的 weighted voting
           Ev2 只做了等權重（soft = 平均），缺少 "weighted" 核心。
           本版新增 inverse-std weighting (_istd)：
             w_i ∝ 1 / std(s_i_train_maj_normalized)
           代表「訓練集分數越穩定的 OCC，分配越高的權重」。

  [修正 2] 修正 F1/Recall/G-mean 的 Bug
           Ev2 中 soft_metrics 的 binary prediction 誤用了 hard 版本的 y_pred，
           導致兩者的 F1/Recall/G-mean 完全相同，只有 AUC 不同。
           本版每個 variant 都用自己的連續分數取 90th percentile 當 threshold，
           再轉成各自獨立的 binary predictions，確保四個指標都正確。

  [修正 3] 加回 hard voting（兩種 tie-break 規則並列）
           Ev3 原本完全移除 hard voting；本版補回，提供完整的方法比較：
             *_hard_or  : ≥1 票判 anomaly（任何一個 OCC 偵測到就視為異常）
                          → recall 高，precision 容易偏低
             *_hard_and : ≥2 票判 anomaly（兩個 OCC 都同意才視為異常）
                          → precision 高，recall 容易偏低
           兩個 variant 共用同一個連續 vote count（0/1/2）作為 AUC ranking 分數，
           因此 *_hard_or 和 *_hard_and 的 AUC 會相同；F1/Recall/G-mean 不同。

核心概念：同一個 OCC 方法（OCSVM / LOF / iForest）分別跑在三種特徵空間上：
  OCC1 = OCC on OF_maj           （對應 Baseline A）
  OCC2 = OCC on DF_maj           （對應 Baseline B）
  OCC3 = OCC on OF_maj + DF_maj  （對應 Baseline C）

三種 pair（每對各有 4 個 method：eq / istd / hard_or / hard_and，共 12 組 ensemble）：
  Pair_AB = OCC1 ⊕ OCC2
  Pair_AC = OCC1 ⊕ OCC3
  Pair_BC = OCC2 ⊕ OCC3

Weighted soft voting 流程（_eq / _istd 共用）：
  1. 各 OCC 的 anomaly score 先用 train-maj min-max normalize → 大致 [0,1]
     （不使用 test set 統計，避免 data leak）
  2. 計算權重：
       _eq   → w_i = w_j = 0.5
       _istd → w_i ∝ 1/std(s_i_train_maj_normalized)，再歸一化 w_i + w_j = 1
               std 以 EPS 保護，避免 zero-div
  3. 加權後的連續分數：
       s_pair_maj = w_i * s_i_maj_n + w_j * s_j_maj_n
       s_pair_tst = w_i * s_i_tst_n + w_j * s_j_tst_n
  4. Threshold = train-maj s_pair 的第 90 百分位數（與 A/B/C 一致）
  5. y_pred   = (s_pair_tst >= threshold).astype(int)
  6. AUC 用連續 s_pair_tst；F1 / Recall / G-mean 用 y_pred（各 variant 各自計算）

Hard voting 流程（_hard_or / _hard_and 共用）：
  1. 各 OCC 用自己的 train-maj 第 90 百分位數獨立轉成 0/1（即 baseline A/B/C 的 y_pred）
  2. vote_count = y_i + y_j ∈ {0, 1, 2}
  3. _hard_or  → y_pred = (vote_count >= 1)  # OR：任一票
     _hard_and → y_pred = (vote_count >= 2)  # AND：全票
  4. AUC 用 vote_count（0~2）當連續 ranking 分數
     → _hard_or 和 _hard_and 的 AUC 必然相同（同一連續分數）
     → F1 / Recall / G-mean 因二元預測不同而有差異

每筆 ensemble 記錄同時附帶 A / B / C 三條 baseline 在同一
(AE, config, fold, OCC) 下的結果，方便直接在 xlsx 對比。

Grid 搜尋空間（對齊 C_baseline_grid 的完整空間）：
  n_layers   : [1, 2, 3]
  bottleneck : ["1/4", "1/3", "1/2", "1/1", "2/1", "3/1", "4/1"]
  → 共 21 種 AE 架構 per (AE_type, OCC, fold)

與 B/C grid 的一致性：
  - AE 模型定義、train_and_extract 完全相同
  - OCC 超參數（nu=0.1、contamination=0.1）完全相同
  - MinMaxScaler fit 只用 training majority、threshold 用 train-maj 第 90 百分位
  - 隨機種子 torch.manual_seed(42) + np.random.seed(42)

輸出：results/E ensemble pairs grid.xlsx
  分頁：
    all_per_fold      所有 (AE, OCC, Config, Fold, Method) 的原始結果
    all_summary       mean ± std across folds（per Dataset × Method × AE × OCC × Config）
    all_overall       全域平均（per Method × AE × OCC × Config）
    best_per_fold     每 (Dataset, AE, OCC, Method) 中『5-fold 平均 AUC 最高』Config 的 fold 資料 (per-dataset)
    best_summary      best mean ± std across folds
    best_overall      best 全域平均
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT   = Path("data")
RESULTS_DIR = Path("results")
OUTPUT_FILE = RESULTS_DIR / "E ensemble pairs grid.xlsx"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5

AE_EPOCHS     = 100
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1
SAE_SPARSITY  = 1e-3
VAE_BETA      = 1.0

AE_TYPES    = ["AE", "DAE", "SAE", "VAE"]
OCC_TYPES   = ["OCSVM", "LOF", "iForest"]
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

N_LAYERS_LIST     = [1, 2, 3]
BOTTLENECK_RATIOS = {
    "1/4": 0.25, "1/3": 1/3, "1/2": 0.5, "1/1": 1.0,
    "2/1": 2.0,  "3/1": 3.0, "4/1": 4.0,
}
ALL_CONFIGS = [f"h{nl}-{rl}" for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]

EPS = 1e-12  # 防止 inverse-std 除 0

# Methods：先 baseline reference，再 12 組 pair ensemble（3 pair × 4 voting scheme）
METHODS = [
    "A", "B", "C",
    "Pair_AB_eq",      "Pair_AB_istd",
    "Pair_AB_hard_or", "Pair_AB_hard_and",
    "Pair_AC_eq",      "Pair_AC_istd",
    "Pair_AC_hard_or", "Pair_AC_hard_and",
    "Pair_BC_eq",      "Pair_BC_istd",
    "Pair_BC_hard_or", "Pair_BC_hard_and",
]

# 三對的組成：(pair_name, idx_i, idx_j)
# idx 對應 (0=OF/A, 1=DF/B, 2=COMB/C)
PAIRS = [("AB", 0, 1), ("AC", 0, 2), ("BC", 1, 2)]


# ─────────────────────────── AE 模型定義 ─────────────────────────────────────
class AEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        for i in range(len(dims) - 1):
            d_in, d_out = dims[-(i+1)], dims[-(i+2)]
            act = nn.Sigmoid() if i == len(dims) - 2 else nn.ReLU()
            dec += [nn.Linear(d_in, d_out), act]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


class VAEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        base = []
        for i in range(len(dims) - 2):
            base += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        self.enc_base  = nn.Sequential(*base) if base else nn.Identity()
        mid = dims[-2] if len(dims) >= 2 else input_dim
        self.fc_mu     = nn.Linear(mid, dims[-1])
        self.fc_logvar = nn.Linear(mid, dims[-1])
        dec_dims = dims[::-1]
        dec = []
        for i in range(len(dec_dims) - 1):
            act = nn.Sigmoid() if i == len(dec_dims) - 2 else nn.ReLU()
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), act]
        self.decoder = nn.Sequential(*dec)

    def reparameterize(self, mu, lv):
        return mu + torch.exp(0.5 * lv) * torch.randn_like(lv)

    def forward(self, x):
        h  = self.enc_base(x)
        mu = self.fc_mu(h)
        lv = self.fc_logvar(h)
        z  = self.reparameterize(mu, lv)
        return self.decoder(z), z, mu, lv


# ─────────────────────────── AE 訓練 + 特徵提取 ──────────────────────────────
def train_and_extract(ae_type, X_maj_s, X_test_s, n_layers, n_units):
    """與 B/C grid 完全相同；回傳 (feat_maj, feat_test)"""
    input_dim = X_maj_s.shape[1]
    model = VAEModel(input_dim, n_layers, n_units) if ae_type == "VAE" \
            else AEModel(input_dim, n_layers, n_units)

    optim  = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))
    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim.zero_grad()
            if ae_type == "DAE":
                xb_n = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
                xr, _ = model(xb_n); loss = mse(xr, xb)
            elif ae_type == "SAE":
                xr, z = model(xb); loss = mse(xr, xb) + SAE_SPARSITY * z.abs().mean()
            elif ae_type == "VAE":
                xr, _, mu, lv = model(xb)
                kl = -0.5 * (1 + lv - mu.pow(2) - lv.exp()).mean()
                loss = mse(xr, xb) + VAE_BETA * kl
            else:
                xr, _ = model(xb); loss = mse(xr, xb)
            loss.backward(); optim.step()

    model.eval()
    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            if ae_type == "VAE":
                _, _, mu, _ = model(xt); return mu.numpy()
            else:
                _, z = model(xt); return z.numpy()

    return extract(X_maj_s), extract(X_test_s)


# ─────────────────────────── OCC 分數取得 ────────────────────────────────────
def train_occ_scores(occ_type, X_maj, X_test, n_neighbors_cap):
    """
    在 X_maj 上訓練 OCC，回傳 (scores_maj, scores_test)。
    scores 越大越像 anomaly（= -decision_function，與 A/B/C baseline 一致）。
    """
    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(X_maj)
        s_maj  = -clf.decision_function(X_maj)
        s_test = -clf.decision_function(X_test)
    elif occ_type == "LOF":
        k = min(20, n_neighbors_cap)
        clf = LocalOutlierFactor(n_neighbors=k, novelty=True, contamination=0.1)
        clf.fit(X_maj)
        s_maj  = -clf.decision_function(X_maj)
        s_test = -clf.decision_function(X_test)
    else:  # iForest
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(X_maj)
        s_maj  = -clf.decision_function(X_maj)
        s_test = -clf.decision_function(X_test)
    return np.asarray(s_maj, dtype=float), np.asarray(s_test, dtype=float)


def normalize_by_majority(scores_maj, scores_test):
    """
    以 training majority 分數的 min-max 做線性變換 → 大致 [0,1]。
    不使用 test 統計，避免 data leak。
    """
    lo, hi = float(scores_maj.min()), float(scores_maj.max())
    if hi - lo < EPS:
        return np.zeros_like(scores_maj), np.zeros_like(scores_test)
    return (scores_maj - lo) / (hi - lo), (scores_test - lo) / (hi - lo)


# ─────────────────────────── 評估指標 ────────────────────────────────────────
def gmean_score(y_true, y_pred_binary):
    cm = confusion_matrix(y_true, y_pred_binary, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return float(np.sqrt(sens * spec))
    return 0.0


def metrics_from(y_true, y_pred_binary, scores_test_for_auc):
    """AUC 用連續分數；F1/Recall/G-mean 用 binary pred（兩者均獨立計算）。"""
    try:
        auc = (roc_auc_score(y_true, scores_test_for_auc)
               if len(np.unique(y_true)) >= 2 else float("nan"))
    except Exception:
        auc = float("nan")
    f1  = f1_score(y_true, y_pred_binary, pos_label=1, zero_division=0)
    rec = recall_score(y_true, y_pred_binary, pos_label=1, zero_division=0)
    gm  = gmean_score(y_true, y_pred_binary)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── Weighted pair voting 核心 ───────────────────────
def weighted_pair(s_i_maj_n, s_i_tst_n, s_j_maj_n, s_j_tst_n,
                  y_test, scheme):
    """
    對一對已 normalize 到 [0,1] 的 OCC 分數做 weighted soft voting。

    scheme = "eq"   : equal weights (0.5, 0.5)
    scheme = "istd" : inverse-std weights — w_k ∝ 1/std(s_k_train_maj_normalized)，
                      再歸一化使 w_i + w_j = 1。
                      代表「訓練集分數越穩定（區分力越一致）的 OCC，權重越高」。

    流程（兩個 scheme 共用）：
      1. 計算權重
      2. s_pair_maj = w_i * s_i_maj_n + w_j * s_j_maj_n
         s_pair_tst = w_i * s_i_tst_n + w_j * s_j_tst_n
      3. threshold = 90th percentile of s_pair_maj（與 A/B/C 一致）
      4. y_pred    = (s_pair_tst >= threshold)
      5. AUC 用 s_pair_tst；F1/Recall/G-mean 用 y_pred

    回傳 metrics dict。
    """
    if scheme == "eq":
        w_i = w_j = 0.5
    elif scheme == "istd":
        std_i = float(np.std(s_i_maj_n)) + EPS
        std_j = float(np.std(s_j_maj_n)) + EPS
        inv_i, inv_j = 1.0 / std_i, 1.0 / std_j
        total = inv_i + inv_j
        w_i, w_j = inv_i / total, inv_j / total
    else:
        raise ValueError(f"未知的 scheme: {scheme}")

    s_pair_maj = w_i * s_i_maj_n + w_j * s_j_maj_n
    s_pair_tst = w_i * s_i_tst_n + w_j * s_j_tst_n

    threshold = np.percentile(s_pair_maj, 90)
    y_pred    = (s_pair_tst >= threshold).astype(int)

    return metrics_from(y_test, y_pred, s_pair_tst)


# ─────────────────────────── Hard pair voting 核心 ───────────────────────────
def hard_pair(y_i_pred, y_j_pred, y_test, rule):
    """
    對一對已產生 binary prediction 的 OCC 結果做 hard voting。

    參數：
      y_i_pred, y_j_pred : 各 OCC 在 test 上的二元預測（用各自 train-maj
                           90th percentile 作為閾值產生，與 baseline A/B/C 一致）
      y_test             : 真實標籤
      rule               : "or" → ≥1 票判 anomaly；"and" → ≥2 票判 anomaly

    流程：
      1. vote_count = y_i + y_j ∈ {0, 1, 2}
      2. _or  → y_pred = (vote_count >= 1)  ← 任一 OCC 偵測到就視為異常（recall 高）
         _and → y_pred = (vote_count >= 2)  ← 兩 OCC 都同意才視為異常（precision 高）
      3. AUC 用 vote_count（連續 0/1/2）作為 ranking 分數
         → "or" 與 "and" 的 AUC 必然相同（同一 ranking）
         → F1 / Recall / G-mean 因二元預測不同而不同

    注意：
      Ev3 docstring 提及「OR 嚴重 over-predict」，這是事實 — _hard_or 通常
      Recall 偏高、Precision 偏低，F1 不一定好；_hard_and 則相反。兩者並列
      可呈現完整 trade-off，由實驗結果評估哪種對特定資料集較合適。

    回傳 metrics dict。
    """
    vote_count = y_i_pred.astype(int) + y_j_pred.astype(int)  # 0 / 1 / 2

    if rule == "or":
        y_pred = (vote_count >= 1).astype(int)
    elif rule == "and":
        y_pred = (vote_count >= 2).astype(int)
    else:
        raise ValueError(f"未知的 rule: {rule}")

    # AUC 用 vote_count 當連續分數（與 D 的 Vote3_hard 一致）
    return metrics_from(y_test, y_pred, vote_count.astype(float))


# ─────────────────────────── KEEL .dat 解析 ──────────────────────────────────
def parse_keel_dat(filepath, minority_label=None):
    """
    解析 KEEL .dat。

    修正重點：
      - 類別特徵：不可轉數值時使用 pd.Categorical(...).codes，與 B/C/D/E 對齊。
      - label：train fold 先決定 minority_label；test fold 沿用 train 的 minority_label，
        避免 train/test 各自判定少數類造成 y 反轉。

    回傳：X, y, minority_label
      y = 1 表示 minority/anomaly；y = 0 表示 majority/normal。
    """
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"):
            continue
        if s.lower() == "@data":
            data_start = True
            continue
        if data_start:
            rows.append(s)
    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    y_raw = df[label_col].astype(str).str.strip().values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    if minority_label is None:
        unique, counts = np.unique(y_raw, return_counts=True)
        minority_label = unique[np.argmin(counts)]

    y = (y_raw == minority_label).astype(int)
    return X, y, minority_label

# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment():
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    param_configs = [(nl, rl) for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]
    all_records   = []

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n{'='*68}")
        print(f"▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            file_prefix  = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]
            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)

            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案"); continue

            try:
                X_tra, y_tra, minority_label = parse_keel_dat(tra_file)
                X_tst, y_tst, _ = parse_keel_dat(tst_file, minority_label=minority_label)
                input_dim = X_tra.shape[1]
                X_maj     = X_tra[y_tra == 0]

                if len(X_maj) < 5:
                    print(f"  [SKIP] Fold {fold}: 訓練集正常樣本不足 ({len(X_maj)})"); continue
                if y_tst.sum() == 0:
                    print(f"  [SKIP] Fold {fold}: 測試集無少數類樣本"); continue

                # OCC1 用的特徵空間（OF）MinMax，fit 只用 training majority
                scaler_of = MinMaxScaler()
                OF_maj    = scaler_of.fit_transform(X_maj)
                OF_tst    = scaler_of.transform(X_tst)
                n_nb_cap  = max(1, len(X_maj) - 1)

            except Exception as e:
                print(f"  [ERROR] Fold {fold} 資料載入失敗: {e}"); continue

            for ae_type in AE_TYPES:
                # 每個 (ae_type, config) 的 AE 只訓練一次，快取三種特徵空間
                feat_cache = {}
                for n_layers, ratio_label in param_configs:
                    ratio   = BOTTLENECK_RATIOS[ratio_label]
                    n_units = max(2, round(input_dim * ratio))
                    try:
                        # 提取深度特徵 DF
                        DF_maj, DF_tst = train_and_extract(
                            ae_type, OF_maj, OF_tst, n_layers, n_units)

                        # DF 自身 MinMax（與 B grid 一致）
                        sc_df    = MinMaxScaler()
                        DF_maj_s = sc_df.fit_transform(DF_maj)
                        DF_tst_s = sc_df.transform(DF_tst)

                        # COMB = hstack(OF, DF) → 再 MinMax（與 C grid 一致）
                        sc_comb  = MinMaxScaler()
                        COMB_maj = sc_comb.fit_transform(np.hstack([OF_maj, DF_maj]))
                        COMB_tst = sc_comb.transform(np.hstack([OF_tst, DF_tst]))

                        feat_cache[(n_layers, ratio_label)] = (
                            DF_maj_s, DF_tst_s, COMB_maj, COMB_tst)

                    except Exception as e:
                        print(f"  [ERROR] AE={ae_type} h{n_layers}-{ratio_label} "
                              f"Fold{fold}: {e}")

                for occ_type in OCC_TYPES:
                    for (n_layers, ratio_label), (
                            DF_maj_s, DF_tst_s, COMB_maj, COMB_tst
                    ) in feat_cache.items():

                        cfg_label = f"h{n_layers}-{ratio_label}"
                        try:
                            # ── 取得三個 OCC 的 raw scores ─────────────────────
                            s_raw = []   # [(s_maj, s_tst), ...]  idx 0=A 1=B 2=C
                            for X_maj_feat, X_tst_feat in [
                                (OF_maj,    OF_tst),
                                (DF_maj_s,  DF_tst_s),
                                (COMB_maj,  COMB_tst),
                            ]:
                                sm, st = train_occ_scores(
                                    occ_type, X_maj_feat, X_tst_feat, n_nb_cap)
                                s_raw.append((sm, st))

                            # ── Baseline A / B / C（單一 representation）─────
                            # 保留各 baseline 的 binary y_pred 供後續 hard voting 使用
                            baseline_metrics = {}
                            baseline_ypred   = {}   # tag → y_pred (np.ndarray)
                            for tag, (sm, st) in zip(["A", "B", "C"], s_raw):
                                t      = np.percentile(sm, 90)
                                y_pred = (st >= t).astype(int)
                                baseline_metrics[tag] = metrics_from(y_tst, y_pred, st)
                                baseline_ypred[tag]   = y_pred

                            # ── Normalize（train-maj min-max，無 data leak）────
                            s_norm = []   # [(sm_n, st_n), ...]
                            for sm, st in s_raw:
                                sm_n, st_n = normalize_by_majority(sm, st)
                                s_norm.append((sm_n, st_n))

                            # ── 12 組 pair ensemble（3 pair × 4 scheme）────────
                            # 4 schemes: eq / istd（soft）+ hard_or / hard_and（hard）
                            pair_metrics = {}
                            tag_map = {0: "A", 1: "B", 2: "C"}   # idx → baseline tag
                            for pair_name, idx_i, idx_j in PAIRS:
                                sm_i_n, st_i_n = s_norm[idx_i]
                                sm_j_n, st_j_n = s_norm[idx_j]
                                # Soft：等權重 + inverse-std 加權
                                for scheme in ("eq", "istd"):
                                    key = f"Pair_{pair_name}_{scheme}"
                                    pair_metrics[key] = weighted_pair(
                                        sm_i_n, st_i_n,
                                        sm_j_n, st_j_n,
                                        y_tst, scheme)
                                # Hard：or（≥1 票）+ and（≥2 票）
                                # 使用 baseline 算好的 binary y_pred（與 baseline 同閾值）
                                y_i_pred = baseline_ypred[tag_map[idx_i]]
                                y_j_pred = baseline_ypred[tag_map[idx_j]]
                                for rule in ("or", "and"):
                                    key = f"Pair_{pair_name}_hard_{rule}"
                                    pair_metrics[key] = hard_pair(
                                        y_i_pred, y_j_pred, y_tst, rule)

                            method_metrics = {**baseline_metrics, **pair_metrics}

                        except Exception as e:
                            print(f"  [ERROR] {ae_type}×{occ_type}({cfg_label}) "
                                  f"Fold{fold}: {e}")
                            method_metrics = {
                                m: {k: float("nan") for k in METRIC_COLS}
                                for m in METHODS
                            }

                        for method in METHODS:
                            all_records.append({
                                "Dataset": ds_name,
                                "AE":      ae_type,
                                "OCC":     occ_type,
                                "Config":  cfg_label,
                                "Fold":    fold,
                                "Method":  method,
                                **method_metrics[method],
                            })

                    # ── 每個 (ae, occ, fold) 列印 6 組 pair 的 best config ────
                    sub = pd.DataFrame([
                        r for r in all_records
                        if r["Dataset"] == ds_name and r["AE"] == ae_type
                        and r["OCC"] == occ_type and r["Fold"] == fold
                        and r["Method"].startswith("Pair_")
                    ])
                    for method in [m for m in METHODS if m.startswith("Pair_")]:
                        sub_m = sub[sub["Method"] == method].dropna(subset=["AUC"])
                        if sub_m.empty: continue
                        best = sub_m.loc[sub_m["AUC"].idxmax()]
                        print(
                            f"  {ae_type:4s}×{occ_type:8s} [{method:14s}] "
                            f"best={best['Config']:10s} Fold{fold}  "
                            f"AUC={best['AUC']:.4f}  F1={best['F1']:.4f}  "
                            f"Recall={best['Recall']:.4f}  G-mean={best['G-mean']:.4f}"
                        )

    df_all = pd.DataFrame(all_records)

    # ── df_best：per-dataset 選法（與 F 一致，避免 per-fold leakage）──
    df_clean = df_all.dropna(subset=["AUC"])
    best_cfg = (
        df_clean.groupby(["Dataset", "AE", "OCC", "Method", "Config"])["AUC"]
                .mean().reset_index()
                .sort_values("AUC", ascending=False)
                .drop_duplicates(["Dataset", "AE", "OCC", "Method"])
    )
    df_best = df_clean.merge(
        best_cfg[["Dataset", "AE", "OCC", "Method", "Config"]],
        on=["Dataset", "AE", "OCC", "Method", "Config"],
    )

    return df_all, df_best


# ─────────────────────────── Excel 樣式定義 ──────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

AE_FILL = {
    "AE":  PatternFill("solid", fgColor="DAEEF3"),
    "DAE": PatternFill("solid", fgColor="E2EFDA"),
    "SAE": PatternFill("solid", fgColor="FFF2CC"),
    "VAE": PatternFill("solid", fgColor="FCE4D6"),
}
# Baseline reference 用淡色；三對 pair 各一色系：
#   _eq        淡色，_istd   深色（soft 雙胞胎）
#   _hard_or   暖色，_hard_and 暖色加深（hard 雙胞胎）
METHOD_FILL = {
    "A":                  PatternFill("solid", fgColor="F8CBAD"),
    "B":                  PatternFill("solid", fgColor="C6E0B4"),
    "C":                  PatternFill("solid", fgColor="BDD7EE"),
    "Pair_AB_eq":         PatternFill("solid", fgColor="FFE699"),
    "Pair_AB_istd":       PatternFill("solid", fgColor="FFD966"),
    "Pair_AB_hard_or":    PatternFill("solid", fgColor="F4B084"),
    "Pair_AB_hard_and":   PatternFill("solid", fgColor="ED7D31"),
    "Pair_AC_eq":         PatternFill("solid", fgColor="B4C7E7"),
    "Pair_AC_istd":       PatternFill("solid", fgColor="8FAADC"),
    "Pair_AC_hard_or":    PatternFill("solid", fgColor="9DC3E6"),
    "Pair_AC_hard_and":   PatternFill("solid", fgColor="2E75B6"),
    "Pair_BC_eq":         PatternFill("solid", fgColor="D9D2E9"),
    "Pair_BC_istd":       PatternFill("solid", fgColor="B4A7D6"),
    "Pair_BC_hard_or":    PatternFill("solid", fgColor="C5B4E3"),
    "Pair_BC_hard_and":   PatternFill("solid", fgColor="7030A0"),
}

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def sc(cell, value, font=None, fill=None, align=None, fmt=None):
    cell.value  = value
    cell.border = THIN_BORDER
    if font:  cell.font          = font
    if fill:  cell.fill          = fill
    if align: cell.alignment     = align
    if fmt:   cell.number_format = fmt


def col_w(ws, letter, width):
    ws.column_dimensions[letter].width = width


# ─────────────────────────── per_fold 分頁 ───────────────────────────────────
def write_per_fold(ws, df, title):
    ws.title = title
    headers  = ["Dataset", "AE", "OCC", "Config", "Fold", "Method"] + METRIC_COLS

    for c, h in enumerate(headers, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        m_fill = METHOD_FILL.get(row["Method"])
        a_fill = AE_FILL.get(row["AE"])
        for c, col in enumerate(headers, 1):
            cell_fill = m_fill if (col == "Method" or col in METRIC_COLS) else a_fill
            sc(ws.cell(r, c), row[col],
               font=BODY_FONT, fill=cell_fill,
               align=LEFT_ALIGN if c <= 2 else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS else None)

    for i, w in enumerate([28, 6, 10, 14, 6, 16, 10, 10, 10, 10], 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


# ─────────────────────────── summary 分頁 ────────────────────────────────────
def write_summary_all(ws, df):
    ws.title  = "all_summary"
    configs   = ALL_CONFIGS
    grouped   = (df.groupby(["Dataset", "Method", "AE", "OCC", "Config"])[METRIC_COLS]
                   .agg(["mean", "std"]).reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for method in METHODS:
        m_span = len(AE_TYPES) * len(OCC_TYPES) * len(configs) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + m_span - 1)
        sc(ws.cell(1, col), method,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=METHOD_FILL.get(method), align=CENTER_ALIGN)
        for ae in AE_TYPES:
            ae_span = len(OCC_TYPES) * len(configs) * len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + ae_span - 1)
            sc(ws.cell(2, col), ae, font=SUBHDR_FONT, fill=AE_FILL.get(ae), align=CENTER_ALIGN)
            for occ in OCC_TYPES:
                occ_span = len(configs) * len(METRIC_COLS)
                ws.merge_cells(start_row=3, start_column=col, end_row=3,
                               end_column=col + occ_span - 1)
                sc(ws.cell(3, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                for cfg in configs:
                    ws.merge_cells(start_row=4, start_column=col, end_row=4,
                                   end_column=col + len(METRIC_COLS) - 1)
                    sc(ws.cell(4, col), cfg, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                    for metric in METRIC_COLS:
                        sc(ws.cell(5, col), metric, font=SUBHDR_FONT,
                           fill=SUBHDR_FILL, align=CENTER_ALIGN)
                        col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 6):
        fill = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=fill, align=LEFT_ALIGN)
        col = 2
        for method in METHODS:
            for ae in AE_TYPES:
                for occ in OCC_TYPES:
                    for cfg in configs:
                        sub = grouped[
                            (grouped["Dataset"] == ds) & (grouped["Method"] == method) &
                            (grouped["AE"] == ae) & (grouped["OCC"] == occ) &
                            (grouped["Config"] == cfg)
                        ]
                        for metric in METRIC_COLS:
                            try:
                                m = sub[(metric, "mean")].values[0]
                                s = sub[(metric, "std")].values[0]
                                display = f"{m:.4f} ± {s:.4f}"
                            except Exception:
                                display = "N/A"
                            sc(ws.cell(r, col), display, font=BODY_FONT,
                               fill=METHOD_FILL.get(method) if fill is None else fill,
                               align=CENTER_ALIGN)
                            col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(METHODS) * len(AE_TYPES) * len(OCC_TYPES) *
                   len(configs) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 16)
    ws.freeze_panes = "B6"


def write_summary_best(ws, df):
    ws.title = "best_summary"
    grouped  = (df.groupby(["Dataset", "Method", "AE", "OCC"])[METRIC_COLS]
                  .agg(["mean", "std"]).reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for method in METHODS:
        m_span = len(AE_TYPES) * len(OCC_TYPES) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + m_span - 1)
        sc(ws.cell(1, col), method,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=METHOD_FILL.get(method), align=CENTER_ALIGN)
        for ae in AE_TYPES:
            ae_span = len(OCC_TYPES) * len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2,
                           end_column=col + ae_span - 1)
            sc(ws.cell(2, col), ae, font=SUBHDR_FONT, fill=AE_FILL.get(ae), align=CENTER_ALIGN)
            for occ in OCC_TYPES:
                ws.merge_cells(start_row=3, start_column=col, end_row=3,
                               end_column=col + len(METRIC_COLS) - 1)
                sc(ws.cell(3, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                col += len(METRIC_COLS)

    col = 2
    for method in METHODS:
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                for metric in METRIC_COLS:
                    sc(ws.cell(4, col), metric, font=SUBHDR_FONT,
                       fill=SUBHDR_FILL, align=CENTER_ALIGN)
                    col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 5):
        fill = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=fill, align=LEFT_ALIGN)
        col = 2
        for method in METHODS:
            for ae in AE_TYPES:
                for occ in OCC_TYPES:
                    sub = grouped[
                        (grouped["Dataset"] == ds) & (grouped["Method"] == method) &
                        (grouped["AE"] == ae)       & (grouped["OCC"] == occ)
                    ]
                    for metric in METRIC_COLS:
                        try:
                            m = sub[(metric, "mean")].values[0]
                            s = sub[(metric, "std")].values[0]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, col), display, font=BODY_FONT,
                           fill=METHOD_FILL.get(method) if fill is None else fill,
                           align=CENTER_ALIGN)
                        col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(METHODS) * len(AE_TYPES) * len(OCC_TYPES) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 16)
    ws.freeze_panes = "B5"


# ─────────────────────────── overall 分頁 ────────────────────────────────────
def write_overall_all(ws, df):
    ws.title = "all_overall"
    configs  = ALL_CONFIGS

    total_cols = 4 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（all configs, all datasets & folds）- E: Pairs Weighted Voting Grid"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["Method", "AE", "OCC", "Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["Method", "AE", "OCC", "Config"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for method in METHODS:
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                for cfg in configs:
                    m_fill = METHOD_FILL.get(method)
                    sc(ws.cell(r, 1), method, font=BOLD_FONT, fill=m_fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 2), ae,     font=BODY_FONT, fill=AE_FILL.get(ae), align=CENTER_ALIGN)
                    sc(ws.cell(r, 3), occ,    font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 4), cfg,    font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                    for i, metric in enumerate(METRIC_COLS, 5):
                        try:
                            m = overall.loc[(method, ae, occ, cfg), (metric, "mean")]
                            s = overall.loc[(method, ae, occ, cfg), (metric, "std")]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, i), display, font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                    r += 1

    for i, w in enumerate([16, 8, 12, 14] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


def write_overall_best(ws, df):
    ws.title = "best_overall"

    total_cols = 4 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（best config per dataset, all datasets）- E: Pairs Weighted Voting Grid"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["Method", "AE", "OCC", "Most Freq Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["Method", "AE", "OCC"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for method in METHODS:
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                m_fill = METHOD_FILL.get(method)
                sc(ws.cell(r, 1), method, font=BOLD_FONT, fill=m_fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 2), ae,     font=BODY_FONT, fill=AE_FILL.get(ae), align=CENTER_ALIGN)
                sc(ws.cell(r, 3), occ,    font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                cfg_ser = df[(df["Method"] == method) &
                             (df["AE"] == ae) & (df["OCC"] == occ)]["Config"]
                cfg = cfg_ser.mode().iloc[0] if not cfg_ser.empty else "N/A"
                sc(ws.cell(r, 4), cfg, font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                for i, metric in enumerate(METRIC_COLS, 5):
                    try:
                        m = overall.loc[(method, ae, occ), (metric, "mean")]
                        s = overall.loc[(method, ae, occ), (metric, "std")]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, i), display, font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                r += 1

    for i, w in enumerate([16, 8, 12, 18] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


# ─────────────────────────── Excel 存檔 ──────────────────────────────────────
def save_excel(df_all, df_best):
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet(); ws3 = wb.create_sheet()
    ws4 = wb.create_sheet(); ws5 = wb.create_sheet(); ws6 = wb.create_sheet(); ws7 = wb.create_sheet()

    write_per_fold(    ws1, df_all,  "all_per_fold")
    write_summary_all( ws2, df_all)
    write_overall_all( ws3, df_all)
    write_per_fold(    ws4, df_best, "best_per_fold")
    write_summary_best(ws5, df_best)
    write_overall_best(ws6, df_best)

    # 給論文畫圖 / 最終整合用：只保留 ensemble 方法，避免 A/B/C reference baseline
    # 因為隨 AE、Config 被重複記錄而在 Method-level 統計時被重複加權。
    df_plot_ready = df_best[~df_best["Method"].isin(["A", "B", "C"])].copy()
    write_per_fold(ws7, df_plot_ready, "plot_ready_best")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 68)
    print("Ensemble E Grid Search：pairs of OCCs by weighted voting（全參數搜尋）")
    print("AE  類型：AE / DAE / SAE / VAE")
    print("OCC 方法：OCSVM / LOF / iForest（三個特徵空間各一個 OCC）")
    print(f"搜尋空間：n_layers={N_LAYERS_LIST} × "
          f"bottleneck={list(BOTTLENECK_RATIOS.keys())}")
    print(f"          共 {len(N_LAYERS_LIST) * len(BOTTLENECK_RATIOS)} 種組合 per (AE, OCC, fold)")
    print(f"Methods ：{METHODS}")
    print("Pairs   ：AB = (OCC1/OF, OCC2/DF)")
    print("          AC = (OCC1/OF, OCC3/OF+DF)")
    print("          BC = (OCC2/DF, OCC3/OF+DF)")
    print("Schemes ：_eq        = equal weights soft voting (0.5, 0.5)")
    print("          _istd      = inverse-std weighted soft voting (w ∝ 1/std)")
    print("          _hard_or   = binary OR  voting (≥1 vote → anomaly, recall ↑)")
    print("          _hard_and  = binary AND voting (≥2 votes → anomaly, precision ↑)")
    print("Threshold：")
    print("  Soft → train-maj weighted scores 的第 90 百分位（與 A/B/C 一致）")
    print("  Hard → 各 OCC 各自用 train-maj 90 percentile 切，再 OR/AND 合併")
    print("選擇準則：AUC 最高")
    print("分頁：all_per_fold / all_summary / all_overall")
    print("      best_per_fold / best_summary / best_overall")
    print("=" * 68)

    df_all, df_best = run_experiment()

    if df_all.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_all, df_best)
        print("\n── Best Config Overall Mean（all datasets, pair ensemble methods）──")
        ens = df_best[df_best["Method"].str.startswith("Pair_")]
        print(ens.groupby(["Method", "AE", "OCC"])[METRIC_COLS]
              .mean().round(4).to_string())
