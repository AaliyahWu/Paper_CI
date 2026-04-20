"""
D_ensemble_voting_grid.py
=========================
Ensemble 方法一：OCC1 + OCC2 + OCC3 by voting（對應投影片 "a. OCC1+2+3 by voting"）

核心概念：同一個 OCC 方法（OCSVM / LOF / iForest）分別跑在三種特徵空間上：
  OCC1 = OCC on OF_maj           （對應 Baseline A）
  OCC2 = OCC on DF_maj           （對應 Baseline B）
  OCC3 = OCC on OF_maj + DF_maj  （對應 Baseline C）
然後對三個 OCC 的輸出做 ensemble。

本檔案記錄兩種 voting 策略：
  1) Vote3_hard  ─ majority hard voting（≥ 2/3 投票為 anomaly）
                   AUC 用 vote count (0–3) 當 ranking score
  2) Vote3_soft  ─ soft voting：各 OCC 的 anomaly score 經 train-maj min-max
                   normalize 後平均，threshold 同樣用 train-maj 第 90 百分位數

為了方便比較，每筆 ensemble 記錄同時附帶 A / B / C 三條 baseline 在同一
(AE, config, fold, OCC) 下的結果（Method = "A" / "B" / "C"），可直接在同一
xlsx 中對比「單一 representation」vs「三人 ensemble」的差異。

Grid 搜尋空間（沿用 C_baseline_grid 的完整空間）：
  n_layers   : [1, 2, 3]
  bottleneck : ["1/4", "1/3", "1/2", "1/1", "2/1", "3/1", "4/1"]
  → 共 21 種 AE 架構 per (AE_type, OCC, fold)

輸入資料夾結構（DATA_ROOT）：
  data/
  └── <dataset_name>/
      ├── <prefix>-1tra.dat  ...  <prefix>-5tra.dat
      └── <prefix>-1tst.dat  ...  <prefix>-5tst.dat

輸出：results/D ensemble voting grid.xlsx
  分頁：
    all_per_fold      所有 (AE, OCC, Config, Fold, Method) 的原始結果
    all_summary       mean ± std across folds（per Dataset × Method × AE × OCC × Config）
    all_overall       全域平均（per Method × AE × OCC × Config）
    best_per_fold     每 (Dataset, AE, OCC, Method) 中 AUC 最高 config 的 fold 資料
    best_summary      best mean ± std across folds
    best_overall      best 全域平均

與 B/C grid 的一致性：
  - AE 模型定義、train_and_extract 函式完全相同
  - OCC 訓練超參數（nu=0.1、contamination=0.1 等）完全相同
  - MinMaxScaler fit 只用 training majority、threshold 用 train-maj 第 90 百分位
  - 隨機種子 torch.manual_seed(42) + np.random.seed(42)
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT   = Path("data")
RESULTS_DIR = Path("results")
OUTPUT_FILE = RESULTS_DIR / "D ensemble voting grid.xlsx"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5

# AE 訓練超參數（與 B / C grid 一致）
AE_EPOCHS     = 100
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1
SAE_SPARSITY  = 1e-3
VAE_BETA      = 1.0

AE_TYPES    = ["AE", "DAE", "SAE", "VAE"]
OCC_TYPES   = ["OCSVM", "LOF", "iForest"]
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

# Grid 搜尋空間（對齊 C_baseline_grid）
N_LAYERS_LIST     = [1, 2, 3]
BOTTLENECK_RATIOS = {
    "1/4": 0.25, "1/3": 1/3, "1/2": 0.5, "1/1": 1.0,
    "2/1": 2.0,  "3/1": 3.0, "4/1": 4.0,
}
ALL_CONFIGS = [f"h{nl}-{rl}" for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]

# 本檔產出的 method 清單（順序決定 Excel 排列）
METHODS = ["A", "B", "C", "Vote3_hard", "Vote3_soft"]


# ─────────────────────────── AE 模型定義 ─────────────────────────────────────
class AEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        for i in range(len(dims) - 1):
            d_in, d_out = dims[-(i+1)], dims[-(i+2)]
            act = nn.Sigmoid() if i == len(dims) - 2 else nn.ReLU()
            dec += [nn.Linear(d_in, d_out), act]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


class VAEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        base = []
        for i in range(len(dims) - 2):
            base += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        self.enc_base  = nn.Sequential(*base) if base else nn.Identity()
        mid = dims[-2] if len(dims) >= 2 else input_dim
        self.fc_mu     = nn.Linear(mid, dims[-1])
        self.fc_logvar = nn.Linear(mid, dims[-1])
        dec_dims = dims[::-1]
        dec = []
        for i in range(len(dec_dims) - 1):
            act = nn.Sigmoid() if i == len(dec_dims) - 2 else nn.ReLU()
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), act]
        self.decoder = nn.Sequential(*dec)

    def reparameterize(self, mu, lv):
        return mu + torch.exp(0.5 * lv) * torch.randn_like(lv)

    def forward(self, x):
        h  = self.enc_base(x)
        mu = self.fc_mu(h)
        lv = self.fc_logvar(h)
        z  = self.reparameterize(mu, lv)
        return self.decoder(z), z, mu, lv


# ─────────────────────────── AE 訓練 + 特徵提取 ──────────────────────────────
def train_and_extract(ae_type, X_maj_s, X_test_s, n_layers, n_units):
    """與 B/C grid 完全相同；回傳 (feat_maj, feat_test)"""
    input_dim = X_maj_s.shape[1]
    model = VAEModel(input_dim, n_layers, n_units) if ae_type == "VAE" \
            else AEModel(input_dim, n_layers, n_units)

    optim  = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))
    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim.zero_grad()
            if ae_type == "DAE":
                xb_n = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
                xr, _ = model(xb_n)
                loss  = mse(xr, xb)
            elif ae_type == "SAE":
                xr, z = model(xb)
                loss  = mse(xr, xb) + SAE_SPARSITY * z.abs().mean()
            elif ae_type == "VAE":
                xr, _, mu, lv = model(xb)
                kl   = -0.5 * (1 + lv - mu.pow(2) - lv.exp()).mean()
                loss = mse(xr, xb) + VAE_BETA * kl
            else:
                xr, _ = model(xb)
                loss  = mse(xr, xb)
            loss.backward()
            optim.step()

    model.eval()

    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            if ae_type == "VAE":
                _, _, mu, _ = model(xt)
                return mu.numpy()
            else:
                _, z = model(xt)
                return z.numpy()

    return extract(X_maj_s), extract(X_test_s)


# ─────────────────────────── OCC 分數取得 ────────────────────────────────────
def train_occ_scores(occ_type, X_maj, X_test, n_neighbors_cap):
    """
    在 X_maj 上訓練 OCC，回傳 (scores_maj, scores_test)。
    X_maj / X_test 都須為已 MinMaxScaled 的矩陣。
    scores 越大代表越像 anomaly（等同 -decision_function，與 baseline 一致）。
    """
    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(X_maj)
        s_maj  = -clf.decision_function(X_maj)
        s_test = -clf.decision_function(X_test)
    elif occ_type == "LOF":
        k = min(20, n_neighbors_cap)
        clf = LocalOutlierFactor(n_neighbors=k, novelty=True, contamination=0.1)
        clf.fit(X_maj)
        s_maj  = -clf.decision_function(X_maj)
        s_test = -clf.decision_function(X_test)
    else:  # iForest
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(X_maj)
        s_maj  = -clf.decision_function(X_maj)
        s_test = -clf.decision_function(X_test)
    return np.asarray(s_maj, dtype=float), np.asarray(s_test, dtype=float)


def normalize_by_majority(scores_maj, scores_test):
    """
    以 training majority 分數的 min-max 做線性變換，
    training maj → [0, 1]（大致），test 可能超出（異常樣本分數可能 > 1，正常）。
    不使用 test set 統計，避免 data leak。
    """
    lo, hi = float(scores_maj.min()), float(scores_maj.max())
    if hi - lo < 1e-12:
        return np.zeros_like(scores_maj), np.zeros_like(scores_test)
    return (scores_maj - lo) / (hi - lo), (scores_test - lo) / (hi - lo)


# ─────────────────────────── 評估指標 ────────────────────────────────────────
def gmean_score(y_true, y_pred_binary):
    cm = confusion_matrix(y_true, y_pred_binary, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return float(np.sqrt(sens * spec))
    return 0.0


def metrics_from(y_true, y_pred_binary, scores_test_for_auc):
    """
    用 continuous scores 算 AUC，用 binary preds 算 F1、Recall、G-mean。
    與 A/B/C baseline 完全一致的做法。
    """
    try:
        auc = (roc_auc_score(y_true, scores_test_for_auc)
               if len(np.unique(y_true)) >= 2 else float("nan"))
    except Exception:
        auc = float("nan")
    f1  = f1_score(y_true, y_pred_binary, pos_label=1, zero_division=0)
    rec = recall_score(y_true, y_pred_binary, pos_label=1, zero_division=0)
    gm  = gmean_score(y_true, y_pred_binary)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── KEEL .dat 解析 ──────────────────────────────────
def parse_keel_dat(filepath):
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"):
            continue
        if s.lower() == "@data":
            data_start = True
            continue
        if data_start:
            rows.append(s)
    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    y_raw = df[label_col].values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    unique, counts = np.unique(y_raw, return_counts=True)
    minority_label = unique[np.argmin(counts)]
    y = (y_raw == minority_label).astype(int)
    return X, y


# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment():
    """
    回傳：
      df_all  — 全部 (AE, OCC, Config, Fold, Method) 的結果
      df_best — 每 (Dataset, AE, OCC, Method) 中 AUC 最高 config 的 fold 資料
    """
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    param_configs = [(nl, rl) for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]
    all_records = []

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n{'='*65}")
        print(f"▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            # 檔名格式多版支援（與 B/C grid 一致）
            file_prefix  = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]

            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)

            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案")
                continue

            try:
                X_tra, y_tra = parse_keel_dat(tra_file)
                X_tst, y_tst = parse_keel_dat(tst_file)
                input_dim = X_tra.shape[1]
                X_maj = X_tra[y_tra == 0]

                if len(X_maj) < 5:
                    print(f"  [SKIP] Fold {fold}: 訓練集正常樣本不足 ({len(X_maj)})")
                    continue
                if y_tst.sum() == 0:
                    print(f"  [SKIP] Fold {fold}: 測試集無少數類樣本")
                    continue

                # OF（原始特徵）MinMax，fit 只用訓練集多數類
                scaler_of = MinMaxScaler()
                OF_maj = scaler_of.fit_transform(X_maj)
                OF_tst = scaler_of.transform(X_tst)
                n_nb_cap = max(1, len(X_maj) - 1)

            except Exception as e:
                print(f"  [ERROR] Fold {fold} 資料載入失敗: {e}")
                continue

            for ae_type in AE_TYPES:
                # 每個 (ae_type, config) 的 AE 只訓練一次，快取三種特徵空間
                feat_cache = {}
                for n_layers, ratio_label in param_configs:
                    ratio   = BOTTLENECK_RATIOS[ratio_label]
                    n_units = max(2, round(input_dim * ratio))
                    try:
                        DF_maj, DF_tst = train_and_extract(
                            ae_type, OF_maj, OF_tst, n_layers, n_units)

                        # DF 自身重新 MinMax（與 B grid 一致）
                        sc_df = MinMaxScaler()
                        DF_maj_s = sc_df.fit_transform(DF_maj)
                        DF_tst_s = sc_df.transform(DF_tst)

                        # COMB = [OF_scaled, DF_raw] → 再 MinMax（與 C grid 一致）
                        COMB_maj_raw = np.hstack([OF_maj, DF_maj])
                        COMB_tst_raw = np.hstack([OF_tst, DF_tst])
                        sc_comb = MinMaxScaler()
                        COMB_maj = sc_comb.fit_transform(COMB_maj_raw)
                        COMB_tst = sc_comb.transform(COMB_tst_raw)

                        feat_cache[(n_layers, ratio_label)] = (
                            DF_maj_s, DF_tst_s, COMB_maj, COMB_tst)
                    except Exception as e:
                        print(f"  [ERROR] AE={ae_type} h{n_layers}-{ratio_label} "
                              f"Fold{fold}: {e}")

                for occ_type in OCC_TYPES:
                    for (n_layers, ratio_label), (
                            DF_maj_s, DF_tst_s, COMB_maj, COMB_tst) in feat_cache.items():

                        cfg_label = f"h{n_layers}-{ratio_label}"
                        try:
                            # ─── OCC1 on OF ─────────────────────────────────
                            s1_maj, s1_tst = train_occ_scores(
                                occ_type, OF_maj, OF_tst, n_nb_cap)
                            t1 = np.percentile(s1_maj, 90)
                            y1 = (s1_tst >= t1).astype(int)
                            m_A = metrics_from(y_tst, y1, s1_tst)

                            # ─── OCC2 on DF ─────────────────────────────────
                            s2_maj, s2_tst = train_occ_scores(
                                occ_type, DF_maj_s, DF_tst_s, n_nb_cap)
                            t2 = np.percentile(s2_maj, 90)
                            y2 = (s2_tst >= t2).astype(int)
                            m_B = metrics_from(y_tst, y2, s2_tst)

                            # ─── OCC3 on COMB ───────────────────────────────
                            s3_maj, s3_tst = train_occ_scores(
                                occ_type, COMB_maj, COMB_tst, n_nb_cap)
                            t3 = np.percentile(s3_maj, 90)
                            y3 = (s3_tst >= t3).astype(int)
                            m_C = metrics_from(y_tst, y3, s3_tst)

                            # ─── Ensemble: Hard voting ─────────────────────
                            # binary: ≥ 2/3 投票為 anomaly
                            # continuous: vote count (0, 1, 2, 3) 作為 AUC 的 score
                            vote_cnt_tst = y1 + y2 + y3
                            y_hard = (vote_cnt_tst >= 2).astype(int)
                            m_hard = metrics_from(y_tst, y_hard, vote_cnt_tst)

                            # ─── Ensemble: Soft voting ─────────────────────
                            # 各 OCC scores 用 train-maj min-max normalize 後平均
                            s1_mn, s1_tn = normalize_by_majority(s1_maj, s1_tst)
                            s2_mn, s2_tn = normalize_by_majority(s2_maj, s2_tst)
                            s3_mn, s3_tn = normalize_by_majority(s3_maj, s3_tst)
                            s_soft_maj = (s1_mn + s2_mn + s3_mn) / 3.0
                            s_soft_tst = (s1_tn + s2_tn + s3_tn) / 3.0
                            t_soft = np.percentile(s_soft_maj, 90)
                            y_soft = (s_soft_tst >= t_soft).astype(int)
                            m_soft = metrics_from(y_tst, y_soft, s_soft_tst)

                            method_metrics = {
                                "A":          m_A,
                                "B":          m_B,
                                "C":          m_C,
                                "Vote3_hard": m_hard,
                                "Vote3_soft": m_soft,
                            }

                        except Exception as e:
                            print(f"  [ERROR] {ae_type}×{occ_type}({cfg_label}) "
                                  f"Fold{fold}: {e}")
                            method_metrics = {
                                m: {k: float("nan") for k in METRIC_COLS}
                                for m in METHODS
                            }

                        for method in METHODS:
                            all_records.append({
                                "Dataset": ds_name,
                                "AE":      ae_type,
                                "OCC":     occ_type,
                                "Config":  cfg_label,
                                "Fold":    fold,
                                "Method":  method,
                                **method_metrics[method],
                            })

                    # 每個 (ae_type, occ_type, fold) 組合結束，列印 5 method × 1 best config
                    sub = pd.DataFrame([
                        r for r in all_records
                        if r["Dataset"] == ds_name and r["AE"] == ae_type
                        and r["OCC"] == occ_type and r["Fold"] == fold
                    ])
                    if sub.empty:
                        continue
                    for method in ["Vote3_hard", "Vote3_soft"]:
                        sub_m = sub[sub["Method"] == method].dropna(subset=["AUC"])
                        if sub_m.empty:
                            continue
                        best = sub_m.loc[sub_m["AUC"].idxmax()]
                        print(
                            f"  {ae_type:4s}×{occ_type:8s} [{method:10s}] "
                            f"best={best['Config']:10s} Fold{fold}  "
                            f"AUC={best['AUC']:.4f}  F1={best['F1']:.4f}  "
                            f"Recall={best['Recall']:.4f}  G-mean={best['G-mean']:.4f}"
                        )

    df_all = pd.DataFrame(all_records)

    # df_best：每 (Dataset, AE, OCC, Method) 在該 fold 內 AUC 最高的 config
    best_rows = []
    if not df_all.empty:
        for keys, grp in df_all.groupby(["Dataset", "AE", "OCC", "Method", "Fold"]):
            g = grp.dropna(subset=["AUC"])
            if g.empty:
                continue
            best_rows.append(g.loc[g["AUC"].idxmax()].to_dict())
    df_best = pd.DataFrame(best_rows)

    return df_all, df_best


# ─────────────────────────── Excel 樣式定義 ──────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

AE_FILL = {
    "AE":  PatternFill("solid", fgColor="DAEEF3"),
    "DAE": PatternFill("solid", fgColor="E2EFDA"),
    "SAE": PatternFill("solid", fgColor="FFF2CC"),
    "VAE": PatternFill("solid", fgColor="FCE4D6"),
}
METHOD_FILL = {
    "A":          PatternFill("solid", fgColor="F8CBAD"),
    "B":          PatternFill("solid", fgColor="C6E0B4"),
    "C":          PatternFill("solid", fgColor="BDD7EE"),
    "Vote3_hard": PatternFill("solid", fgColor="FFE699"),
    "Vote3_soft": PatternFill("solid", fgColor="FFD966"),
}

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def sc(cell, value, font=None, fill=None, align=None, fmt=None):
    cell.value  = value
    cell.border = THIN_BORDER
    if font:  cell.font         = font
    if fill:  cell.fill         = fill
    if align: cell.alignment    = align
    if fmt:   cell.number_format = fmt


def col_w(ws, letter, width):
    ws.column_dimensions[letter].width = width


# ─────────────────────────── per_fold 分頁 ───────────────────────────────────
def write_per_fold(ws, df, title):
    ws.title = title
    headers = ["Dataset", "AE", "OCC", "Config", "Fold", "Method"] + METRIC_COLS

    for c, h in enumerate(headers, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        method_fill = METHOD_FILL.get(row["Method"])
        ae_fill     = AE_FILL.get(row["AE"])
        for c, col in enumerate(headers, 1):
            if col == "Method":
                cell_fill = method_fill
            elif col in ("AUC", "F1", "Recall", "G-mean"):
                cell_fill = method_fill
            else:
                cell_fill = ae_fill
            sc(ws.cell(r, c), row[col],
               font=BODY_FONT, fill=cell_fill,
               align=LEFT_ALIGN if c <= 2 else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS else None)

    for i, w in enumerate([28, 6, 10, 14, 6, 12, 10, 10, 10, 10], 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


# ─────────────────────────── summary 分頁 ────────────────────────────────────
def write_summary_all(ws, df):
    """all_summary：Dataset | (Method × AE × OCC × Config × Metric) mean ± std"""
    ws.title = "all_summary"
    configs = ALL_CONFIGS

    grouped = (df.groupby(["Dataset", "Method", "AE", "OCC", "Config"])[METRIC_COLS]
                 .agg(["mean", "std"])
                 .reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for method in METHODS:
        m_span = len(AE_TYPES) * len(OCC_TYPES) * len(configs) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + m_span - 1)
        sc(ws.cell(1, col), method,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=METHOD_FILL.get(method), align=CENTER_ALIGN)
        for ae in AE_TYPES:
            ae_span = len(OCC_TYPES) * len(configs) * len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + ae_span - 1)
            sc(ws.cell(2, col), ae,
               font=SUBHDR_FONT, fill=AE_FILL.get(ae), align=CENTER_ALIGN)
            for occ in OCC_TYPES:
                occ_span = len(configs) * len(METRIC_COLS)
                ws.merge_cells(start_row=3, start_column=col, end_row=3,
                               end_column=col + occ_span - 1)
                sc(ws.cell(3, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                for cfg in configs:
                    cfg_span = len(METRIC_COLS)
                    ws.merge_cells(start_row=4, start_column=col, end_row=4,
                                   end_column=col + cfg_span - 1)
                    sc(ws.cell(4, col), cfg, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                    for metric in METRIC_COLS:
                        sc(ws.cell(5, col), metric, font=SUBHDR_FONT,
                           fill=SUBHDR_FILL, align=CENTER_ALIGN)
                        col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 6):
        fill = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=fill, align=LEFT_ALIGN)
        col = 2
        for method in METHODS:
            for ae in AE_TYPES:
                for occ in OCC_TYPES:
                    for cfg in configs:
                        sub = grouped[
                            (grouped["Dataset"] == ds) &
                            (grouped["Method"]  == method) &
                            (grouped["AE"]      == ae)  &
                            (grouped["OCC"]     == occ) &
                            (grouped["Config"]  == cfg)
                        ]
                        for metric in METRIC_COLS:
                            try:
                                m = sub[(metric, "mean")].values[0]
                                s = sub[(metric, "std")].values[0]
                                display = f"{m:.4f} ± {s:.4f}"
                            except Exception:
                                display = "N/A"
                            sc(ws.cell(r, col), display,
                               font=BODY_FONT,
                               fill=METHOD_FILL.get(method) if fill is None else fill,
                               align=CENTER_ALIGN)
                            col += 1

    col_w(ws, "A", 28)
    total_cols = len(METHODS) * len(AE_TYPES) * len(OCC_TYPES) * len(configs) * len(METRIC_COLS)
    for i in range(2, 2 + total_cols):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B6"


def write_summary_best(ws, df):
    """best_summary：Dataset | (Method × AE × OCC × Metric) mean ± std"""
    ws.title = "best_summary"

    grouped = (df.groupby(["Dataset", "Method", "AE", "OCC"])[METRIC_COLS]
                 .agg(["mean", "std"])
                 .reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for method in METHODS:
        m_span = len(AE_TYPES) * len(OCC_TYPES) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + m_span - 1)
        sc(ws.cell(1, col), method,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=METHOD_FILL.get(method), align=CENTER_ALIGN)
        for ae in AE_TYPES:
            ae_span = len(OCC_TYPES) * len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2,
                           end_column=col + ae_span - 1)
            sc(ws.cell(2, col), ae, font=SUBHDR_FONT, fill=AE_FILL.get(ae), align=CENTER_ALIGN)
            for occ in OCC_TYPES:
                ws.merge_cells(start_row=3, start_column=col, end_row=3,
                               end_column=col + len(METRIC_COLS) - 1)
                sc(ws.cell(3, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                col += len(METRIC_COLS)

    col = 2
    for method in METHODS:
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                for metric in METRIC_COLS:
                    sc(ws.cell(4, col), metric, font=SUBHDR_FONT,
                       fill=SUBHDR_FILL, align=CENTER_ALIGN)
                    col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 5):
        fill = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=fill, align=LEFT_ALIGN)
        col = 2
        for method in METHODS:
            for ae in AE_TYPES:
                for occ in OCC_TYPES:
                    sub = grouped[
                        (grouped["Dataset"] == ds) &
                        (grouped["Method"]  == method) &
                        (grouped["AE"]      == ae)  &
                        (grouped["OCC"]     == occ)
                    ]
                    for metric in METRIC_COLS:
                        try:
                            m = sub[(metric, "mean")].values[0]
                            s = sub[(metric, "std")].values[0]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, col), display,
                           font=BODY_FONT,
                           fill=METHOD_FILL.get(method) if fill is None else fill,
                           align=CENTER_ALIGN)
                        col += 1

    col_w(ws, "A", 28)
    total_cols = len(METHODS) * len(AE_TYPES) * len(OCC_TYPES) * len(METRIC_COLS)
    for i in range(2, 2 + total_cols):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B5"


# ─────────────────────────── overall 分頁 ────────────────────────────────────
def write_overall_all(ws, df):
    """all_overall：(Method, AE, OCC, Config) 全域平均"""
    ws.title = "all_overall"
    configs = ALL_CONFIGS

    total_cols = 4 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（all configs, all datasets & folds）- D: OCC1+2+3 Voting Grid"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["Method", "AE", "OCC", "Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["Method", "AE", "OCC", "Config"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for method in METHODS:
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                for cfg in configs:
                    m_fill = METHOD_FILL.get(method)
                    sc(ws.cell(r, 1), method, font=BOLD_FONT, fill=m_fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 2), ae,     font=BODY_FONT, fill=AE_FILL.get(ae),
                       align=CENTER_ALIGN)
                    sc(ws.cell(r, 3), occ,    font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 4), cfg,    font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                    for i, metric in enumerate(METRIC_COLS, 5):
                        try:
                            m = overall.loc[(method, ae, occ, cfg), (metric, "mean")]
                            s = overall.loc[(method, ae, occ, cfg), (metric, "std")]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, i), display, font=BODY_FONT, fill=m_fill,
                           align=CENTER_ALIGN)
                    r += 1

    for i, w in enumerate([14, 8, 12, 14] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


def write_overall_best(ws, df):
    """best_overall：(Method, AE, OCC) 全域平均 + 最常選中的 Config"""
    ws.title = "best_overall"

    total_cols = 4 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（best config per fold, all datasets）- D: OCC1+2+3 Voting Grid"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["Method", "AE", "OCC", "Most Freq Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["Method", "AE", "OCC"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for method in METHODS:
        for ae in AE_TYPES:
            for occ in OCC_TYPES:
                m_fill = METHOD_FILL.get(method)
                sc(ws.cell(r, 1), method, font=BOLD_FONT, fill=m_fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 2), ae,     font=BODY_FONT, fill=AE_FILL.get(ae),
                   align=CENTER_ALIGN)
                sc(ws.cell(r, 3), occ,    font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                cfg_series = df[(df["Method"] == method) &
                                (df["AE"]     == ae)     &
                                (df["OCC"]    == occ)]["Config"]
                cfg = cfg_series.mode().iloc[0] if not cfg_series.empty else "N/A"
                sc(ws.cell(r, 4), cfg, font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                for i, metric in enumerate(METRIC_COLS, 5):
                    try:
                        m = overall.loc[(method, ae, occ), (metric, "mean")]
                        s = overall.loc[(method, ae, occ), (metric, "std")]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, i), display, font=BODY_FONT, fill=m_fill, align=CENTER_ALIGN)
                r += 1

    for i, w in enumerate([14, 8, 12, 18] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


# ─────────────────────────── Excel 存檔 ──────────────────────────────────────
def save_excel(df_all, df_best):
    wb = Workbook()

    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()
    ws6 = wb.create_sheet()

    write_per_fold(     ws1, df_all,  "all_per_fold")
    write_summary_all(  ws2, df_all)
    write_overall_all(  ws3, df_all)
    write_per_fold(     ws4, df_best, "best_per_fold")
    write_summary_best( ws5, df_best)
    write_overall_best( ws6, df_best)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 72)
    print("Ensemble D Grid Search：OCC1 + OCC2 + OCC3 by voting（全參數搜尋）")
    print("AE  類型：AE / DAE / SAE / VAE")
    print("OCC 方法：OCSVM / LOF / iForest（三個特徵空間共用同一 OCC 方法）")
    print(f"搜尋空間：n_layers={N_LAYERS_LIST} × bottleneck={list(BOTTLENECK_RATIOS.keys())}")
    print(f"          共 {len(N_LAYERS_LIST) * len(BOTTLENECK_RATIOS)} 種組合 per (AE, OCC, fold)")
    print(f"Methods ：{METHODS}")
    print("Voting  ：Vote3_hard = binary majority voting (≥2/3)")
    print("          Vote3_soft = normalize + mean score soft voting")
    print("選擇準則：AUC 最高")
    print("分頁：all_per_fold / all_summary / all_overall")
    print("      best_per_fold / best_summary / best_overall")
    print("=" * 72)

    df_all, df_best = run_experiment()

    if df_all.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_all, df_best)
        print("\n── Best Config Overall Mean（all datasets, ensemble methods only）──")
        ens_only = df_best[df_best["Method"].isin(["Vote3_hard", "Vote3_soft"])]
        print(ens_only.groupby(["Method", "AE", "OCC"])[METRIC_COLS]
              .mean().round(4).to_string())
