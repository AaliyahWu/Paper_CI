"""
G_baseline_B_fw_OF.py
=====================
實驗二 Phase 1：Feature Weighting 應用方式 1 —— OF 端 weighting

Pipeline:
    OF  ──MinMaxScaler──>  OF_s  ──Feature Weighting──>  OF_s_weighted
        ──DAE 訓練/提取──>  DF  ──MinMaxScaler──>  ──OCC──>  prediction

對照 B_baseline_grid_final.py 的差異：
  1. AE 固定為 DAE（按 Phase 1 規劃，先收斂變數）
  2. 在 MinMax 後、AE 訓練前，對 OF_s 套用 feature weighting
  3. 新增 FW 維度，取代 AE 在輸出表格的位置
  4. 共 5 種 weighting（含 "none" 作為內部 baseline，應等於 B baseline 的 DAE 結果）

Feature Weighting 方法（全部只在 train_maj 上計算，無 leakage）：
  • none : w_j = 1 全部相等（baseline，sanity check）
  • var  : w_j ∝ Var(x_j)                     變異大 → 鑑別力強
  • ivar : w_j ∝ 1 / Var(x_j)                 變異小 → majority 穩定標識
  • mad  : w_j ∝ median|x_j - median(x_j)|    對 outlier robust 的離散度
  • lap  : w_j ∝ 1 / LaplacianScore(x_j)      保留 local structure 的能力

權重歸一化：mean(w) = 1，保持特徵整體 scale 不變（不破壞 MinMax 的 [0,1] 結構）。

搜尋空間（與 B/C/D/E/F 對齊）：
  AE        : 固定 DAE
  FW 方法   : none / var / ivar / mad / lap     ── 新增維度
  OCC       : OCSVM / LOF / iForest
  n_layers  : [1, 2, 3]
  bottleneck: ["1/4", "1/3", "1/2", "1/1", "2/1", "3/1", "4/1"]
  → 5 FW × 3 OCC × 21 configs = 315 種組合 per (fold)

輸出：results/G_baseline_B_fw_OF.xlsx
  分頁（全部組合）：
    all_per_fold      所有 315 種組合每 fold 的原始結果
    all_summary       所有組合 mean ± std（across folds，per dataset）
    all_overall       所有組合全域平均
  分頁（最佳組合，per-dataset 選法）：
    best_per_fold     每 (Dataset, FW, OCC) 中『5-fold 平均 AUC 最高』的 Config
    best_summary      最佳組合 mean ± std
    best_overall      最佳組合全域平均 + 最常選中的 Config
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor, kneighbors_graph
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT   = Path("data")
RESULTS_DIR = Path("results")
OUTPUT_FILE = RESULTS_DIR / "G_baseline_B_fw_OF.xlsx"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5
EPS     = 1e-12

# ── AE 訓練超參數（與 B 完全一致）──
AE_EPOCHS     = 100
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1

# ── Phase 1：固定 DAE ──
AE_TYPE     = "DAE"
OCC_TYPES   = ["OCSVM", "LOF", "iForest"]
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

# ── Feature Weighting 方法（含 "none" 內部 baseline）──
FW_METHODS  = ["none", "var", "ivar", "mad", "lap"]
LAPLACIAN_K = 5     # Laplacian Score 用的 KNN 大小

# ── Grid（與 B 對齊）──
N_LAYERS_LIST     = [1, 2, 3]
BOTTLENECK_RATIOS = {
    "1/4": 0.25, "1/3": 1/3, "1/2": 0.5, "1/1": 1.0,
    "2/1": 2.0,  "3/1": 3.0, "4/1": 4.0,
}
ALL_CONFIGS = [f"h{nl}-{rl}" for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]


# ─────────────────────────── Feature Weighting ───────────────────────────────
def compute_feature_weights(X, method, k=LAPLACIAN_K, eps=EPS):
    """
    在 train majority 上計算每個特徵的權重（非負）。

    參數：
      X      : (n, d) array，已 MinMax-scaled 的 train majority
      method : "none" | "var" | "ivar" | "mad" | "lap"
      k      : Laplacian Score 用的 KNN 鄰居數
      eps    : 數值穩定保護

    返回：
      w : (d,) array，非負原始權重（尚未歸一化）
    """
    n, d = X.shape

    if method == "none":
        return np.ones(d, dtype=float)

    elif method == "var":
        # 變異越大 → 權重越大
        return np.clip(X.var(axis=0), 0, None)

    elif method == "ivar":
        # 變異越小 → 權重越大（majority 內穩定 = anomaly 標識）
        return 1.0 / (X.var(axis=0) + eps)

    elif method == "mad":
        # Median Absolute Deviation：對 outlier robust
        med = np.median(X, axis=0)
        return np.clip(np.median(np.abs(X - med), axis=0), 0, None)

    elif method == "lap":
        # Laplacian Score (He et al., NIPS 2005)
        # 分數越小 → 越能保留 local structure → 越重要
        # 權重取倒數：w_j ∝ 1 / LS_j
        if n < 3:
            return np.ones(d, dtype=float)

        k_eff = max(1, min(k, n - 1))
        # 取鄰近距離當 graph weight 的基底（heat kernel）
        try:
            S = kneighbors_graph(
                X, n_neighbors=k_eff, mode="distance",
                include_self=False,
            ).toarray()
        except Exception:
            return np.ones(d, dtype=float)

        # heat kernel: S_ij = exp(-||x_i - x_j||^2 / t)，t 用 median 距離平方
        pos = S[S > 0]
        t = (np.median(pos) ** 2 + eps) if len(pos) > 0 else 1.0
        S = np.where(S > 0, np.exp(-(S ** 2) / t), 0.0)
        # 對稱化（KNN graph 不對稱）
        S = 0.5 * (S + S.T)

        D_diag = S.sum(axis=1)
        D = np.diag(D_diag)
        L = D - S

        w = np.zeros(d, dtype=float)
        D_sum = D_diag.sum() + eps
        for j in range(d):
            f = X[:, j]
            f_bar = (f * D_diag).sum() / D_sum
            f_tilde = f - f_bar
            num = float(f_tilde @ L @ f_tilde)
            den = float(f_tilde @ D @ f_tilde) + eps
            ls = num / den
            w[j] = 1.0 / (ls + eps)

        return np.clip(w, 0, None)

    else:
        raise ValueError(f"Unknown FW method: {method}")


def normalize_weights(w):
    """歸一化使 max(w) = 1。

    物理意義：最重要的 feature 保留 100%，其他 < 100% 是衰減。

    為什麼用 max=1 而非 mean=1：
      1. 保證 weighted X ∈ [0, X_original] ⊆ [0, 1]
         → DAE 的 noise clamp [0,1] 與 sigmoid output 假設都不被破壞
         → 高權重 dim 不會因為 weighted > 1 而被 sigmoid 永遠重建不出來
      2. mean=1 歸一化在 ivar 模式下，遇到 var≈0 的 dim 會讓
         該 dim weight 爆增、其他 dim 趨近 0；max=1 不會（最大固定 1，
         其他相對縮放）
      3. 純 linear scaling，可逆，保留所有 dim 的相對比例
    """
    w = np.asarray(w, dtype=float)
    w = np.clip(w, 0, None)
    m = w.max()
    if m <= EPS:
        return np.ones_like(w)
    return w / m


def apply_weights(X_maj_s, X_tst_s, method):
    """
    在 OF（已 MinMax-scaled）上計算 weights 並套用至 train + test。
    weights 只用 train_maj 計算，test 沿用 → 無 data leakage。
    """
    w_raw  = compute_feature_weights(X_maj_s, method)
    w_norm = normalize_weights(w_raw)
    return X_maj_s * w_norm, X_tst_s * w_norm, w_norm


# ─────────────────────────── AE 模型（DAE）─────────────────────────────────
class AEModel(nn.Module):
    """DAE 共用架構（與 B_baseline_grid_final.py 完全一致）。"""
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        for i in range(len(dims) - 1):
            d_in, d_out = dims[-(i+1)], dims[-(i+2)]
            act = nn.Sigmoid() if i == len(dims) - 2 else nn.ReLU()
            dec += [nn.Linear(d_in, d_out), act]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


def train_and_extract_dae(X_maj_s, X_test_s, n_layers, n_units):
    """訓練 DAE 並提取深度特徵（與 B 的 train_and_extract 在 DAE 分支完全一致）。"""
    input_dim = X_maj_s.shape[1]
    model = AEModel(input_dim, n_layers, n_units)
    optim_ = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))
    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True,
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim_.zero_grad()
            # DAE: 加 noise → 重建乾淨 x
            xb_n = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
            xr, _ = model(xb_n)
            loss  = mse(xr, xb)
            loss.backward()
            optim_.step()

    model.eval()

    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            _, z = model(xt)
            return z.numpy()

    return extract(X_maj_s), extract(X_test_s)


# ─────────────────────────── 評估指標（與 B 一致）──────────────────────────
def gmean_score(y_true, y_pred_binary):
    cm = confusion_matrix(y_true, y_pred_binary, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return float(np.sqrt(sens * spec))
    return 0.0


def run_occ_eval(occ_type, feat_maj, feat_test, y_test, n_neighbors_cap):
    scaler      = MinMaxScaler()
    feat_maj_s  = scaler.fit_transform(feat_maj)
    feat_test_s = scaler.transform(feat_test)

    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)
    elif occ_type == "LOF":
        k = min(20, n_neighbors_cap)
        clf = LocalOutlierFactor(n_neighbors=k, novelty=True, contamination=0.1)
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)
    else:
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)

    threshold = np.percentile(scores_maj, 90)
    y_pred    = (scores_test >= threshold).astype(int)

    try:
        auc = roc_auc_score(y_test, scores_test) if len(np.unique(y_test)) >= 2 else float("nan")
    except Exception:
        auc = float("nan")

    f1  = f1_score(y_test, y_pred, pos_label=1, zero_division=0)
    rec = recall_score(y_test, y_pred, pos_label=1, zero_division=0)
    gm  = gmean_score(y_test, y_pred)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── KEEL .dat 解析（與 B 一致）──────────────────
def parse_keel_dat(filepath, minority_label=None):
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"): continue
        if s.lower() == "@data":
            data_start = True; continue
        if data_start:
            rows.append(s)
    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    y_raw = df[label_col].astype(str).str.strip().values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    if minority_label is None:
        unique, counts = np.unique(y_raw, return_counts=True)
        minority_label = unique[np.argmin(counts)]

    y = (y_raw == minority_label).astype(int)
    return X, y, minority_label


# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment():
    """
    執行順序：
      dataset → fold → FW → config → OCC

    為什麼 FW 在外、config 在內？
      因為 FW 改變 AE 的 input → 每個 FW 都要重新訓練全部 21 configs 的 DAE。
      把 FW 放外層較容易追蹤進度，且 print log 比較清楚。
    """
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    param_configs = [(nl, rl) for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]
    all_records   = []

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n{'='*68}")
        print(f"▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            file_prefix = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]
            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)

            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案"); continue

            try:
                X_tra, y_tra, minority_label = parse_keel_dat(tra_file)
                X_tst, y_tst, _ = parse_keel_dat(tst_file, minority_label=minority_label)
                input_dim = X_tra.shape[1]
                X_maj = X_tra[y_tra == 0]

                if len(X_maj) < 5:
                    print(f"  [SKIP] Fold {fold}: 訓練集正常樣本不足 ({len(X_maj)})")
                    continue
                if y_tst.sum() == 0:
                    print(f"  [SKIP] Fold {fold}: 測試集無少數類樣本")
                    continue

                # MinMax fit 只用 training majority
                scaler   = MinMaxScaler()
                X_maj_s  = scaler.fit_transform(X_maj)
                X_tst_s  = scaler.transform(X_tst)
                n_nb_cap = max(1, len(X_maj) - 1)

            except Exception as e:
                print(f"  [ERROR] Fold {fold} 資料載入失敗: {e}")
                continue

            for fw in FW_METHODS:
                # === Phase 1 核心步驟：在 OF 端套用 weighting ===
                try:
                    X_maj_w, X_tst_w, w_used = apply_weights(X_maj_s, X_tst_s, fw)
                except Exception as e:
                    print(f"  [ERROR] FW={fw} Fold{fold} weighting 失敗: {e}")
                    continue

                # 對每個 config 訓練一次 DAE（DAE 看到的是 weighted OF）
                feat_cache = {}
                for n_layers, ratio_label in param_configs:
                    ratio   = BOTTLENECK_RATIOS[ratio_label]
                    n_units = max(2, round(input_dim * ratio))
                    try:
                        feat_maj, feat_tst = train_and_extract_dae(
                            X_maj_w, X_tst_w, n_layers, n_units)
                        feat_cache[(n_layers, ratio_label)] = (feat_maj, feat_tst)
                    except Exception as e:
                        print(f"  [ERROR] FW={fw} h{n_layers}-{ratio_label} "
                              f"Fold{fold}: DAE 失敗 {e}")

                # 對每個 OCC × config 計算 metrics
                for occ_type in OCC_TYPES:
                    for (n_layers, ratio_label), (feat_maj, feat_tst) in feat_cache.items():
                        cfg_label = f"h{n_layers}-{ratio_label}"
                        try:
                            metrics = run_occ_eval(
                                occ_type, feat_maj, feat_tst, y_tst, n_nb_cap)
                        except Exception as e:
                            print(f"  [ERROR] FW={fw} {occ_type}({cfg_label}) "
                                  f"Fold{fold}: {e}")
                            metrics = {m: float("nan") for m in METRIC_COLS}

                        all_records.append({
                            "Dataset": ds_name,
                            "AE":      AE_TYPE,        # 固定 DAE
                            "FW":      fw,
                            "OCC":     occ_type,
                            "Config":  cfg_label,
                            "Fold":    fold,
                            **metrics,
                        })

                print(f"  [fold {fold}] FW={fw:5s}  ✓")

    df_all = pd.DataFrame(all_records)

    # ── df_best：per-dataset 選法（與 B/F 一致，避免 per-fold leakage）──
    # 對每個 (Dataset, FW, OCC)，從 21 個 config 中選 5-fold 平均 AUC 最高者，
    # 保留該 config 全部 5 fold 紀錄。
    df_clean = df_all.dropna(subset=["AUC"])
    best_cfg = (
        df_clean.groupby(["Dataset", "FW", "OCC", "Config"])["AUC"]
                .mean().reset_index()
                .sort_values("AUC", ascending=False)
                .drop_duplicates(["Dataset", "FW", "OCC"])
    )
    df_best = df_clean.merge(
        best_cfg[["Dataset", "FW", "OCC", "Config"]],
        on=["Dataset", "FW", "OCC", "Config"],
    )

    print("\n── Best Config 列表 ──")
    for (ds_name, fw, occ_type), grp in df_best.groupby(["Dataset", "FW", "OCC"]):
        cfg      = grp["Config"].iloc[0]
        auc_mean = grp["AUC"].mean()
        print(f"  [{ds_name}] FW={fw:5s} × {occ_type:8s}  best={cfg:12s}  "
              f"5-fold avg AUC={auc_mean:.4f}")

    return df_all, df_best


# ─────────────────────────── Excel 樣式 ──────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

FW_FILL = {
    "none": PatternFill("solid", fgColor="EEEEEE"),
    "var":  PatternFill("solid", fgColor="DAEEF3"),
    "ivar": PatternFill("solid", fgColor="E2EFDA"),
    "mad":  PatternFill("solid", fgColor="FFF2CC"),
    "lap":  PatternFill("solid", fgColor="FCE4D6"),
}

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def sc(cell, value, font=None, fill=None, align=None, fmt=None):
    cell.value  = value
    cell.border = THIN_BORDER
    if font:  cell.font          = font
    if fill:  cell.fill          = fill
    if align: cell.alignment     = align
    if fmt:   cell.number_format = fmt


def col_w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ─────────────────────────── per_fold 分頁 ───────────────────────────────────
def write_per_fold(ws, df, title):
    ws.title = title
    headers = ["Dataset", "AE", "FW", "OCC", "Config", "Fold"] + METRIC_COLS

    for c, h in enumerate(headers, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        fill = FW_FILL.get(row["FW"])
        for c, col in enumerate(headers, 1):
            sc(ws.cell(r, c), row[col],
               font=BODY_FONT, fill=fill,
               align=LEFT_ALIGN if c <= 2 else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS else None)

    for i, w in enumerate([28, 6, 6, 10, 14, 6] + [10] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


# ─────────────────────────── summary 分頁 ────────────────────────────────────
def write_summary_all(ws, df):
    """all_summary：Dataset | FW × OCC × Config × Metric"""
    ws.title = "all_summary"
    configs = ALL_CONFIGS

    grouped = (df.groupby(["Dataset", "FW", "OCC", "Config"])[METRIC_COLS]
                 .agg(["mean", "std"])
                 .reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for fw in FW_METHODS:
        fw_span = len(OCC_TYPES) * len(configs) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + fw_span - 1)
        sc(ws.cell(1, col), f"FW={fw}",
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=FW_FILL.get(fw), align=CENTER_ALIGN)
        for occ in OCC_TYPES:
            occ_span = len(configs) * len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + occ_span - 1)
            sc(ws.cell(2, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            for cfg in configs:
                cfg_span = len(METRIC_COLS)
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + cfg_span - 1)
                sc(ws.cell(3, col), cfg, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                for metric in METRIC_COLS:
                    sc(ws.cell(4, col), metric, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                    col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 5):
        row_alt = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=row_alt, align=LEFT_ALIGN)
        col = 2
        for fw in FW_METHODS:
            for occ in OCC_TYPES:
                for cfg in configs:
                    sub = grouped[
                        (grouped["Dataset"] == ds) &
                        (grouped["FW"]      == fw) &
                        (grouped["OCC"]     == occ) &
                        (grouped["Config"]  == cfg)
                    ]
                    for metric in METRIC_COLS:
                        try:
                            m = sub[(metric, "mean")].values[0]
                            s = sub[(metric, "std")].values[0]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, col), display,
                           font=BODY_FONT,
                           fill=FW_FILL.get(fw) if row_alt is None else row_alt,
                           align=CENTER_ALIGN)
                        col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(FW_METHODS) * len(OCC_TYPES) * len(configs) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B5"


def write_summary_best(ws, df):
    """best_summary：Dataset | FW × OCC × Metric"""
    ws.title = "best_summary"

    grouped = (df.groupby(["Dataset", "FW", "OCC"])[METRIC_COLS]
                 .agg(["mean", "std"])
                 .reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for fw in FW_METHODS:
        fw_span = len(OCC_TYPES) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + fw_span - 1)
        sc(ws.cell(1, col), f"FW={fw}",
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=FW_FILL.get(fw), align=CENTER_ALIGN)
        for occ in OCC_TYPES:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + len(METRIC_COLS) - 1)
            sc(ws.cell(2, col), occ, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            col += len(METRIC_COLS)

    col = 2
    for fw in FW_METHODS:
        for occ in OCC_TYPES:
            for metric in METRIC_COLS:
                sc(ws.cell(3, col), metric, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 4):
        row_alt = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=row_alt, align=LEFT_ALIGN)
        col = 2
        for fw in FW_METHODS:
            for occ in OCC_TYPES:
                sub = grouped[
                    (grouped["Dataset"] == ds) &
                    (grouped["FW"]      == fw) &
                    (grouped["OCC"]     == occ)
                ]
                for metric in METRIC_COLS:
                    try:
                        m = sub[(metric, "mean")].values[0]
                        s = sub[(metric, "std")].values[0]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, col), display,
                       font=BODY_FONT,
                       fill=FW_FILL.get(fw) if row_alt is None else row_alt,
                       align=CENTER_ALIGN)
                    col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(FW_METHODS) * len(OCC_TYPES) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B4"


# ─────────────────────────── overall 分頁 ────────────────────────────────────
def write_overall_all(ws, df):
    """all_overall：FW × OCC × Config 全域平均"""
    ws.title = "all_overall"

    total_cols = 3 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（all configs, all datasets & folds）- G: OF-side Feature Weighting (DAE)"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["FW", "OCC", "Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["FW", "OCC", "Config"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for fw in FW_METHODS:
        for occ in OCC_TYPES:
            for cfg in ALL_CONFIGS:
                fill = FW_FILL.get(fw)
                sc(ws.cell(r, 1), fw,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 2), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 3), cfg, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                for i, metric in enumerate(METRIC_COLS, 4):
                    try:
                        m = overall.loc[(fw, occ, cfg), (metric, "mean")]
                        s = overall.loc[(fw, occ, cfg), (metric, "std")]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, i), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                r += 1

    for i, w in enumerate([8, 12, 14] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


def write_overall_best(ws, df):
    """best_overall：FW × OCC 全域平均 + 最常選中的 Config"""
    ws.title = "best_overall"

    total_cols = 3 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（best config per dataset）- G: OF-side Feature Weighting (DAE)"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["FW", "OCC", "Most Freq Config"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["FW", "OCC"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for fw in FW_METHODS:
        for occ in OCC_TYPES:
            fill = FW_FILL.get(fw)
            sc(ws.cell(r, 1), fw,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
            sc(ws.cell(r, 2), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            cfg_series = df[(df["FW"] == fw) & (df["OCC"] == occ)]["Config"]
            cfg = cfg_series.mode().iloc[0] if not cfg_series.empty else "N/A"
            sc(ws.cell(r, 3), cfg, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            for i, metric in enumerate(METRIC_COLS, 4):
                try:
                    m = overall.loc[(fw, occ), (metric, "mean")]
                    s = overall.loc[(fw, occ), (metric, "std")]
                    display = f"{m:.4f} ± {s:.4f}"
                except Exception:
                    display = "N/A"
                sc(ws.cell(r, i), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            r += 1

    for i, w in enumerate([8, 12, 18] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


# ─────────────────────────── Excel 存檔 ──────────────────────────────────────
def save_excel(df_all, df_best):
    wb  = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()
    ws6 = wb.create_sheet()

    write_per_fold(    ws1, df_all,  "all_per_fold")
    write_summary_all( ws2, df_all)
    write_overall_all( ws3, df_all)
    write_per_fold(    ws4, df_best, "best_per_fold")
    write_summary_best(ws5, df_best)
    write_overall_best(ws6, df_best)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 68)
    print("實驗二 / Phase 1：G — OF 端 Feature Weighting × DAE")
    print(f"AE        : 固定 {AE_TYPE}")
    print(f"FW 方法   : {FW_METHODS}（含 none 內部 baseline）")
    print(f"OCC       : {OCC_TYPES}")
    print(f"搜尋空間  : n_layers={N_LAYERS_LIST} × bottleneck={list(BOTTLENECK_RATIOS.keys())}")
    print(f"            共 {len(N_LAYERS_LIST) * len(BOTTLENECK_RATIOS)} 種 config")
    print(f"總組合數  : {len(FW_METHODS)} × {len(OCC_TYPES)} × "
          f"{len(N_LAYERS_LIST) * len(BOTTLENECK_RATIOS)} = "
          f"{len(FW_METHODS) * len(OCC_TYPES) * len(N_LAYERS_LIST) * len(BOTTLENECK_RATIOS)} "
          f"per fold")
    print("選擇準則  : AUC 最高")
    print("分頁      : all_per_fold / all_summary / all_overall")
    print("            best_per_fold / best_summary / best_overall")
    print("Pipeline  : OF → MinMax → FW → DAE → DF → MinMax → OCC")
    print("=" * 68)

    df_all, df_best = run_experiment()

    if df_all.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_all, df_best)
        print("\n── Best Config Overall Mean（all datasets, by FW × OCC）──")
        print(df_best.groupby(["FW", "OCC"])[METRIC_COLS].mean().round(4).to_string())
