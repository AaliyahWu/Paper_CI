"""
Study2_undersampling_OF.py
==========================
實驗二（Study Two）／對應老師研究方向圖「OF_maj feature set」這一格：
    "the under-sampling effect on the OF_maj feature set"

== Pipeline（對齊 baseline A = OF → OCC1，無 AE）==
    OF (train: maj+min)
        ──MinMax(fit on uncleaned X_maj)──> X_maj_s / X_min_s / X_tst_s
        ──Under-sampling(在 X_*_s 空間，用 y_tra 標籤，只刪 X_maj_s)──> 乾淨 X_maj_clean_s
        ──OCC(OCSVM / LOF / iForest；不再二次 MinMax)──> prediction
    順序：under-sampling → OCC（OCC 在最後；本格無 AE）。

== 與 baseline A 的對齊 ==
    A：OF → MinMax(fit on X_maj) → OCC（OCSVM / LOF / iForest）（n_neighbors=min(20,cap), contamination=0.1）
    本支只在 A 的 baseline 座標空間中插入「OF 多數類清理」，不在清理後重新 fit MinMax。
    Sampler=none ⇒ 乾淨 OF_maj == 原始 X_maj_s ⇒ 重現 baseline A 的對應 OCC
    （無 AE、無 torch 亂數；iForest 用固定 random_state，LOF 確定性）。

== 參數一致性（與 A/B/C 對齊）==
    MinMaxScaler fit on majority；LOF n_neighbors=min(20, n_neighbors_cap),
    novelty=True, contamination=0.1；threshold = 訓練多數類分數第 90 百分位。
    OCC 跑完整三種：OCSVM、LOF、iForest，方便與 A/B/C 對應 OCC 做完整比較。

== Under-sampling 方法（只用 train fold 標籤計算）==
    none / ENN(n_neighbors=3,kind_sel="all") / CNN(n_neighbors=1,random_state=42) / TL
    sampling_strategy="auto"（只刪 majority，完整保留 minority）。

退化保護：清理後 OF_maj < 5 → 跳過該 (Sampler, fold)；清理後重算 LOF k 上限。

本格無 AE、無 config grid，故 best-config 概念不適用；輸出分頁（比照 baseline A + A~K 統整 export）：
    per_fold / summary（Dataset × Sampler × Metric）/ overall（Sampler × Metric）
輸出：results/Study2_undersampling_OF_3OCC.xlsx
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

from imblearn.under_sampling import (
    EditedNearestNeighbours, CondensedNearestNeighbour, TomekLinks,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT   = Path("data")
RESULTS_DIR = Path("results")
OUTPUT_FILE = RESULTS_DIR / "Study2_undersampling_OF_3OCC.xlsx"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5

OCC_TYPES   = ["OCSVM", "LOF", "iForest"]
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

SAMPLERS  = ["none", "ENN", "CNN", "TL"]
ENN_K     = 3
CNN_K     = 1
CNN_SEED  = 42

# ── A~K 統整與參數對齊 metadata（不影響實驗計算）──
STUDY_ID           = "I"
METHOD_ID          = "I_OF_US"
FEATURE_SET        = "OF_maj"
AE_LABEL           = "N/A"
CONFIG_LABEL       = "N/A"
BASELINE_REF       = "A_matched_OCC"
CONFIG_POLICY      = "no_config"
OCC_SCOPE          = "all_three_occ"
SAMPLER_SCALE_MODE = "baseline_majority_minmax_once_before_sampler_and_occ"

COMPARISON_EXPORT_COLS = [
    "Study", "Method", "FeatureSet", "Dataset", "AE", "Sampler", "OCC",
    "Config", "Fold", "ConfigPolicy", "MajKept", "MajRemoved", "RemovedRate",
    "SamplerStatus", "BaselineRef", "OCCScope", "SamplerScaleMode",
] + METRIC_COLS


def safe_removed_rate(n_removed, n_kept):
    """回傳 majority 刪除比例；避免 0 除錯，方便後續分析 sampler 影響。"""
    denom = int(n_removed) + int(n_kept)
    return float(n_removed / denom) if denom > 0 else 0.0


# ─────────────────────────── Under-sampling（在 OF 空間）─────────────────────
def make_sampler(name):
    if name == "ENN":
        return EditedNearestNeighbours(
            n_neighbors=ENN_K, kind_sel="all", sampling_strategy="auto")
    if name == "CNN":
        return CondensedNearestNeighbour(
            n_neighbors=CNN_K, random_state=CNN_SEED, sampling_strategy="auto")
    if name == "TL":
        return TomekLinks(sampling_strategy="auto")
    return None


def clean_majority(X_maj_s, X_min_s, sampler_name):
    """在【已依 baseline majority fit 完成 MinMax 的 OF 空間】清理 majority。

    重點：
      1. 外部先用 uncleaned X_maj fit MinMax，得到 X_maj_s / X_min_s / X_tst_s。
      2. Sampler 直接在這個 baseline 座標尺度上做距離判斷。
      3. 回傳的 X_maj_clean_s 已是可直接送進 OCC 的 scaled 特徵；OCC 內不可再二次 fit MinMax。

    回傳：(X_maj_clean_s, n_removed, sampler_status)
    """
    if sampler_name == "none":
        return X_maj_s, 0, "none_baseline"

    n_maj = len(X_maj_s)
    X_all_s = np.vstack([X_maj_s, X_min_s])
    y_all = np.array([0] * n_maj + [1] * len(X_min_s))

    sampler = make_sampler(sampler_name)
    if sampler is None:
        return X_maj_s, 0, "fallback_unknown_sampler"
    sampler.fit_resample(X_all_s, y_all)
    idx = sampler.sample_indices_

    if int(np.sum(idx >= n_maj)) != len(X_min_s):
        return X_maj_s, 0, "fallback_minority_changed"  # auto 理論上只刪 majority

    keep_maj_local = idx[idx < n_maj]
    X_maj_clean_s = X_maj_s[keep_maj_local]
    n_removed = n_maj - len(X_maj_clean_s)
    status = "ok_removed" if n_removed > 0 else "ok_no_removed"
    return X_maj_clean_s, n_removed, status


# ─────────────────────────── 評估指標（與 A/B 一致）──────────────────────────
def gmean_score(y_true, y_pred_binary):
    cm = confusion_matrix(y_true, y_pred_binary, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return float(np.sqrt(sens * spec))
    return 0.0


def run_occ_eval(occ_type, X_maj_s, X_test_s, y_test, n_neighbors_cap):
    """OCC 評估：輸入必須已在 baseline majority scaler 座標下，不再二次 MinMax。"""
    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(X_maj_s)
        scores_maj  = -clf.decision_function(X_maj_s)
        scores_test = -clf.decision_function(X_test_s)
    elif occ_type == "LOF":
        k = min(20, n_neighbors_cap)
        clf = LocalOutlierFactor(n_neighbors=k, novelty=True, contamination=0.1)
        clf.fit(X_maj_s)
        scores_maj  = -clf.decision_function(X_maj_s)
        scores_test = -clf.decision_function(X_test_s)
    else:
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(X_maj_s)
        scores_maj  = -clf.decision_function(X_maj_s)
        scores_test = -clf.decision_function(X_test_s)

    threshold = np.percentile(scores_maj, 90)
    y_pred    = (scores_test >= threshold).astype(int)

    try:
        auc = roc_auc_score(y_test, scores_test) if len(np.unique(y_test)) >= 2 else float("nan")
    except Exception:
        auc = float("nan")

    f1  = f1_score(y_test, y_pred, pos_label=1, zero_division=0)
    rec = recall_score(y_test, y_pred, pos_label=1, zero_division=0)
    gm  = gmean_score(y_test, y_pred)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── KEEL .dat 解析（與 A/B 一致）────────────────────
def parse_keel_dat(filepath, minority_label=None):
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"):
            continue
        if s.lower() == "@data":
            data_start = True
            continue
        if data_start:
            rows.append(s)
    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    y_raw = df[label_col].astype(str).str.strip().values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    if minority_label is None:
        unique, counts = np.unique(y_raw, return_counts=True)
        minority_label = unique[np.argmin(counts)]

    y = (y_raw == minority_label).astype(int)
    return X, y, minority_label


# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment():
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    all_records = []

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n{'='*68}")
        print(f"▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            file_prefix = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]
            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)
            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案")
                continue

            try:
                X_tra, y_tra, minority_label = parse_keel_dat(tra_file)
                X_tst, y_tst, _ = parse_keel_dat(tst_file, minority_label=minority_label)

                if int(np.sum(y_tra == 0)) < 5:
                    print(f"  [SKIP] Fold {fold}: 訓練集正常樣本不足")
                    continue
                if int(np.sum(y_tra == 1)) < 1:
                    print(f"  [SKIP] Fold {fold}: 訓練集無少數類（sampler 需參考點）")
                    continue
                if y_tst.sum() == 0:
                    print(f"  [SKIP] Fold {fold}: 測試集無少數類樣本")
                    continue

                X_maj = X_tra[y_tra == 0]
                X_min = X_tra[y_tra == 1]
                scaler_of = MinMaxScaler().fit(X_maj)   # baseline A 的唯一座標基準
                X_maj_s = scaler_of.transform(X_maj)
                X_min_s = scaler_of.transform(X_min)
                X_tst_s = scaler_of.transform(X_tst)
            except Exception as e:
                print(f"  [ERROR] Fold {fold} 資料載入失敗: {e}")
                continue

            for sampler_name in SAMPLERS:
                try:
                    X_maj_clean, n_removed, sampler_status = clean_majority(X_maj_s, X_min_s, sampler_name)
                except Exception as e:
                    print(f"  [ERROR] Fold{fold} Sampler={sampler_name}: 清理失敗 {e}")
                    continue

                if len(X_maj_clean) < 5:
                    print(f"  [SKIP] Fold{fold} Sampler={sampler_name}: "
                          f"清理後 OF_maj 不足 ({len(X_maj_clean)})")
                    continue

                n_nb_cap = max(1, len(X_maj_clean) - 1)

                for occ_type in OCC_TYPES:
                    try:
                        metrics = run_occ_eval(
                            occ_type, X_maj_clean, X_tst_s, y_tst, n_nb_cap)
                    except Exception as e:
                        print(f"  [ERROR] Fold{fold} Sampler={sampler_name} "
                              f"{occ_type}: {e}")
                        metrics = {m: float("nan") for m in METRIC_COLS}

                    all_records.append({
                        "Dataset":    ds_name,
                        "Sampler":    sampler_name,
                        "OCC":        occ_type,
                        "Fold":       fold,
                        "MajKept":       len(X_maj_clean),
                        "MajRemoved":    n_removed,
                        "RemovedRate":   safe_removed_rate(n_removed, len(X_maj_clean)),
                        "SamplerStatus": sampler_status,
                        "BaselineCheck": f"A_{occ_type}" if sampler_name == "none" else "",
                        **metrics,
                    })

            print(f"  [fold {fold}] 完成 {len(SAMPLERS)} Sampler × {len(OCC_TYPES)} OCC")

    return pd.DataFrame(all_records)


# ─────────────────────────── Excel 樣式（與 A/B 共用）────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

SAMPLER_FILL = {
    "none": PatternFill("solid", fgColor="EEEEEE"),
    "ENN":  PatternFill("solid", fgColor="DAEEF3"),
    "CNN":  PatternFill("solid", fgColor="FCE4D6"),
    "TL":   PatternFill("solid", fgColor="E2EFDA"),
}

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def sc(cell, value, font=None, fill=None, align=None, fmt=None):
    cell.value  = value
    cell.border = THIN_BORDER
    if font:  cell.font          = font
    if fill:  cell.fill          = fill
    if align: cell.alignment     = align
    if fmt:   cell.number_format = fmt


def col_w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ─────────────────────────── A~K 統整友善輸出 ────────────────────────────────
def make_ak_export_df(df, config_policy=CONFIG_POLICY):
    """建立固定欄位的 A~K 統整用資料表。

    目的：後續統整 A~K 時只要讀取 ak_*_export 分頁並用欄位名稱抓值，
    不會受到 per_fold / summary 額外 metadata 欄位增減影響。
    """
    if df.empty:
        return pd.DataFrame(columns=COMPARISON_EXPORT_COLS)

    out = pd.DataFrame({
        "Study": STUDY_ID,
        "Method": METHOD_ID,
        "FeatureSet": FEATURE_SET,
        "Dataset": df["Dataset"],
        "AE": AE_LABEL,
        "Sampler": df["Sampler"],
        "OCC": df["OCC"],
        "Config": CONFIG_LABEL,
        "Fold": df["Fold"],
        "ConfigPolicy": config_policy,
        "MajKept": df["MajKept"],
        "MajRemoved": df["MajRemoved"],
        "RemovedRate": df["RemovedRate"],
        "SamplerStatus": df["SamplerStatus"],
        "BaselineRef": df["BaselineCheck"],
        "OCCScope": OCC_SCOPE,
        "SamplerScaleMode": SAMPLER_SCALE_MODE,
    })
    for metric in METRIC_COLS:
        out[metric] = df[metric]
    return out[COMPARISON_EXPORT_COLS]


def write_ak_export(ws, df, title="ak_all_export", config_policy=CONFIG_POLICY):
    ws.title = title
    out = make_ak_export_df(df, config_policy=config_policy)
    for c, h in enumerate(COMPARISON_EXPORT_COLS, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    for r, (_, row) in enumerate(out.iterrows(), 2):
        fill = SAMPLER_FILL.get(row["Sampler"])
        for c, col in enumerate(COMPARISON_EXPORT_COLS, 1):
            sc(ws.cell(r, c), row[col], font=BODY_FONT, fill=fill,
               align=LEFT_ALIGN if col in ["Dataset", "SamplerScaleMode"] else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS + ["RemovedRate"] else None)
    widths = [8, 14, 14, 28, 8, 10, 8, 14, 8, 22, 10, 12, 12, 24, 14, 12, 42] + [10] * len(METRIC_COLS)
    for i, w in enumerate(widths, 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


def write_alignment_notes(ws):
    """記錄此檔與 baseline 的對齊關係，方便論文與 A~K 統整追溯。"""
    ws.title = "alignment_notes"
    notes = [
        ("Study", STUDY_ID),
        ("Method", METHOD_ID),
        ("FeatureSet", FEATURE_SET),
        ("BaselineRef", BASELINE_REF),
        ("AE", AE_LABEL),
        ("Config", CONFIG_LABEL),
        ("OCCScope", OCC_SCOPE),
        ("OCC", ", ".join(OCC_TYPES)),
        ("Sampler", ", ".join(SAMPLERS)),
        ("SamplerScaleMode", SAMPLER_SCALE_MODE),
        ("None sanity check", "Sampler=none + each OCC should match the corresponding Baseline A OCC under the same data/fold."),
        ("Threshold", "90th percentile of training majority anomaly scores."),
        ("LOF n_neighbors", "min(20, len(cleaned majority)-1)."),
        ("Contamination", "0.1"),
        ("Data leakage guard", "The single scaler is fit on uncleaned training majority only; sampler uses train fold only; test is transformed only."),
    ]
    sc(ws.cell(1, 1), "Item", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    sc(ws.cell(1, 2), "Value", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    for r, (k, v) in enumerate(notes, 2):
        sc(ws.cell(r, 1), k, font=BOLD_FONT, fill=ALT_FILL if r % 2 == 0 else None, align=LEFT_ALIGN)
        sc(ws.cell(r, 2), v, font=BODY_FONT, fill=ALT_FILL if r % 2 == 0 else None, align=LEFT_ALIGN)
    col_w(ws, "A", 24)
    col_w(ws, "B", 100)
    ws.freeze_panes = "A2"


def write_per_fold(ws, df):
    ws.title = "per_fold"
    headers = ["Dataset", "Sampler", "OCC", "Fold", "MajKept", "MajRemoved",
               "RemovedRate", "SamplerStatus", "BaselineCheck"] + METRIC_COLS
    for c, h in enumerate(headers, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        fill = SAMPLER_FILL.get(row["Sampler"])
        for c, col in enumerate(headers, 1):
            sc(ws.cell(r, c), row[col], font=BODY_FONT, fill=fill,
               align=LEFT_ALIGN if c == 1 else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS + ["RemovedRate"] else None)
    for i, w in enumerate([28, 9, 8, 6, 9, 11, 12, 24, 16] + [10] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


def write_summary(ws, df):
    """summary：Dataset | Sampler × Metric（mean ± std over folds）"""
    ws.title = "summary"
    grouped = (df.groupby(["Dataset", "Sampler"])[METRIC_COLS]
                 .agg(["mean", "std"]).reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for sp in SAMPLERS:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(METRIC_COLS) - 1)
        sc(ws.cell(1, col), f"US={sp}",
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=SAMPLER_FILL.get(sp), align=CENTER_ALIGN)
        for metric in METRIC_COLS:
            sc(ws.cell(2, col), metric, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 3):
        row_alt = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=row_alt, align=LEFT_ALIGN)
        col = 2
        for sp in SAMPLERS:
            sub = grouped[(grouped["Dataset"] == ds) & (grouped["Sampler"] == sp)]
            for metric in METRIC_COLS:
                try:
                    m = sub[(metric, "mean")].values[0]
                    s = sub[(metric, "std")].values[0]
                    display = f"{m:.4f} ± {s:.4f}"
                except Exception:
                    display = "N/A"
                sc(ws.cell(r, col), display, font=BODY_FONT,
                   fill=SAMPLER_FILL.get(sp) if row_alt is None else row_alt,
                   align=CENTER_ALIGN)
                col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(SAMPLERS) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B3"


def write_overall(ws, df):
    """overall：Sampler × OCC 全域平均 + 平均刪除量（核心比較表）"""
    ws.title = "overall"
    total_cols = 4 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean - Study2 OF-side Under-sampling (no AE, 3 OCC)"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["Sampler", "OCC", "Avg Removed", "Avg Removed Rate"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["Sampler", "OCC"])[METRIC_COLS].agg(["mean", "std"])
    r = 3
    for sp in SAMPLERS:
        for occ in OCC_TYPES:
            fill = SAMPLER_FILL.get(sp)
            sub = df[(df["Sampler"] == sp) & (df["OCC"] == occ)]
            avg_rm = sub["MajRemoved"].mean() if not sub.empty else 0.0
            avg_rate = sub["RemovedRate"].mean() if not sub.empty else 0.0
            sc(ws.cell(r, 1), sp,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
            sc(ws.cell(r, 2), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            sc(ws.cell(r, 3), f"{avg_rm:.1f}", font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            sc(ws.cell(r, 4), avg_rate, font=BODY_FONT, fill=fill, align=CENTER_ALIGN, fmt="0.0000")
            for i, metric in enumerate(METRIC_COLS, 5):
                try:
                    m = overall.loc[(sp, occ), (metric, "mean")]
                    s = overall.loc[(sp, occ), (metric, "std")]
                    display = f"{m:.4f} ± {s:.4f}"
                except Exception:
                    display = "N/A"
                sc(ws.cell(r, i), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
            r += 1

    for i, w in enumerate([9, 8, 12, 16] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


def save_excel(df):
    wb  = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()
    write_per_fold(ws1, df)
    write_summary(ws2, df)
    write_overall(ws3, df)
    write_ak_export(ws4, df, title="ak_all_export", config_policy=CONFIG_POLICY)
    # I 無 config grid，best 與 all 相同；保留 ak_best_export 方便 A~K 統一讀取。
    write_ak_export(ws5, df, title="ak_best_export", config_policy=CONFIG_POLICY)
    write_alignment_notes(wb.create_sheet())
    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 68)
    print("Study Two（OF_maj feature set）：OF 側 Under-sampling × 3 OCC（無 AE）")
    print(f"Sampler   : {SAMPLERS}（含 none = baseline A）")
    print(f"OCC       : {OCC_TYPES}")
    print("Pipeline  : OF → MinMax(fit uncleaned maj) → Under-sampling → OCC(OCSVM/LOF/iForest, no second MinMax)")
    print("=" * 68)

    df_results = run_experiment()

    if df_results.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_results)
        print("\n── Overall Mean（by Sampler）──")
        print(df_results.groupby("Sampler")[METRIC_COLS].mean().round(4).to_string())
