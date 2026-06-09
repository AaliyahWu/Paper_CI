"""
Study2_undersampling_DF_4AE.py
==============================
實驗二（Study Two）／對應老師研究方向圖「DF_maj feature set」這一格：
    "the under-sampling effect on the DF_maj feature set"

== 與上一版（I_baseline_B_undersampling_OF.py）的關鍵差異 ==
    上一版：under-sampling 作用在 OF（原始特徵）→ 再進 AE → OCC
            ＝「OF 側清理」，對應的是「OF_maj feature set」那一格。
    本  版：under-sampling 作用在 DF（深度特徵）→ OCC
            ＝「DF 側清理」，才符合老師圖中「DF_maj feature set」的規劃。

== Pipeline（DF 側清理，對齊 baseline B = DF → OCC2）==
    OF (train: maj+min)
        ──MinMax(fit on X_maj)──> X_maj_s / X_min_s / X_tst_s
        ──AE 訓練(只用 X_maj_s)──> 抽取 DF_maj / DF_min / DF_tst
        ──MinMax(fit on uncleaned DF_maj)──> DF_maj_s / DF_min_s / DF_tst_s
        ──在 DF_s 空間用標籤 under-sample(只刪 DF_maj_s)──> 乾淨 DF_maj_clean_s
        ──OCC(OCSVM / LOF / iForest；不再二次 MinMax)──> prediction

順序釐清（回答「是不是 under-sampling → OCC → AE？」）：
    不是。OCC 永遠在最後。正確順序是
        AE(抽 DF) → under-sampling(在 DF 上清理 majority) → OCC
    under-sampling 夾在「特徵抽取之後、OCC 之前」，作用在要丟進 OCC 的那組特徵。

== 概念釐清：under-sampling 與 OCC 的相容性 ==
    ENN/CNN/TL 都需要兩類標籤才能運作（靠 minority 的位置判斷該刪哪些 majority）。
    OCC 只用 majority 訓練，訓練 fold 的 minority 原本丟掉不用。
    本實驗把 minority 當「參考點」：不拿來訓練 OCC，而是用來決定哪些 DF_maj
    （normal 的深度特徵）是雜訊／邊界／與異常重疊 → 移除它們，讓 one-class
    邊界更乾淨。test set 全程不碰，無 leakage。

== 4 個 AE 全跑（本次需求）==
    AE_TYPES = ["AE", "DAE", "SAE", "VAE"]，與 baseline B/C 一致。
    效率：因為清理發生在 AE 之後，AE 的 input（未清理 majority）不受 sampler 影響，
          所以每個 (AE, config) 的 AE 只訓練一次，sampler 只是刪列（很便宜）。
          AE 訓練數 = 4 AE × 21 config × 5 fold per dataset（與 B 相同）。

== Under-sampling 方法（只用 train fold 標籤計算）==
    • none : 不清理（內部 baseline；應重現 baseline B 的對應 AE×LOF 數值）
    • ENN  : Edited Nearest Neighbours（Wilson 1972）—刪掉落在異常鄰域中的 normal（邊界雜訊）
    • CNN  : Condensed Nearest Neighbour（Hart 1968）—留邊界、丟內部冗餘；可能傷害密度型 LOF
    • TL   : Tomek Links（Tomek 1976）—只清最糾纏的重疊點，最溫和

Sampler 超參數（固定，可重現）：
    ENN n_neighbors=3, kind_sel="all" / CNN n_neighbors=1, random_state=42 / TL 預設
    sampling_strategy="auto"（只從 majority 移除，完整保留 minority）

退化保護：清理後 DF_maj < 5 → 跳過該 (AE, config, sampler, fold)；
         清理後重算 n_neighbors_cap，避免 LOF k 超界（CNN 尤需）。

搜尋空間：4 AE × 4 Sampler × 1 OCC(LOF) × 21 config = 336 組 per fold
best 選法：per-(Dataset, AE, Sampler, OCC) 取 5-fold 平均 AUC 最高的 config，
          保留該 config 全部 5 fold（per-dataset 選法，與 B/F/G/H 一致）。

輸出：results/Study2_undersampling_DF_4AE_3OCC.xlsx（6 分頁，與 B 對齊，多 Sampler 維度）

備註：這只是 Study Two 三格中的「DF_maj」。完整 Study Two 還需要：
      • OF_maj  ：OF → under-sample → OCC（無 AE，對應 baseline A）
      • OF+DF   ：AE→DF、串接 [OF,DF]、在串接上 under-sample → OCC（對應 baseline C）
      這兩支結構不同（OF 那支無 AE/config），建議比照 A/C 各自獨立成檔。
"""

import re
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import OneClassSVM
from sklearn.neighbors import LocalOutlierFactor
from sklearn.ensemble import IsolationForest
from sklearn.metrics import roc_auc_score, f1_score, recall_score, confusion_matrix

# imbalanced-learn（需 pip install imbalanced-learn）。注意 0.14 是 CondensedNearestNeighbour（單數）
from imblearn.under_sampling import (
    EditedNearestNeighbours, CondensedNearestNeighbour, TomekLinks,
)

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
torch.manual_seed(42)
np.random.seed(42)

# ─────────────────────────── 路徑設定 ────────────────────────────────────────
DATA_ROOT   = Path("data")
RESULTS_DIR = Path("results")
OUTPUT_FILE = RESULTS_DIR / "Study2_undersampling_DF_4AE_3OCC.xlsx"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

N_FOLDS = 5
EPS     = 1e-12

# ── AE 訓練超參數（與 B/C 完全一致）──
AE_EPOCHS     = 100
AE_BATCH_SIZE = 64
AE_LR         = 1e-3
DAE_NOISE     = 0.1
SAE_SPARSITY  = 1e-3
VAE_BETA      = 1.0

# ── 4 個 AE 全跑（本次需求；與 baseline B/C 一致）──
AE_TYPES    = ["AE", "DAE", "SAE", "VAE"]

# ── OCC 跑完整三種：OCSVM、LOF、iForest，方便與 baseline 做完整對齊 ──
# OCC_TYPES = ["OCSVM", "LOF", "iForest"]
OCC_TYPES   = ["OCSVM", "LOF", "iForest"]
METRIC_COLS = ["AUC", "F1", "Recall", "G-mean"]

# ── Under-sampling 方法（含 none 內部 baseline）──
SAMPLERS  = ["none", "ENN", "CNN", "TL"]
ENN_K     = 3
CNN_K     = 1
CNN_SEED  = 42

# ── A~K 統整與參數對齊 metadata（不影響實驗計算）──
STUDY_ID           = "J"
METHOD_ID          = "J_DF_US"
FEATURE_SET        = "DF_maj"
BASELINE_REF       = "B_matched_OCC"
OCC_SCOPE          = "all_three_occ"
SAMPLER_SCALE_MODE = "df_majority_minmax_once_before_sampler_and_occ"

COMPARISON_EXPORT_COLS = [
    "Study", "Method", "FeatureSet", "Dataset", "AE", "Sampler", "OCC",
    "Config", "Fold", "ConfigPolicy", "MajKept", "MajRemoved", "RemovedRate",
    "SamplerStatus", "BaselineRef", "OCCScope", "SamplerScaleMode",
] + METRIC_COLS


def safe_removed_rate(n_removed, n_kept):
    """回傳 majority 刪除比例；避免 0 除錯，方便後續分析 sampler 影響。"""
    denom = int(n_removed) + int(n_kept)
    return float(n_removed / denom) if denom > 0 else 0.0


# ── Grid（與 B/C/G/H 對齊）──
N_LAYERS_LIST     = [1, 2, 3]
BOTTLENECK_RATIOS = {
    "1/4": 0.25, "1/3": 1/3, "1/2": 0.5, "1/1": 1.0,
    "2/1": 2.0,  "3/1": 3.0, "4/1": 4.0,
}
ALL_CONFIGS = [f"h{nl}-{rl}" for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]


# ─────────────────────────── Under-sampling（在 DF 空間）─────────────────────
def make_sampler(name):
    if name == "ENN":
        return EditedNearestNeighbours(
            n_neighbors=ENN_K, kind_sel="all", sampling_strategy="auto")
    if name == "CNN":
        return CondensedNearestNeighbour(
            n_neighbors=CNN_K, random_state=CNN_SEED, sampling_strategy="auto")
    if name == "TL":
        return TomekLinks(sampling_strategy="auto")
    return None


def undersample_df(DF_maj_s, DF_min_s, sampler_name):
    """在【已依 uncleaned DF_maj fit 完成 MinMax 的 DF 空間】清理 majority。

    重點：
      1. 外部先以未清理的 DF_maj fit MinMax，得到 DF_maj_s / DF_min_s / DF_tst_s。
      2. Sampler 直接在 baseline B 的 DF 座標尺度上做距離判斷。
      3. 回傳的 DF_maj_clean_s 已可直接送進 OCC；OCC 內不可再二次 fit MinMax。

    回傳：(DF_maj_clean_s, n_removed, sampler_status)
    """
    if sampler_name == "none":
        return DF_maj_s, 0, "none_baseline"

    n_maj = len(DF_maj_s)
    DF_all_s = np.vstack([DF_maj_s, DF_min_s])
    y_all = np.array([0] * n_maj + [1] * len(DF_min_s))

    sampler = make_sampler(sampler_name)
    if sampler is None:
        return DF_maj_s, 0, "fallback_unknown_sampler"
    sampler.fit_resample(DF_all_s, y_all)
    idx = sampler.sample_indices_

    if int(np.sum(idx >= n_maj)) != len(DF_min_s):
        return DF_maj_s, 0, "fallback_minority_changed"

    keep_maj_local = idx[idx < n_maj]
    DF_maj_clean_s = DF_maj_s[keep_maj_local]
    n_removed = n_maj - len(DF_maj_clean_s)
    status = "ok_removed" if n_removed > 0 else "ok_no_removed"
    return DF_maj_clean_s, n_removed, status


# ─────────────────────────── AE 模型（與 B/C 完全一致）──────────────────────
class AEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        enc, dec = [], []
        for i in range(len(dims) - 1):
            enc += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        for i in range(len(dims) - 1):
            d_in, d_out = dims[-(i+1)], dims[-(i+2)]
            act = nn.Sigmoid() if i == len(dims) - 2 else nn.ReLU()
            dec += [nn.Linear(d_in, d_out), act]
        self.encoder = nn.Sequential(*enc)
        self.decoder = nn.Sequential(*dec)

    def forward(self, x):
        z = self.encoder(x)
        return self.decoder(z), z


class VAEModel(nn.Module):
    def __init__(self, input_dim, n_layers, n_units):
        super().__init__()
        dims = [input_dim] + [n_units] * n_layers
        base = []
        for i in range(len(dims) - 2):
            base += [nn.Linear(dims[i], dims[i+1]), nn.ReLU()]
        self.enc_base  = nn.Sequential(*base) if base else nn.Identity()
        mid = dims[-2] if len(dims) >= 2 else input_dim
        self.fc_mu     = nn.Linear(mid, dims[-1])
        self.fc_logvar = nn.Linear(mid, dims[-1])
        dec_dims = dims[::-1]
        dec = []
        for i in range(len(dec_dims) - 1):
            act = nn.Sigmoid() if i == len(dec_dims) - 2 else nn.ReLU()
            dec += [nn.Linear(dec_dims[i], dec_dims[i+1]), act]
        self.decoder = nn.Sequential(*dec)

    def reparameterize(self, mu, lv):
        return mu + torch.exp(0.5 * lv) * torch.randn_like(lv)

    def forward(self, x):
        h  = self.enc_base(x)
        mu = self.fc_mu(h)
        lv = self.fc_logvar(h)
        z  = self.reparameterize(mu, lv)
        return self.decoder(z), z, mu, lv


def train_ae_and_get_extractor(ae_type, X_maj_s, n_layers, n_units):
    """訓練 AE（只用 majority），回傳 extract 函式（與 B 的 train_and_extract 一致）。

    回傳的 extract(X) 對 VAE 取 μ、其餘取 encoder z（與 B 完全一致）。
    這裡回傳 extractor 而非固定兩個輸出，方便對 maj/min/test 三者都抽 DF。
    """
    input_dim = X_maj_s.shape[1]
    model = VAEModel(input_dim, n_layers, n_units) if ae_type == "VAE" \
            else AEModel(input_dim, n_layers, n_units)

    optim_ = torch.optim.Adam(model.parameters(), lr=AE_LR)
    mse    = nn.MSELoss()
    bs     = min(AE_BATCH_SIZE, len(X_maj_s))
    loader = DataLoader(
        TensorDataset(torch.tensor(X_maj_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True,
    )

    for _ in range(AE_EPOCHS):
        model.train()
        for (xb,) in loader:
            optim_.zero_grad()
            if ae_type == "DAE":
                xb_n  = torch.clamp(xb + DAE_NOISE * torch.randn_like(xb), 0, 1)
                xr, _ = model(xb_n)
                loss  = mse(xr, xb)
            elif ae_type == "SAE":
                xr, z = model(xb)
                loss  = mse(xr, xb) + SAE_SPARSITY * z.abs().mean()
            elif ae_type == "VAE":
                xr, _, mu, lv = model(xb)
                kl   = -0.5 * (1 + lv - mu.pow(2) - lv.exp()).mean()
                loss = mse(xr, xb) + VAE_BETA * kl
            else:  # 標準 AE
                xr, _ = model(xb)
                loss  = mse(xr, xb)
            loss.backward()
            optim_.step()

    model.eval()

    def extract(X):
        xt = torch.tensor(X, dtype=torch.float32)
        with torch.no_grad():
            if ae_type == "VAE":
                _, _, mu, _ = model(xt)
                return mu.numpy()
            else:
                _, z = model(xt)
                return z.numpy()
    return extract


# ─────────────────────────── 評估指標（與 B 一致）──────────────────────────
def gmean_score(y_true, y_pred_binary):
    cm = confusion_matrix(y_true, y_pred_binary, labels=[1, 0])
    if cm.shape == (2, 2):
        tp, fn, fp, tn = cm[0, 0], cm[0, 1], cm[1, 0], cm[1, 1]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0.0
        return float(np.sqrt(sens * spec))
    return 0.0


def run_occ_eval(occ_type, feat_maj_s, feat_test_s, y_test, n_neighbors_cap, do_scale=False):
    """OCC 評估：預設輸入已在 baseline B 的 DF scaler 座標下，不再二次 MinMax。

    do_scale 只保留向後相容；Study2 修正版呼叫時一律使用 do_scale=False。
    """
    if do_scale:
        scaler      = MinMaxScaler()
        feat_maj_s  = scaler.fit_transform(feat_maj_s)
        feat_test_s = scaler.transform(feat_test_s)

    if occ_type == "OCSVM":
        clf = OneClassSVM(nu=0.1, kernel="rbf")
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)
    elif occ_type == "LOF":
        k = min(20, n_neighbors_cap)
        clf = LocalOutlierFactor(n_neighbors=k, novelty=True, contamination=0.1)
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)
    else:
        clf = IsolationForest(n_estimators=100, contamination=0.1, random_state=42)
        clf.fit(feat_maj_s)
        scores_maj  = -clf.decision_function(feat_maj_s)
        scores_test = -clf.decision_function(feat_test_s)

    threshold = np.percentile(scores_maj, 90)
    y_pred    = (scores_test >= threshold).astype(int)

    try:
        auc = roc_auc_score(y_test, scores_test) if len(np.unique(y_test)) >= 2 else float("nan")
    except Exception:
        auc = float("nan")

    f1  = f1_score(y_test, y_pred, pos_label=1, zero_division=0)
    rec = recall_score(y_test, y_pred, pos_label=1, zero_division=0)
    gm  = gmean_score(y_test, y_pred)
    return {"AUC": auc, "F1": f1, "Recall": rec, "G-mean": gm}


# ─────────────────────────── KEEL .dat 解析（與 B 一致）──────────────────────
def parse_keel_dat(filepath, minority_label=None):
    lines = Path(filepath).read_text(encoding="utf-8", errors="replace").splitlines()
    data_start = False
    rows = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("%"):
            continue
        if s.lower() == "@data":
            data_start = True
            continue
        if data_start:
            rows.append(s)
    if not rows:
        raise ValueError(f"No data found in {filepath}")

    records = [[p.strip() for p in r.split(",")] for r in rows]
    df = pd.DataFrame(records)
    label_col = df.columns[-1]
    y_raw = df[label_col].astype(str).str.strip().values

    feat_df = df.iloc[:, :-1].copy()
    for col in feat_df.columns:
        conv = pd.to_numeric(feat_df[col], errors="coerce")
        if conv.isna().all():
            feat_df[col] = pd.Categorical(feat_df[col]).codes.astype(float)
        else:
            feat_df[col] = conv
    X = feat_df.values.astype(float)

    if minority_label is None:
        unique, counts = np.unique(y_raw, return_counts=True)
        minority_label = unique[np.argmin(counts)]

    y = (y_raw == minority_label).astype(int)
    return X, y, minority_label


# ─────────────────────────── 主流程 ──────────────────────────────────────────
def run_experiment():
    """
    執行順序：
      dataset → fold → [scale] → AE → config
              → [AE 訓練一次、抽 DF_maj/DF_min/DF_tst] → sampler → [DF 清理] → OCC

    為什麼 AE 在外、sampler 在內？
      sampler 清理的是 DF（AE 的輸出），不影響 AE 的 input（未清理 majority），
      所以每個 (AE, config) 的 AE 只訓練一次，sampler 只是刪列 → 省時（同 H）。
    """
    dataset_dirs = sorted([d for d in DATA_ROOT.iterdir() if d.is_dir()])
    if not dataset_dirs:
        raise FileNotFoundError(f"找不到任何資料夾於 {DATA_ROOT.resolve()}")

    param_configs = [(nl, rl) for nl in N_LAYERS_LIST for rl in BOTTLENECK_RATIOS]
    all_records   = []

    for ds_dir in dataset_dirs:
        ds_name = ds_dir.name
        print(f"\n{'='*68}")
        print(f"▶ Dataset: {ds_name}")

        for fold in range(1, N_FOLDS + 1):
            file_prefix = re.sub(r'-fold.*$', '', ds_name)
            patterns_tra = [
                ds_dir / f"{file_prefix}-{fold}tra.dat",
                ds_dir / f"{ds_name}-5-fold-tra{fold}.dat",
                ds_dir / f"{ds_name}-5-tra{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_train.dat",
            ]
            patterns_tst = [
                ds_dir / f"{file_prefix}-{fold}tst.dat",
                ds_dir / f"{ds_name}-5-fold-tst{fold}.dat",
                ds_dir / f"{ds_name}-5-tst{fold}.dat",
                ds_dir / f"{ds_name}_fold{fold}_test.dat",
            ]
            tra_file = next((p for p in patterns_tra if p.exists()), None)
            tst_file = next((p for p in patterns_tst if p.exists()), None)
            if tra_file is None or tst_file is None:
                print(f"  [SKIP] Fold {fold}: 找不到檔案")
                continue

            try:
                X_tra, y_tra, minority_label = parse_keel_dat(tra_file)
                X_tst, y_tst, _ = parse_keel_dat(tst_file, minority_label=minority_label)
                input_dim = X_tra.shape[1]
                X_maj = X_tra[y_tra == 0]
                X_min = X_tra[y_tra == 1]

                if len(X_maj) < 5:
                    print(f"  [SKIP] Fold {fold}: 訓練集正常樣本不足 ({len(X_maj)})")
                    continue
                if len(X_min) < 1:
                    print(f"  [SKIP] Fold {fold}: 訓練集無少數類（sampler 需參考點）")
                    continue
                if y_tst.sum() == 0:
                    print(f"  [SKIP] Fold {fold}: 測試集無少數類樣本")
                    continue

                scaler  = MinMaxScaler()
                X_maj_s = scaler.fit_transform(X_maj)         # fit 只用 majority（與 B 一致）
                X_min_s = scaler.transform(X_min)             # minority 僅作 sampler 參考
                X_tst_s = scaler.transform(X_tst)
            except Exception as e:
                print(f"  [ERROR] Fold {fold} 資料載入失敗: {e}")
                continue

            for ae_type in AE_TYPES:
                for n_layers, ratio_label in param_configs:
                    ratio     = BOTTLENECK_RATIOS[ratio_label]
                    n_units   = max(2, round(input_dim * ratio))
                    cfg_label = f"h{n_layers}-{ratio_label}"

                    # AE 只訓練一次，抽 maj/min/test 三組 DF
                    #
                    # ★ RNG 路徑對齊（讓 Sampler=none 精準重現 baseline B）★
                    #   VAE 的 forward 會 sample z（randn_like），故每次 extract()
                    #   對 VAE 會消耗一次 torch 亂數。B 的順序是 extract(maj), extract(test)。
                    #   這裡多抽一個 DF_min 給 sampler 當參考點，若不處理會：
                    #     (1) 讓 DF_tst 拿到錯位的亂數 → none ≠ B；
                    #     (2) 讓後續每個 config 的 AE 權重與 B 漂移。
                    #   解法：先 maj→test（對齊 B），再用 get/set_rng_state 包住 DF_min，
                    #   使「抽 min」對 torch 亂數路徑零影響。
                    try:
                        extract = train_ae_and_get_extractor(
                            ae_type, X_maj_s, n_layers, n_units)
                        DF_maj = extract(X_maj_s)          # 對齊 B：第 1 次 extract
                        DF_tst = extract(X_tst_s)          # 對齊 B：第 2 次 extract
                        _rng_state = torch.get_rng_state()  # 存檔
                        DF_min = extract(X_min_s)          # 額外抽取（給 sampler 用）
                        torch.set_rng_state(_rng_state)    # 還原 → 路徑與 B 一致
                    except Exception as e:
                        print(f"  [ERROR] Fold{fold} {ae_type} {cfg_label}: AE 失敗 {e}")
                        continue

                    # 以 uncleaned DF_maj 作為唯一 MinMax 基準，對齊 baseline B 的 OCC 輸入尺度。
                    scaler_df = MinMaxScaler().fit(DF_maj)
                    DF_maj_s = scaler_df.transform(DF_maj)
                    DF_min_s = scaler_df.transform(DF_min)
                    DF_tst_s = scaler_df.transform(DF_tst)

                    for sampler_name in SAMPLERS:
                        # 在已對齊 baseline B 尺度的 DF 空間清理 majority
                        try:
                            DF_maj_clean, n_removed, sampler_status = undersample_df(
                                DF_maj_s, DF_min_s, sampler_name)
                        except Exception as e:
                            print(f"  [ERROR] Fold{fold} {ae_type} {cfg_label} "
                                  f"Sampler={sampler_name}: 清理失敗 {e}")
                            continue

                        if len(DF_maj_clean) < 5:
                            print(f"  [SKIP] Fold{fold} {ae_type} {cfg_label} "
                                  f"Sampler={sampler_name}: 清理後 DF_maj 不足 "
                                  f"({len(DF_maj_clean)})")
                            continue

                        n_nb_cap = max(1, len(DF_maj_clean) - 1)  # 清理後重算 LOF k 上限

                        for occ_type in OCC_TYPES:
                            try:
                                metrics = run_occ_eval(
                                    occ_type, DF_maj_clean, DF_tst_s, y_tst,
                                    n_nb_cap, do_scale=False)
                            except Exception as e:
                                print(f"  [ERROR] Fold{fold} {ae_type} {cfg_label} "
                                      f"Sampler={sampler_name} {occ_type}: {e}")
                                metrics = {m: float("nan") for m in METRIC_COLS}

                            all_records.append({
                                "Dataset":    ds_name,
                                "AE":         ae_type,
                                "Sampler":    sampler_name,
                                "OCC":        occ_type,
                                "Config":     cfg_label,
                                "Fold":       fold,
                                "MajKept":       len(DF_maj_clean),
                                "MajRemoved":    n_removed,
                                "RemovedRate":   safe_removed_rate(n_removed, len(DF_maj_clean)),
                                "SamplerStatus": sampler_status,
                                "BaselineCheck": f"B_{occ_type}" if sampler_name == "none" else "",
                                **metrics,
                            })

            print(f"  [fold {fold}] 完成 {len(param_configs)} configs × "
                  f"{len(AE_TYPES)} AE × {len(SAMPLERS)} Sampler × {len(OCC_TYPES)} OCC")

    df_all = pd.DataFrame(all_records)

    # ── df_best：per-(Dataset, AE, Sampler, OCC) 選 5-fold 平均 AUC 最高的 config ──
    if df_all.empty or "AUC" not in df_all.columns:
        print("\n⚠️  沒有可用的 AUC 結果，略過 best config 選取。")
        return df_all, pd.DataFrame()

    df_clean = df_all.dropna(subset=["AUC"])
    if df_clean.empty:
        print("\n⚠️  AUC 全為 NaN，略過 best config 選取。")
        return df_all, pd.DataFrame(columns=df_all.columns)

    best_cfg = (
        df_clean.groupby(["Dataset", "AE", "Sampler", "OCC", "Config"])["AUC"]
                .mean().reset_index()
                .sort_values("AUC", ascending=False)
                .drop_duplicates(["Dataset", "AE", "Sampler", "OCC"])
    )
    df_best = df_clean.merge(
        best_cfg[["Dataset", "AE", "Sampler", "OCC", "Config"]],
        on=["Dataset", "AE", "Sampler", "OCC", "Config"],
    )

    print("\n── Best Config 列表（節錄）──")
    for (ds_name, ae, sp, occ), grp in list(
            df_best.groupby(["Dataset", "AE", "Sampler", "OCC"]))[:20]:
        print(f"  [{ds_name}] {ae:4s} Sampler={sp:4s} × {occ:8s}  "
              f"best={grp['Config'].iloc[0]:12s}  5-fold avg AUC={grp['AUC'].mean():.4f}")

    return df_all, df_best


# ─────────────────────────── Excel 樣式（與 B 共用設計）──────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

AE_FILL = {
    "AE":  PatternFill("solid", fgColor="DAEEF3"),
    "DAE": PatternFill("solid", fgColor="E2EFDA"),
    "SAE": PatternFill("solid", fgColor="FFF2CC"),
    "VAE": PatternFill("solid", fgColor="FCE4D6"),
}

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="1F3864", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def sc(cell, value, font=None, fill=None, align=None, fmt=None):
    cell.value  = value
    cell.border = THIN_BORDER
    if font:  cell.font          = font
    if fill:  cell.fill          = fill
    if align: cell.alignment     = align
    if fmt:   cell.number_format = fmt


def col_w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ─────────────────────────── A~K 統整友善輸出 ────────────────────────────────
def make_ak_export_df(df, config_policy):
    """建立固定欄位的 A~K 統整用資料表。

    後續統整 A~K 時建議讀取 ak_all_export / ak_best_export，並用欄位名稱抓值；
    這樣 per_fold / summary 分頁新增 metadata 欄位時，不會造成欄位位置錯位。
    """
    if df.empty:
        return pd.DataFrame(columns=COMPARISON_EXPORT_COLS)

    out = pd.DataFrame({
        "Study": STUDY_ID,
        "Method": METHOD_ID,
        "FeatureSet": FEATURE_SET,
        "Dataset": df["Dataset"],
        "AE": df["AE"],
        "Sampler": df["Sampler"],
        "OCC": df["OCC"],
        "Config": df["Config"],
        "Fold": df["Fold"],
        "ConfigPolicy": config_policy,
        "MajKept": df["MajKept"],
        "MajRemoved": df["MajRemoved"],
        "RemovedRate": df["RemovedRate"],
        "SamplerStatus": df["SamplerStatus"],
        "BaselineRef": df["BaselineCheck"],
        "OCCScope": OCC_SCOPE,
        "SamplerScaleMode": SAMPLER_SCALE_MODE,
    })
    for metric in METRIC_COLS:
        out[metric] = df[metric]
    return out[COMPARISON_EXPORT_COLS]


def write_ak_export(ws, df, title, config_policy):
    ws.title = title
    out = make_ak_export_df(df, config_policy=config_policy)
    for c, h in enumerate(COMPARISON_EXPORT_COLS, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    for r, (_, row) in enumerate(out.iterrows(), 2):
        fill = AE_FILL.get(row["AE"])
        for c, col in enumerate(COMPARISON_EXPORT_COLS, 1):
            sc(ws.cell(r, c), row[col], font=BODY_FONT, fill=fill,
               align=LEFT_ALIGN if col in ["Dataset", "SamplerScaleMode"] else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS + ["RemovedRate"] else None)
    widths = [8, 14, 18, 28, 8, 10, 8, 14, 8, 24, 10, 12, 12, 24, 14, 12, 48] + [10] * len(METRIC_COLS)
    for i, w in enumerate(widths, 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


def write_alignment_notes(ws):
    """記錄此檔與 baseline 的對齊關係，方便論文與 A~K 統整追溯。"""
    ws.title = "alignment_notes"
    notes = [
        ("Study", STUDY_ID),
        ("Method", METHOD_ID),
        ("FeatureSet", FEATURE_SET),
        ("BaselineRef", BASELINE_REF),
        ("AE", ", ".join(AE_TYPES)),
        ("AE epochs", str(AE_EPOCHS)),
        ("AE batch size", str(AE_BATCH_SIZE)),
        ("AE learning rate", str(AE_LR)),
        ("DAE noise", str(DAE_NOISE)),
        ("SAE sparsity", str(SAE_SPARSITY)),
        ("VAE beta", str(VAE_BETA)),
        ("Grid", f"{len(N_LAYERS_LIST)} layers × {len(BOTTLENECK_RATIOS)} ratios = {len(ALL_CONFIGS)} configs"),
        ("OCCScope", OCC_SCOPE),
        ("OCC", ", ".join(OCC_TYPES)),
        ("Sampler", ", ".join(SAMPLERS)),
        ("SamplerScaleMode", SAMPLER_SCALE_MODE),
        ("None sanity check", "Sampler=none + AE×OCC×Config should match the corresponding Baseline B AE×OCC×Config under the same data/fold."),
        ("Threshold", "90th percentile of training majority anomaly scores."),
        ("LOF n_neighbors", "min(20, len(cleaned majority)-1)."),
        ("Contamination", "0.1"),
        ("Best config policy", "per-(Dataset, AE, Sampler, OCC), choose Config by 5-fold mean AUC."),
        ("Data leakage guard", "DF scaler is fit on uncleaned training majority only; sampler uses train fold only; test is transformed only."),
    ]
    sc(ws.cell(1, 1), "Item", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    sc(ws.cell(1, 2), "Value", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    for r, (k, v) in enumerate(notes, 2):
        sc(ws.cell(r, 1), k, font=BOLD_FONT, fill=ALT_FILL if r % 2 == 0 else None, align=LEFT_ALIGN)
        sc(ws.cell(r, 2), v, font=BODY_FONT, fill=ALT_FILL if r % 2 == 0 else None, align=LEFT_ALIGN)
    col_w(ws, "A", 24)
    col_w(ws, "B", 110)
    ws.freeze_panes = "A2"


# ─────────────────────────── per_fold 分頁 ───────────────────────────────────
def write_per_fold(ws, df, title):
    ws.title = title
    headers = ["Dataset", "AE", "Sampler", "OCC", "Config", "Fold",
               "MajKept", "MajRemoved", "RemovedRate",
               "SamplerStatus", "BaselineCheck"] + METRIC_COLS

    for c, h in enumerate(headers, 1):
        sc(ws.cell(1, c), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        fill = AE_FILL.get(row["AE"])
        for c, col in enumerate(headers, 1):
            sc(ws.cell(r, c), row[col],
               font=BODY_FONT, fill=fill,
               align=LEFT_ALIGN if c == 1 else CENTER_ALIGN,
               fmt="0.0000" if col in METRIC_COLS + ["RemovedRate"] else None)

    for i, w in enumerate([28, 6, 9, 8, 14, 6, 9, 11, 12, 24, 16] + [10] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A2"


# ─────────────────────────── all_summary 分頁 ────────────────────────────────
def write_summary_all(ws, df):
    """all_summary：Dataset | AE × Sampler × Config × Metric（OCC 單一，省略不列）"""
    ws.title = "all_summary"
    configs = ALL_CONFIGS

    grouped = (df.groupby(["Dataset", "AE", "Sampler", "Config"])[METRIC_COLS]
                 .agg(["mean", "std"]).reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for ae in AE_TYPES:
        ae_span = len(SAMPLERS) * len(configs) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + ae_span - 1)
        sc(ws.cell(1, col), ae,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=AE_FILL.get(ae), align=CENTER_ALIGN)
        for sp in SAMPLERS:
            sp_span = len(configs) * len(METRIC_COLS)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + sp_span - 1)
            sc(ws.cell(2, col), f"US={sp}", font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            for cfg in configs:
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + len(METRIC_COLS) - 1)
                sc(ws.cell(3, col), cfg, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                for metric in METRIC_COLS:
                    sc(ws.cell(4, col), metric, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                    col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 5):
        row_alt = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=row_alt, align=LEFT_ALIGN)
        col = 2
        for ae in AE_TYPES:
            for sp in SAMPLERS:
                for cfg in configs:
                    sub = grouped[
                        (grouped["Dataset"] == ds) & (grouped["AE"] == ae) &
                        (grouped["Sampler"] == sp) & (grouped["Config"] == cfg)
                    ]
                    for metric in METRIC_COLS:
                        try:
                            m = sub[(metric, "mean")].values[0]
                            s = sub[(metric, "std")].values[0]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, col), display, font=BODY_FONT,
                           fill=AE_FILL.get(ae) if row_alt is None else row_alt,
                           align=CENTER_ALIGN)
                        col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(AE_TYPES) * len(SAMPLERS) * len(configs) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B5"


# ─────────────────────────── best_summary 分頁 ───────────────────────────────
def write_summary_best(ws, df):
    """best_summary：Dataset | AE × Sampler × Metric（OCC 單一省略）"""
    ws.title = "best_summary"

    grouped = (df.groupby(["Dataset", "AE", "Sampler"])[METRIC_COLS]
                 .agg(["mean", "std"]).reset_index())

    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    sc(ws.cell(1, 1), "Dataset", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    col = 2
    for ae in AE_TYPES:
        ae_span = len(SAMPLERS) * len(METRIC_COLS)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + ae_span - 1)
        sc(ws.cell(1, col), ae,
           font=Font(name="Arial", bold=True, color="1F3864", size=11),
           fill=AE_FILL.get(ae), align=CENTER_ALIGN)
        for sp in SAMPLERS:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + len(METRIC_COLS) - 1)
            sc(ws.cell(2, col), f"US={sp}", font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
            for metric in METRIC_COLS:
                sc(ws.cell(3, col), metric, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENTER_ALIGN)
                col += 1

    datasets = df["Dataset"].unique()
    for r, ds in enumerate(datasets, 4):
        row_alt = ALT_FILL if r % 2 == 0 else None
        sc(ws.cell(r, 1), ds, font=BODY_FONT, fill=row_alt, align=LEFT_ALIGN)
        col = 2
        for ae in AE_TYPES:
            for sp in SAMPLERS:
                sub = grouped[
                    (grouped["Dataset"] == ds) & (grouped["AE"] == ae) &
                    (grouped["Sampler"] == sp)
                ]
                for metric in METRIC_COLS:
                    try:
                        m = sub[(metric, "mean")].values[0]
                        s = sub[(metric, "std")].values[0]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, col), display, font=BODY_FONT,
                       fill=AE_FILL.get(ae) if row_alt is None else row_alt,
                       align=CENTER_ALIGN)
                    col += 1

    col_w(ws, "A", 28)
    for i in range(2, 2 + len(AE_TYPES) * len(SAMPLERS) * len(METRIC_COLS)):
        col_w(ws, get_column_letter(i), 18)
    ws.freeze_panes = "B4"


# ─────────────────────────── all_overall 分頁 ────────────────────────────────
def write_overall_all(ws, df):
    """all_overall：AE × Sampler × OCC × Config 全域平均（flat）"""
    ws.title = "all_overall"

    total_cols = 6 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（all configs）- Study2 DF-side Under-sampling × 4 AE"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["AE", "Sampler", "OCC", "Config", "Avg Removed", "Avg Removed Rate"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["AE", "Sampler", "OCC", "Config"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for ae in AE_TYPES:
        for sp in SAMPLERS:
            for occ in OCC_TYPES:
                for cfg in ALL_CONFIGS:
                    fill = AE_FILL.get(ae)
                    sub = df[(df["AE"] == ae) & (df["Sampler"] == sp) &
                             (df["OCC"] == occ) & (df["Config"] == cfg)]
                    avg_rm = sub["MajRemoved"].mean() if not sub.empty else 0.0
                    avg_rate = sub["RemovedRate"].mean() if not sub.empty else 0.0
                    sc(ws.cell(r, 1), ae,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 2), sp,  font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 3), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 4), cfg, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 5), f"{avg_rm:.1f}", font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                    sc(ws.cell(r, 6), avg_rate, font=BODY_FONT, fill=fill, align=CENTER_ALIGN, fmt="0.0000")
                    for i, metric in enumerate(METRIC_COLS, 7):
                        try:
                            m = overall.loc[(ae, sp, occ, cfg), (metric, "mean")]
                            s = overall.loc[(ae, sp, occ, cfg), (metric, "std")]
                            display = f"{m:.4f} ± {s:.4f}"
                        except Exception:
                            display = "N/A"
                        sc(ws.cell(r, i), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                    r += 1

    for i, w in enumerate([6, 9, 8, 14, 12, 16] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"


# ─────────────────────────── best_overall 分頁 ───────────────────────────────
def write_overall_best(ws, df):
    """best_overall：AE × Sampler × OCC 全域平均 + 最常選中 Config + 平均刪除量（核心比較表）"""
    ws.title = "best_overall"

    total_cols = 6 + len(METRIC_COLS)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value     = "Overall Mean（best config per dataset）- Study2 DF-side Under-sampling × 4 AE"
    c.font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    c.alignment = CENTER_ALIGN

    headers = ["AE", "Sampler", "OCC", "Most Freq Config", "Avg Removed", "Avg Removed Rate"] + METRIC_COLS
    for i, h in enumerate(headers, 1):
        sc(ws.cell(2, i), h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)

    overall = df.groupby(["AE", "Sampler", "OCC"])[METRIC_COLS].agg(["mean", "std"])

    r = 3
    for ae in AE_TYPES:
        for sp in SAMPLERS:
            for occ in OCC_TYPES:
                fill = AE_FILL.get(ae)
                sub = df[(df["AE"] == ae) & (df["Sampler"] == sp) & (df["OCC"] == occ)]
                cfg = sub["Config"].mode().iloc[0] if not sub.empty else "N/A"
                avg_rm = sub["MajRemoved"].mean() if not sub.empty else 0.0
                avg_rate = sub["RemovedRate"].mean() if not sub.empty else 0.0
                sc(ws.cell(r, 1), ae,  font=BOLD_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 2), sp,  font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 3), occ, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 4), cfg, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 5), f"{avg_rm:.1f}", font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                sc(ws.cell(r, 6), avg_rate, font=BODY_FONT, fill=fill, align=CENTER_ALIGN, fmt="0.0000")
                for i, metric in enumerate(METRIC_COLS, 7):
                    try:
                        m = overall.loc[(ae, sp, occ), (metric, "mean")]
                        s = overall.loc[(ae, sp, occ), (metric, "std")]
                        display = f"{m:.4f} ± {s:.4f}"
                    except Exception:
                        display = "N/A"
                    sc(ws.cell(r, i), display, font=BODY_FONT, fill=fill, align=CENTER_ALIGN)
                r += 1

    for i, w in enumerate([6, 9, 8, 18, 12, 16] + [22] * len(METRIC_COLS), 1):
        col_w(ws, get_column_letter(i), w)
    ws.freeze_panes = "A3"



def write_empty_sheet(ws, title, message):
    """在沒有 best 結果時保留分頁，避免整個 Excel 輸出失敗。"""
    ws.title = title
    sc(ws.cell(1, 1), message, font=BOLD_FONT, fill=SUBHDR_FILL, align=LEFT_ALIGN)
    col_w(ws, "A", 80)

# ─────────────────────────── Excel 存檔 ──────────────────────────────────────
def save_excel(df_all, df_best):
    wb  = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()
    ws6 = wb.create_sheet()
    ws7 = wb.create_sheet()
    ws8 = wb.create_sheet()
    ws9 = wb.create_sheet()

    write_per_fold(    ws1, df_all,  "all_per_fold")
    write_summary_all( ws2, df_all)
    write_overall_all( ws3, df_all)

    if df_best.empty:
        write_empty_sheet(ws4, "best_per_fold",  "No valid best config rows. Check AUC results or skipped folds.")
        write_empty_sheet(ws5, "best_summary",   "No valid best config rows. Check AUC results or skipped folds.")
        write_empty_sheet(ws6, "best_overall",   "No valid best config rows. Check AUC results or skipped folds.")
        write_empty_sheet(ws8, "ak_best_export", "No valid best config rows. Use ak_all_export or check skipped folds.")
    else:
        write_per_fold(    ws4, df_best, "best_per_fold")
        write_summary_best(ws5, df_best)
        write_overall_best(ws6, df_best)
        write_ak_export(ws8, df_best, title="ak_best_export", config_policy="best_config_per_dataset")

    write_ak_export(ws7, df_all, title="ak_all_export", config_policy="all_configs")
    write_alignment_notes(ws9)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 結果已儲存至：{OUTPUT_FILE.resolve()}")


# ─────────────────────────── Entry Point ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 68)
    print("Study Two（DF_maj feature set）：DF 側 Under-sampling × 4 AE × LOF")
    print(f"AE        : {AE_TYPES}（4 個全跑）")
    print(f"Sampler   : {SAMPLERS}（含 none 內部 baseline）")
    print(f"OCC       : {OCC_TYPES}")
    print(f"搜尋空間  : n_layers={N_LAYERS_LIST} × bottleneck={list(BOTTLENECK_RATIOS.keys())}")
    print(f"總組合數  : {len(AE_TYPES)} AE × {len(SAMPLERS)} US × {len(OCC_TYPES)} OCC × "
          f"{len(ALL_CONFIGS)} cfg = "
          f"{len(AE_TYPES)*len(SAMPLERS)*len(OCC_TYPES)*len(ALL_CONFIGS)} per fold")
    print("Pipeline  : OF → MinMax → AE → DF → MinMax(fit uncleaned DF_maj) → US")
    print("            → LOF(no second MinMax)      （OCC 在最後）")
    print("=" * 68)

    df_all, df_best = run_experiment()

    if df_all.empty:
        print("\n⚠️  沒有任何結果，請確認資料路徑與檔名格式。")
    else:
        save_excel(df_all, df_best)
        if df_best.empty:
            print("\n⚠️  已產生 all results，但沒有可用的 best config 結果。")
        else:
            print("\n── Best Config Overall Mean（by AE × Sampler）──")
            print(df_best.groupby(["AE", "Sampler"])[METRIC_COLS].mean().round(4).to_string())
